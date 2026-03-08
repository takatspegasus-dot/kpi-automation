# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Tuple, Optional, Any, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border


# =========================
# フォルダ
# =========================
BASE_DIR = Path(r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート")
INTEGRATED_DIR = BASE_DIR / "03_統合_課別回収ファイル"
ANALYSIS_DIR = BASE_DIR / "04_分析"
ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)

# =========================
# 入力（統合ブック）
# =========================
INPUT_FILE = INTEGRATED_DIR / "KPI振り返り_課別統合_20260205.xlsx"

# =========================
# テンプレ（固定）
# =========================
TEMPLATE_FILE = ANALYSIS_DIR / "KPI振り返り_分析RAW01.xlsx"

# =========================
# 出力（日付入り）
# =========================
OUT_FILE = ANALYSIS_DIR / f"KPI振り返り_分析RAW_{datetime.now():%Y%m%d}.xlsx"


# =========================
# 成果判定しきい値
# =========================
THRESH_OK = 1.00
THRESH_PARTIAL = 0.80


# =========================
# 戦略意図マスタ（5分類）
# =========================
STRATEGY_MASTER = [
    ("土俵転換成功率", "BAC以外の判断軸で商談が進んだか"),
    ("判断持ち帰り率", "即決ではなく、納得した検討で終われたか"),
    ("仮説ヒアリング実施率", "SPIN話法による判断軸探索ができたか"),
    ("BAC回避成功率", "敵の土俵に乗らなかったか"),
    ("設定マスク型ロープレ実施率", "思考型営業の再現性"),
]
LABELS = {x[0] for x in STRATEGY_MASTER}


# =========================
# 統合ブック側ヘッダ探索キーワード
# =========================
KW = {
    "kpi": ["KPI"],
    "kpi_category": ["KPI区分"],
    "period": ["評価対象期間", "対象期間", "評価期間"],
    "target": ["設定目標", "KPI目標", "目標値", "目標"],
    "actual": ["最終実績", "実績"],
    # 行動：シート側の表記揺れ対応
    "action": ["KPI達成のために行った指示・行動", "指示・行動", "行動"],
    # 戦略意図：表記揺れ対応（読点/括弧）
    "strategy_raw": [
        "今回の行動は、どの戦略意図に基づくものか",
        "今回の行動はどの戦略意図に基づくものか",
        "戦略意図（選択）",
        "戦略意図",
    ],
    "kpi_out": ["KPI外で実績を生んだ行動", "KPI外"],
    "next_kpi": ["スタートに戻れるなら設定するKPI", "戻れるなら", "設定するKPI"],
    "comment": ["妥当性の振り返り", "妥当性", "コメント"],
}


# =========================
# ユーティリティ
# =========================
def norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("\r", " ").strip()

def to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = norm(v).replace(",", "").replace("％", "%")
    if not s:
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100.0
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None

def calc_rate01(target: Any, actual: Any) -> Optional[float]:
    t = to_float(target)
    a = to_float(actual)
    if t is None or a is None or t == 0:
        return None
    return a / t

def judge_outcome_symbol(rate01: Optional[float]) -> str:
    if rate01 is None:
        return "—"
    if rate01 >= THRESH_OK:
        return "◎"
    if rate01 >= THRESH_PARTIAL:
        return "○"
    return "△"

def find_header(ws: Worksheet, search_rows: int = 60, search_cols: int = 200) -> Tuple[int, Dict[str, int]]:
    """
    ヘッダ行と各項目の列番号（1-based）を返す
    """
    best_row, best_cols, best_score = 0, {}, -1
    for r in range(1, min(ws.max_row, search_rows) + 1):
        cols: Dict[str, int] = {}
        for c in range(1, min(ws.max_column, search_cols) + 1):
            v = norm(ws.cell(r, c).value)
            if not v:
                continue
            for key, kws in KW.items():
                if key in cols:
                    continue
                if any(k in v for k in kws):
                    cols[key] = c

        score = len(cols)
        if score > best_score:
            best_row, best_cols, best_score = r, cols, score

        # 必須最低限
        if all(k in cols for k in ("kpi", "target", "actual", "strategy_raw")):
            return r, cols

    return best_row, best_cols

def detect_colmap_by_header(row1_ws: Worksheet) -> Dict[str, int]:
    """
    テンプレの00_分析_RAW：1行目ヘッダ名 -> 列番号
    """
    m = {}
    for c in range(1, row1_ws.max_column + 1):
        h = norm(row1_ws.cell(1, c).value)
        if h:
            m[h] = c
    return m

def clear_values(ws: Worksheet, start_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).value = None

def last_row_with_any_value(ws: Worksheet, cols: List[int]) -> int:
    last = 1
    for r in range(2, ws.max_row + 1):
        if any(norm(ws.cell(r, c).value) != "" for c in cols):
            last = r
    return last

def remove_borders_for_empty_cells_after(ws: Worksheet, start_row: int, cols: List[int]) -> None:
    """
    start_row 以降、空セルだけ罫線を消す（列指定）
    """
    no_border = Border()
    for r in range(start_row, ws.max_row + 1):
        for c in cols:
            if norm(ws.cell(r, c).value) == "":
                ws.cell(r, c).border = no_border

def map_strategy(strategy_raw: Any) -> Tuple[str, str]:
    """
    戦略意図のラベル化：事故防止のため、安易に土俵転換へ落とさない
    """
    raw = norm(strategy_raw)

    if not raw:
        return "未判定", ""

    # すでにラベル
    if raw in LABELS:
        return raw, raw

    # 「ラベル：説明」
    for sep in ("：", ":"):
        if sep in raw:
            left = raw.split(sep, 1)[0].strip()
            if left in LABELS:
                return left, raw

    # 文中にラベルが含まれる
    for label in LABELS:
        if label in raw:
            return label, raw

    # ルール分類（保険）
    if "BAC" in raw:
        return "BAC回避成功率", raw
    if "SPIN" in raw or "ヒアリング" in raw or "仮説" in raw:
        return "仮説ヒアリング実施率", raw
    if "ロープレ" in raw or "マスク" in raw:
        return "設定マスク型ロープレ実施率", raw
    if "持ち帰り" in raw or "検討" in raw:
        return "判断持ち帰り率", raw
    if "土俵" in raw or "判断軸" in raw:
        return "土俵転換成功率", raw

    return "未判定", raw


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {INPUT_FILE}")
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"テンプレが見つかりません: {TEMPLATE_FILE}")

    # テンプレをロード（見た目はテンプレ維持）
    out_wb = load_workbook(TEMPLATE_FILE)
    raw_ws = out_wb["00_分析_RAW"]
    ms_ws = out_wb["99_戦略意図マスタ"]

    # 入力：式版/値版
    srcF = load_workbook(INPUT_FILE, data_only=False)
    srcV = load_workbook(INPUT_FILE, data_only=True)

    # 99_戦略意図マスタ更新（値だけ）
    clear_values(ms_ws, 2, 1, 2)
    rr = 2
    for label, desc in STRATEGY_MASTER:
        ms_ws.cell(rr, 1).value = label
        ms_ws.cell(rr, 2).value = desc
        rr += 1

    # 00_分析_RAW：値をクリア（スタイルは残す）
    clear_values(raw_ws, 2, 1, raw_ws.max_column)

    # テンプレ列マップ（このヘッダに合わせて書き込む）
    colmap = detect_colmap_by_header(raw_ws)

    # テンプレ必須ヘッダ（添付テンプレの構造）
    must = [
        "課", "KPI区分", "評価対象期間", "KPI", "設定目標", "最終実績", "KPI達成率",
        "成果が出たか（自動）", "戦略意図（選択）", "行動（結合）",
        "KPI外だが効いた行動", "次フェーズKPI候補", "コメント原文", "備考"
    ]
    missing = [h for h in must if h not in colmap]
    if missing:
        raise RuntimeError(f"テンプレの00_分析_RAW でヘッダ不足: {missing}")

    # 対象シート：01_～14_ のような課別
    target_sheets = [s for s in srcF.sheetnames if re.match(r"^\d{2}_", s)]

    out_r = 2
    for sh in target_sheets:
        wsF = srcF[sh]
        wsV = srcV[sh]

        header_row, cols = find_header(wsF)
        if header_row == 0:
            raise RuntimeError(f"{sh}：ヘッダ行を検出できません")
        if "strategy_raw" not in cols:
            raise RuntimeError(f"{sh}：戦略意図列を検出できません")

        blank_streak = 0
        saw_any_strategy = False

        for r in range(header_row + 1, wsF.max_row + 1):
            kpi = norm(wsF.cell(r, cols["kpi"]).value)
            if not kpi:
                blank_streak += 1
                if blank_streak >= 10:
                    break
                continue
            blank_streak = 0

            # KPI区分/期間（無ければ空）
            kpi_cat = norm(wsF.cell(r, cols.get("kpi_category", 0)).value) if cols.get("kpi_category") else ""
            period = norm(wsF.cell(r, cols.get("period", 0)).value) if cols.get("period") else ""

            # 目標/実績：値版優先→無ければ式版
            tgt = wsV.cell(r, cols["target"]).value
            if tgt is None:
                tgt = wsF.cell(r, cols["target"]).value

            act = wsV.cell(r, cols["actual"]).value
            if act is None:
                act = wsF.cell(r, cols["actual"]).value

            rate01 = calc_rate01(tgt, act)
            outcome = judge_outcome_symbol(rate01)

            # 戦略意図
            strat_cell = wsF.cell(r, cols["strategy_raw"]).value
            if norm(strat_cell) != "":
                saw_any_strategy = True
            strat_label, strat_raw = map_strategy(strat_cell)

            # 行動（結合）：取れるなら action列
            action = norm(wsF.cell(r, cols.get("action", 0)).value) if cols.get("action") else ""

            # KPI外 / 次KPI / コメント
            kpi_out = norm(wsF.cell(r, cols.get("kpi_out", 0)).value) if cols.get("kpi_out") else ""
            next_kpi = norm(wsF.cell(r, cols.get("next_kpi", 0)).value) if cols.get("next_kpi") else ""
            comment = norm(wsF.cell(r, cols.get("comment", 0)).value) if cols.get("comment") else ""

            # ----- 書き込み（テンプレ列へ）
            raw_ws.cell(out_r, colmap["課"]).value = sh
            raw_ws.cell(out_r, colmap["KPI区分"]).value = kpi_cat
            raw_ws.cell(out_r, colmap["評価対象期間"]).value = period
            raw_ws.cell(out_r, colmap["KPI"]).value = kpi
            raw_ws.cell(out_r, colmap["設定目標"]).value = tgt
            raw_ws.cell(out_r, colmap["最終実績"]).value = act

            # KPI達成率
            c_rate = colmap["KPI達成率"]
            if rate01 is None:
                raw_ws.cell(out_r, c_rate).value = None
            else:
                raw_ws.cell(out_r, c_rate).value = rate01
                raw_ws.cell(out_r, c_rate).number_format = "0%"

            raw_ws.cell(out_r, colmap["成果が出たか（自動）"]).value = outcome
            raw_ws.cell(out_r, colmap["戦略意図（選択）"]).value = strat_label
            raw_ws.cell(out_r, colmap["行動（結合）"]).value = action
            raw_ws.cell(out_r, colmap["KPI外だが効いた行動"]).value = kpi_out
            raw_ws.cell(out_r, colmap["次フェーズKPI候補"]).value = next_kpi
            raw_ws.cell(out_r, colmap["コメント原文"]).value = comment

            # 備考：戦略意図の原文は必ず残す（監査用）
            note_parts = []
            if strat_raw:
                note_parts.append(f"戦略意図原文={strat_raw}")
            if strat_label == "未判定" and strat_raw:
                note_parts.append("※戦略意図がラベル化できていません（要確認）")
            raw_ws.cell(out_r, colmap["備考"]).value = " / ".join(note_parts)

            out_r += 1

        if not saw_any_strategy:
            # 事故防止：戦略意図列が空なら止める（全件土俵転換になりがち）
            raise RuntimeError(f"{sh}：戦略意図列が空です（列違い/結合セルの可能性）")

    # -------------------------
    # 罫線消し：最終データ行以降の「空セル」は全列で罫線OFF
    # -------------------------
    # データの最終行は「課」「KPI」などで判定
    major_cols = [
        colmap["課"],
        colmap["KPI"],
        colmap["設定目標"],
        colmap["最終実績"],
        colmap["KPI達成率"],
    ]
    last_data_row = last_row_with_any_value(raw_ws, major_cols)

    # 00_分析_RAWの全列（A～N）を対象に「空セルの罫線」を消す
    all_cols = list(range(1, len(must) + 1))
    remove_borders_for_empty_cells_after(raw_ws, start_row=last_data_row + 1, cols=all_cols)

    # 保存
    out_wb.save(OUT_FILE)
    print(f"✅ 出力完了: {OUT_FILE}")


if __name__ == "__main__":
    main()

