# -*- coding: utf-8 -*-
"""
KPI振り返り_①事実②構造③判断_母艦
【完全版・安定動作（短い保存名：YYYYMMDD + 連番）】

前提：
- 元テンプレは M=① / N=② / O=③ の構造

処理内容：
1. ②を2列化（N=構造タイプ / O=理由補足）
2. ③をP列へ移動
3. 判断ログ（Q,R,S）を追加
4. タイトル（4〜6行）縦結合＋装飾（薄緑・中央・Meiryo UI 11・太字なし）
5. 罫線コピー＋外枠＋4行目実線補正
6. プルダウン設定（矢印表示）
7. 安全保存（短いファイル名：_YYYYMMDD.xlsx、2回目以降 _v2 等）
"""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ===== 入出力 =====
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\研修テンプレート\【研修テンプレ原本】KPI振り返り.xlsx"
OUTPUT_DIR = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート"

OUTPUT_BASENAME = "KPI振り返り_①事実②構造③判断_母艦"
NEW_SHEET_NAME = "00_全課_①②③検討"

# ===== 行定義 =====
TITLE_TOP_ROW = 4
TITLE_BOTTOM_ROW = 6
DATA_START_ROW = 7

# ===== 列定義（最終形：M〜S）=====
COL_M = 13  # ①事実
COL_N = 14  # ②-1 構造タイプ（プルダウン）
COL_O = 15  # ②-2 理由・補足（自由記述）
COL_P = 16  # ③論点（旧③）
COL_Q = 17  # 判断区分（プルダウン）
COL_R = 18  # 判断主体（プルダウン）
COL_S = 19  # 次フェーズ反映（プルダウン）
BLOCK_COLS = (COL_M, COL_N, COL_O, COL_P, COL_Q, COL_R, COL_S)

SRC_BORDER_COL = 11  # K列（罫線コピー元）

COL_WIDTH = 34
ZOOM_SCALE = 80

HEADERS = {
    COL_M: "①何が起きているか（事実）",
    COL_N: "②-1 構造タイプ（ラベル）",
    COL_O: "②-2 構造の理由・補足",
    COL_P: "③次に何を考えるべきか（具体論点）",
    COL_Q: "判断区分",
    COL_R: "判断主体",
    COL_S: "次フェーズ反映",
}

# ===== プルダウン候補（カンマ禁止）=====
STRUCTURE_TYPES = [
    "仕込み成功・転換未設計",
    "プロセス完成・決定理由不足",
    "KPI未達だが成果発生",
    "活動主義",
    "選択と集中型",
    "因果成立（成功型）",
    "戦略仮説ミスマッチ",
    "その他",
]
DECISIONS = ["継続", "変更", "削除", "新設"]
OWNERS = ["課", "営業推進部", "全社"]
REFLECTS = ["○", "△", "×"]


# ---------- 保存（短いファイル名：YYYYMMDD + 連番） ----------
def save_unique_short(wb) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    d = datetime.now().strftime("%Y%m%d")
    base = os.path.join(OUTPUT_DIR, f"{OUTPUT_BASENAME}_{d}.xlsx")

    # 1本目
    if not os.path.exists(base):
        wb.save(base)
        return base

    # 2本目以降：_v2, _v3 ...
    name, ext = os.path.splitext(base)
    for i in range(2, 100):
        cand = f"{name}_v{i}{ext}"
        try:
            wb.save(cand)
            return cand
        except PermissionError:
            continue

    raise PermissionError("保存可能なファイル名が見つかりませんでした。")


# ---------- 既存結合解除（タイトル領域） ----------
def unmerge_title_area(ws) -> None:
    targets = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= TITLE_BOTTOM_ROW and rng.max_row >= TITLE_TOP_ROW:
            if rng.min_col <= COL_S and rng.max_col >= COL_M:
                targets.append(str(rng))
    for addr in targets:
        ws.unmerge_cells(addr)


# ---------- 列構造変更（②2列化・③右へ・判断3列追加） ----------
def expand_columns(ws) -> None:
    """
    前提：元テンプレは M=① / N=② / O=③
    目的：N=構造タイプ、O=理由補足(旧②)、P=論点(旧③)、QRS追加
    """
    # O列に1列挿入 → 旧③(O)がPへ自動シフト
    ws.insert_cols(COL_O, amount=1)

    # 旧②(N) を Oへ移動、Nは空に（構造タイプ用）
    for r in range(DATA_START_ROW, ws.max_row + 1):
        ws.cell(row=r, column=COL_O).value = ws.cell(row=r, column=COL_N).value
        ws.cell(row=r, column=COL_N).value = None

    # 判断ログ（Q,R,S）を追加：Q位置（17列目）から3列挿入
    ws.insert_cols(COL_Q, amount=3)


# ---------- 書式（タイトル） ----------
def format_titles(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAD3")
    font = Font(name="Meiryo UI", size=11, bold=False)
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin = Side(style="thin", color="000000")

    ws.sheet_view.zoomScale = ZOOM_SCALE
    ws.row_dimensions[TITLE_TOP_ROW].height = 30
    ws.row_dimensions[TITLE_TOP_ROW + 1].height = 30
    ws.row_dimensions[TITLE_BOTTOM_ROW].height = 30

    for c in BLOCK_COLS:
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH

        ws.merge_cells(
            start_row=TITLE_TOP_ROW, start_column=c,
            end_row=TITLE_BOTTOM_ROW, end_column=c
        )

        ws.cell(row=TITLE_TOP_ROW, column=c).value = HEADERS[c]

        for r in range(TITLE_TOP_ROW, TITLE_BOTTOM_ROW + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------- 書式（データ行） ----------
def format_data(ws) -> None:
    font = Font(name="Meiryo UI", size=11, bold=False)
    align = Alignment(wrap_text=True, vertical="top")

    for r in range(DATA_START_ROW, ws.max_row + 1):
        for c in BLOCK_COLS:
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.alignment = align


# ---------- 罫線（K列→M〜Sコピー） ----------
def copy_borders_from_k(ws) -> None:
    for r in range(TITLE_TOP_ROW, ws.max_row + 1):
        src = ws.cell(row=r, column=SRC_BORDER_COL).border
        for c in BLOCK_COLS:
            ws.cell(row=r, column=c).border = Border(
                left=src.left, right=src.right,
                top=src.top, bottom=src.bottom,
                diagonal=src.diagonal,
                diagonal_direction=src.diagonal_direction,
                outline=src.outline,
                vertical=src.vertical,
                horizontal=src.horizontal,
            )


# ---------- 4行目を必ず実線（細） ----------
def force_row4_solid(ws) -> None:
    thin = Side(style="thin", color="000000")
    r = TITLE_TOP_ROW
    for c in BLOCK_COLS:
        ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------- 外枠（M〜S）を実線（細） ----------
def apply_outer_border(ws) -> None:
    thin = Side(style="thin", color="000000")
    left_col, right_col = COL_M, COL_S

    for r in range(TITLE_TOP_ROW, ws.max_row + 1):
        for c in BLOCK_COLS:
            b = ws.cell(row=r, column=c).border
            ws.cell(row=r, column=c).border = Border(
                left=thin if c == left_col else b.left,
                right=thin if c == right_col else b.right,
                top=thin if r == TITLE_TOP_ROW else b.top,
                bottom=thin if r == ws.max_row else b.bottom,
            )


# ---------- プルダウン（矢印表示） ----------
def add_dropdown(ws, col: int, options: list[str]) -> None:
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.showDropDown = False  # ★矢印を表示させる（Trueだと隠れる）
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col)
    dv.add(f"{col_letter}{DATA_START_ROW}:{col_letter}{ws.max_row}")


# ---------- main ----------
def main() -> None:
    if not os.path.exists(INPUT_XLSX):
        raise FileNotFoundError(INPUT_XLSX)

    wb = load_workbook(INPUT_XLSX)
    base = wb.active

    # 同名シートは削除して作り直す（混乱防止）
    if NEW_SHEET_NAME in wb.sheetnames:
        del wb[NEW_SHEET_NAME]

    ws = wb.copy_worksheet(base)
    ws.title = NEW_SHEET_NAME

    # 既存の結合があると merge が崩れるので解除
    unmerge_title_area(ws)

    # 列の構造変更（本体）
    expand_columns(ws)

    # 書式
    format_titles(ws)
    format_data(ws)

    # 罫線・外枠
    copy_borders_from_k(ws)
    force_row4_solid(ws)
    apply_outer_border(ws)

    # プルダウン
    add_dropdown(ws, COL_N, STRUCTURE_TYPES)
    add_dropdown(ws, COL_Q, DECISIONS)
    add_dropdown(ws, COL_R, OWNERS)
    add_dropdown(ws, COL_S, REFLECTS)

    # 保存（短い名前）
    saved = save_unique_short(wb)
    print("保存完了:", saved)


if __name__ == "__main__":
    main()
