# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
import math
from typing import Any, Dict

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# =========================
# 入出力（WS19を母艦）
# =========================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_WSシート19.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_WSシート19_V列_次レビュー指標.xlsx"

SHEET_NAME = "00_WS_サンプル回答"
ROW_START = 7
ROW_END = 300

# 元データ参照列（これまでの流れ踏襲）
COL_E = "E"   # KPI
COL_O = "O"   # 自課でコントロール可能な要因（レバー）
COL_S = "S"   # 次フェーズで強化するKPI軸
COL_T = "T"   # 次フェーズKPI（仮）※見出し＋説明
COL_V = "V"   # 次レビューで見る指標（出力）

# 行高は「行内最大列」に合わせる（列幅は手動設定をそのまま使う）
WRAP_COLS = list("LMNOPQRSTUV")

BASE_LINE_HEIGHT = 15.0
ROW_PADDING = 6.0
MIN_ROW_HEIGHT = 18.0


# =========================
# util
# =========================
def col_idx(col: str) -> int:
    n = 0
    for c in col:
        n = n * 26 + ord(c) - 64
    return n


def to_text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def shorten(s: str, max_len: int = 44) -> str:
    t = normalize(s)
    if not t:
        return ""
    if "。" in t:
        head = t.split("。")[0].strip()
        if len(head) >= 10:
            t = head
    return t if len(t) <= max_len else t[:max_len] + "…"


def strip_long_parens(s: str) -> str:
    t = (s or "").strip()
    if len(t) > 34:
        t = re.sub(r"（[^）]{18,}）", "", t)
    return t.strip()


def t_head(t_text: str) -> str:
    t = (t_text or "").strip()
    if not t:
        return ""
    line1 = t.splitlines()[0].replace("【KPI（仮）】", "").strip()
    return line1 if len(line1) <= 70 else line1[:70] + "…"


# =========================
# O列（レバー）解析
# =========================
def parse_levers(o_text: str) -> Dict[str, str]:
    levers = {"行動": "", "対象": "", "タイミング": "", "伝え方": ""}
    for line in (o_text or "").splitlines():
        m = re.match(r"^・(行動|対象|タイミング|伝え方)：(.+)$", line.strip())
        if m:
            levers[m.group(1)] = m.group(2).strip()
    return levers


# =========================
# S列（強化軸）判定
# =========================
def axis_from_s(s_text: str) -> str:
    t = s_text or ""
    if "新たなKPI" in t or "新KPI" in t:
        return "新KPI"
    if "タイミング" in t:
        return "タイミング"
    if "質" in t or "提案" in t or "伝え方" in t:
        return "質"
    if "活動量" in t:
        return "活動量"
    if "対象" in t:
        return "対象"
    return "対象"


def pick_focus(axis: str, levers: Dict[str, str]) -> str:
    if axis == "対象":
        return levers.get("対象") or levers.get("行動") or ""
    if axis == "活動量":
        return levers.get("行動") or ""
    if axis == "タイミング":
        return levers.get("タイミング") or levers.get("行動") or ""
    if axis == "質":
        return levers.get("伝え方") or levers.get("行動") or ""
    return levers.get("行動") or levers.get("対象") or ""


# =========================
# V列生成（自然日本語：あとで読んでわかる）
# =========================
def build_v(kpi_name: str, axis: str, levers: Dict[str, str], t_text: str) -> str:
    unit = "回" if axis == "活動量" else "件"

    kpi_line = t_head(t_text)
    if not kpi_line:
        kpi_line = f"{strip_long_parens(kpi_name)}：○{unit}"

    focus = shorten(pick_focus(axis, levers))
    if not focus:
        focus = "（設計ポイント未記入）"

    result = f"見る①：{kpi_line}（達成/未達と差分は次回ここに記入）"

    if axis == "対象":
        struct = f"見る②：対象の当て方が改善したか（例：{focus}）"
        process = "見る③：重点先への接点比率／重点先での反応・許可の出方"
    elif axis == "活動量":
        struct = f"見る②：実行量が確保できたか（例：{focus}）"
        process = "見る③：週あたり実施回数／継続できた週数（抜けがないか）"
    elif axis == "タイミング":
        struct = f"見る②：初動とフォロー間隔が整ったか（例：{focus}）"
        process = "見る③：初動実施率／フォロー遅延の有無（間隔が空いていないか）"
    elif axis == "質":
        struct = f"見る②：判断材料の出し方が整ったか（例：{focus}）"
        process = "見る③：判断材料提示率／次アクション合意率（Yesの取り方）"
    else:
        struct = f"見る②：前段の行動KPIに置き換えられたか（例：{focus}）"
        process = "見る③：前段行動の実施率／前段→成果への遷移（つながったか）"

    judge = "見る④：達成だけでなく、②③が改善しているかで「続ける/変える」を判断する。"

    return "\n".join([result, struct, process, judge])


# =========================
# wrap & 行高（行内最大列に合わせる）
# =========================
def apply_wrap(cell):
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal,
        vertical="top",
        wrap_text=True
    )


def get_col_width(ws, col_letter: str) -> float:
    w = ws.column_dimensions[col_letter].width
    return float(w) if w is not None else 12.0


def estimate_lines_for_cell(text: str, col_width: float) -> int:
    if not text:
        return 1

    # 列幅→1行あたり文字数（概算・欠け防止で少し控えめ）
    chars_per_line = max(10, int(col_width * 1.05))

    lines = 0
    for seg in text.split("\n"):
        s = seg.strip()
        if not s:
            lines += 1
            continue
        lines += max(1, math.ceil(len(s) / chars_per_line))

    return max(1, lines)


def adjust_row_height_by_max(ws, r: int):
    max_lines = 1
    for col in WRAP_COLS:
        cell = ws.cell(r, col_idx(col))
        apply_wrap(cell)
        if isinstance(cell.value, str) and cell.value.strip():
            col_width = get_col_width(ws, col)
            needed = estimate_lines_for_cell(cell.value, col_width)
            max_lines = max(max_lines, needed)

    h = BASE_LINE_HEIGHT * max_lines + ROW_PADDING
    ws.row_dimensions[r].height = max(h, MIN_ROW_HEIGHT)


# =========================
# main
# =========================
def main():
    if not os.path.exists(INPUT_XLSX):
        raise FileNotFoundError(f"入力ファイルが見つかりません: {INPUT_XLSX}")

    wb = load_workbook(INPUT_XLSX)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SHEET_NAME}")

    ws = wb[SHEET_NAME]

    count = 0
    for r in range(ROW_START, ROW_END + 1):
        kpi_name = to_text(ws[f"{COL_E}{r}"].value)
        if is_blank(kpi_name):
            continue

        axis = axis_from_s(to_text(ws[f"{COL_S}{r}"].value))
        levers = parse_levers(to_text(ws[f"{COL_O}{r}"].value))
        t_text = to_text(ws[f"{COL_T}{r}"].value)

        ws[f"{COL_V}{r}"].value = build_v(kpi_name, axis, levers, t_text)
        adjust_row_height_by_max(ws, r)
        count += 1

    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)
    wb.save(OUTPUT_XLSX)

    print("完了：WS19を母艦としてV列（次レビューで見る指標）を生成し、行高を行内最大列に合わせました。")
    print(f"処理行数：{count}")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
