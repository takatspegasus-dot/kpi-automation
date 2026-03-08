# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from typing import Optional, Tuple, List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ===== 入出力 =====
SRC_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_①事実②構造③判断_母艦_タイトル内ガイド列幅調整_01.xlsx"
OUT_DIR = os.path.dirname(SRC_XLSX)

OUT_NAME = "KPI振り返り_①事実②構造③判断_母艦_HK数式復元_プルダウン_サンプル選択済.xlsx"
OUT_XLSX = os.path.join(OUT_DIR, OUT_NAME)

# ===== 列（A=1 …）=====
COL_TARGET_DEN = 6   # F: 目標軒数
COL_ACTUAL     = 7   # G: 最終実績
COL_RATE       = 8   # H: KPI達成率（G/F）

COL_NYUNYU     = 10  # J: 納入本数
COL_SHARE      = 11  # K: （本数）実績全体に占める割合

COL_KPI        = 4   # D: KPI（データ行判定）
COL_KA         = 2   # B: 課（必要なら拡張用）

# ===== プルダウン対象 =====
SHEET_TEMPLATE = "00_全課_①②③検討"
SHEET_SAMPLE   = "00_全課_①②③検討_サンプル"

COL_STRUCT_TYPE = 14  # N列：②-1 構造タイプ（ラベル）
COL_JUDGE       = 17  # Q列：判断区分
COL_OWNER       = 18  # R列：判断主体
COL_REFLECT     = 19  # S列：次フェーズ反映

LIST_SHEET_NAME = "_dv_lists"  # 隠しシート名


# ===== ヘッダー行検出 =====
def find_header_row(ws) -> Optional[int]:
    """
    「課」「KPI区分」「評価対象期間」を含む行をヘッダー行として検出
    """
    required = {"課", "KPI区分", "評価対象期間"}
    for r in range(1, min(ws.max_row, 160) + 1):
        values = []
        for c in range(1, min(ws.max_column, 80) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                values.append(v.replace("\n", "").replace(" ", "").strip())
        if required.issubset(set(values)):
            return r
    return None


def find_last_data_row(ws, start_row: int) -> int:
    """
    KPI列（D列）が入っている最終行をデータ最終行とする
    """
    last = start_row
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=COL_KPI).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


# ===== H/K 数式復元 =====
def restore_formulas(ws) -> Tuple[int, int]:
    """
    wsのH列/K列を数式で復元。
    返り値：(処理行数, スキップ行数)
    """
    header_row = find_header_row(ws)
    if header_row is None:
        return (0, 0)

    data_start = header_row + 1
    last_row = find_last_data_row(ws, data_start)
    if last_row < data_start:
        return (0, 0)

    processed = 0
    skipped = 0

    j_col_letter = get_column_letter(COL_NYUNYU)
    j_sum_range = f"${j_col_letter}${data_start}:${j_col_letter}${last_row}"

    for r in range(data_start, last_row + 1):
        kpi_val = ws.cell(row=r, column=COL_KPI).value
        if kpi_val is None or str(kpi_val).strip() == "":
            skipped += 1
            continue

        # H列：達成率
        f_addr = f"{get_column_letter(COL_TARGET_DEN)}{r}"
        g_addr = f"{get_column_letter(COL_ACTUAL)}{r}"
        ws.cell(row=r, column=COL_RATE).value = f'=IFERROR({g_addr}/{f_addr},"")'
        if (not ws.cell(row=r, column=COL_RATE).number_format) or ws.cell(row=r, column=COL_RATE).number_format == "General":
            ws.cell(row=r, column=COL_RATE).number_format = "0.0%"

        # K列：構成比
        j_addr = f"{get_column_letter(COL_NYUNYU)}{r}"
        ws.cell(row=r, column=COL_SHARE).value = f'=IFERROR({j_addr}/SUM({j_sum_range}),"")'
        if (not ws.cell(row=r, column=COL_SHARE).number_format) or ws.cell(row=r, column=COL_SHARE).number_format == "General":
            ws.cell(row=r, column=COL_SHARE).number_format = "0.0%"

        processed += 1

    return (processed, skipped)


# ===== プルダウン（隠しシート方式）=====
def ensure_list_sheet(wb):
    if LIST_SHEET_NAME in wb.sheetnames:
        ws = wb[LIST_SHEET_NAME]
    else:
        ws = wb.create_sheet(LIST_SHEET_NAME)
    ws.sheet_state = "hidden"
    return ws


def write_list(ws_list, col: int, values: List[str]) -> str:
    # 既存クリア
    for r in range(1, 600):
        ws_list.cell(row=r, column=col).value = None

    for i, v in enumerate(values, start=1):
        ws_list.cell(row=i, column=col, value=v)

    col_letter = get_column_letter(col)
    return f"'{LIST_SHEET_NAME}'!${col_letter}$1:${col_letter}${len(values)}"


def add_dropdown_by_range(ws, start_row: int, end_row: int, col: int, ref_range: str):
    dv = DataValidation(
        type="list",
        formula1=f"={ref_range}",
        allow_blank=True,
    )
    ws.add_data_validation(dv)

    start_cell = ws.cell(row=start_row, column=col).coordinate
    end_cell = ws.cell(row=end_row, column=col).coordinate
    dv.add(f"{start_cell}:{end_cell}")


def apply_dropdowns(ws, wb):
    """
    1シートにプルダウンを適用する（テンプレ／サンプル共通）
    """
    header_row = find_header_row(ws)
    if header_row is None:
        raise RuntimeError(f"ヘッダー行が特定できません: {ws.title}")

    data_start = header_row + 1
    data_end = ws.max_row

    ws_list = ensure_list_sheet(wb)

    struct_types = [
        "仕込み成功・転換未設計",
        "プロセス完成・決定理由不足",
        "KPI未達だが成果発生",
        "活動主義",
        "選択と集中型",
        "因果成立（成功型）",
        "戦略仮説ミスマッチ",
        "その他",
    ]
    judge_kbn = ["継続", "変更", "削除", "新設"]
    judge_owner = ["課", "営業推進部", "本部"]
    next_reflect = ["次フェーズKPIに反映", "KPI再設計の検討対象", "今回は見送り", "横展開候補"]

    rng_struct = write_list(ws_list, 1, struct_types)   # A
    rng_judge  = write_list(ws_list, 2, judge_kbn)      # B
    rng_owner  = write_list(ws_list, 3, judge_owner)    # C
    rng_reflec = write_list(ws_list, 4, next_reflect)   # D

    add_dropdown_by_range(ws, data_start, data_end, COL_STRUCT_TYPE, rng_struct)  # N
    add_dropdown_by_range(ws, data_start, data_end, COL_JUDGE, rng_judge)         # Q
    add_dropdown_by_range(ws, data_start, data_end, COL_OWNER, rng_owner)         # R
    add_dropdown_by_range(ws, data_start, data_end, COL_REFLECT, rng_reflec)      # S

    return data_start, data_end


# ===== サンプルに「選択済み値」を入れる =====
def fill_sample_values(ws, data_start: int, data_end: int):
    """
    サンプルシートに、プルダウンから選んだ状態の値を入れる
    （まずは全行共通の見本値）
    """
    # 見本（プルダウン候補と完全一致させる）
    SAMPLE_STRUCT  = "仕込み成功・転換未設計"
    SAMPLE_JUDGE   = "変更"
    SAMPLE_OWNER   = "営業推進部"
    SAMPLE_REFLECT = "KPI再設計の検討対象"

    for r in range(data_start, data_end + 1):
        # KPIが空ならスキップ（空行対策）
        kpi_val = ws.cell(row=r, column=COL_KPI).value
        if kpi_val is None or str(kpi_val).strip() == "":
            continue

        ws.cell(row=r, column=COL_STRUCT_TYPE).value = SAMPLE_STRUCT
        ws.cell(row=r, column=COL_JUDGE).value = SAMPLE_JUDGE
        ws.cell(row=r, column=COL_OWNER).value = SAMPLE_OWNER
        ws.cell(row=r, column=COL_REFLECT).value = SAMPLE_REFLECT


def main():
    if not os.path.exists(SRC_XLSX):
        raise FileNotFoundError(SRC_XLSX)

    wb = load_workbook(SRC_XLSX)

    # 1) HK数式復元（全シート）
    total_processed = 0
    for ws in wb.worksheets:
        if ws.title == LIST_SHEET_NAME:
            continue
        p, _ = restore_formulas(ws)
        total_processed += p

    # 2) テンプレにプルダウン
    if SHEET_TEMPLATE not in wb.sheetnames:
        raise RuntimeError(f"対象シートが見つかりません: {SHEET_TEMPLATE}")
    ws_t = wb[SHEET_TEMPLATE]
    apply_dropdowns(ws_t, wb)

    # 3) サンプルにもプルダウン＋選択済み値を投入
    if SHEET_SAMPLE in wb.sheetnames:
        ws_s = wb[SHEET_SAMPLE]
        ds, de = apply_dropdowns(ws_s, wb)
        fill_sample_values(ws_s, ds, de)
    else:
        print(f"⚠ サンプルシートがありません: {SHEET_SAMPLE}")

    wb.save(OUT_XLSX)

    print("保存完了:", OUT_XLSX)
    print("H/K 数式復元 行数:", total_processed)
    print("プルダウン設定: テンプレ/サンプル（N/Q/R/S）")
    print("サンプル: N/Q/R/S に選択済み値を投入")


if __name__ == "__main__":
    main()
