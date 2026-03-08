# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# =========================================================
# 入出力（★ここだけ自分の環境に合わせて変更）
# =========================================================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_WSシート14.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_WSシート14_V列_現状値_自然文_分析版.xlsx"

TARGET_SHEET_CANDIDATES = ["00_WS_サンプル回答", "00_WS_サンプル回答シート"]

START_ROW = 7
END_ROW = 73
COL_V = "V"


# =========================================================
# ユーティリティ
# =========================================================
def pick_sheet(wb):
    for name in TARGET_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    for s in wb.sheetnames:
        if "00_WS_サンプル回答" in s:
            return wb[s]
    raise ValueError(f"指定シートが見つかりません。実在シート: {wb.sheetnames}")


def to_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip()
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if m:
            try:
                return float(m.group(0))
            except:
                return None
    return None


def parse_priority_lever(row_t_text: str | None, row_s_text: str | None) -> str:
    if isinstance(row_t_text, str):
        m = re.search(r"優先レバー：\s*([^\n\r]+)", row_t_text)
        if m:
            return m.group(1).strip()

    if isinstance(row_s_text, str):
        m = re.search(r"【優先レバー】\s*\n([^\n\r]+)", row_s_text)
        if m:
            return m.group(1).strip().rstrip("。")

    return "（未記載）"


def classify_kpi_axis(kpi_name: str | None, lever: str) -> str:
    """
    KPI名（E列想定）と優先レバーから、現状値の“工程上の位置づけ”を推定
    → 事実確認と分解観点を具体化するため
    """
    name = (kpi_name or "")
    if any(w in name for w in ["許可", "アポ", "面談", "次ステップ", "説明会"]):
        return "前工程（次ステップ確定・許可・アポ）"
    if any(w in name for w in ["採用", "納入", "決定", "クロージング"]):
        return "後工程（採用・納入・決定）"
    if any(w in name for w in ["訪問", "接触", "同行"]):
        return "活動量（接点・訪問）"
    # leverから補完
    if "伝え方" in lever or "タイミング" in lever:
        return "質（伝え方・タイミング起因）"
    if "対象" in lever:
        return "対象設計（重点先・配分起因）"
    if "行動" in lever:
        return "行動設計（手順・追いかけ起因）"
    return "（工程未特定）"


def fact_check_sentence(kpi_axis: str, lever: str) -> str:
    """
    Step5「現状値を事実で確認する」：分析に直結する“確定ポイント”を一文に落とす
    """
    # ここが肝：単なる列挙ではなく「母数と分子」「工程の定義」「対象・期間」を確定させる
    if "前工程" in kpi_axis:
        return (
            f"事実確認：この数は「接点→次ステップ確定（許可/アポ/説明会）」のどこまでを1件と数えるかを確定し、"
            f"対象の母数（重点先）と期間を揃えて、分子のブレ要因を潰す（優先レバー：{lever}）"
        )
    if "後工程" in kpi_axis:
        return (
            f"事実確認：この数は「案件化→採用/納入/決定」のどの到達点を1件とするかを確定し、"
            f"前工程（許可/次ステップ確定）からの流入数とセットで整合を取る（優先レバー：{lever}）"
        )
    if "活動量" in kpi_axis:
        return (
            f"事実確認：訪問/接触の“対象（重点先の定義）”と“1回の数え方（同日重複/同行/オンライン）”を確定し、"
            f"期間内の実施回数が正しい母数で出ているかを揃える（優先レバー：{lever}）"
        )
    if "質" in kpi_axis:
        return (
            f"事実確認：同じ対象・同じ回数でも結果が変わるため、"
            f"伝えた論点（判断条件/次ステップ）と実施タイミングが記録で追える状態に揃える（優先レバー：{lever}）"
        )

    return (
        f"事実確認：対象・期間・1件の定義を先に固定し、"
        f"現状値が“同じ物差し”で再現できる状態にする（優先レバー：{lever}）"
    )


def decompose_sentence(kpi_axis: str, lever: str, gap_phrase: str) -> str:
    """
    分解観点：差分を“工程順に切り分ける”文章にする
    """
    # 原則：①対象の偏り → ②工程で止まる位置 → ③やり方の差（伝え方/タイミング）
    if "前工程" in kpi_axis:
        return (
            f"分解観点：{gap_phrase}は、"
            f"①対象（重点先/非重点先）で発生しているのか、②次ステップ確定までの工程（接触→提案→許可）のどこで落ちているのか、"
            f"③落ちている場面の共通パターン（判断条件不足/次アクション未確定）を特定する"
        )
    if "後工程" in kpi_axis:
        return (
            f"分解観点：{gap_phrase}は、"
            f"①前工程の流入不足（許可/次ステップ確定が足りない）の問題か、②後工程の詰め不足（条件確認/不安解消/決裁者接触）の問題かを切り分け、"
            f"③止まっている案件群の共通要因（対象/論点/タイミング）を特定する"
        )
    if "活動量" in kpi_axis:
        return (
            f"分解観点：{gap_phrase}は、"
            f"①重点先への配分不足（重点先に回せていない）の問題か、②実行回数の不足（回数自体が足りない）の問題か、"
            f"③やっているが次に繋がっていない（次ステップ未確定）の問題かを順に切り分ける"
        )
    if "質" in kpi_axis:
        return (
            f"分解観点：{gap_phrase}は、"
            f"①対象の選び方（誰に当てたか）で差が出ているのか、②伝えた論点（判断条件/安全/供給/運用）で差が出ているのか、"
            f"③タイミング（いつ言ったか・追いかけ間隔）で差が出ているのかを比較して特定する"
        )
    # fallback
    return (
        f"分解観点：{gap_phrase}がどこで生じているかを、"
        f"対象→工程→やり方（{lever}）の順で切り分けて特定する"
    )


def build_v_text(kpi_name, target, actual, lever) -> str:
    t = to_number(target)
    a = to_number(actual)

    if t is not None and t != 0 and a is not None:
        rate = a / t
        line1 = f"現状値は{int(a) if a.is_integer() else a:g}（目標{int(t) if t.is_integer() else t:g}、達成率{rate:.1%}）"
        diff = a - t
        if diff < 0:
            gap_phrase = f"未達差分{int(-diff) if (-diff).is_integer() else -diff:g}"
        elif diff > 0:
            gap_phrase = f"上振れ{int(diff) if diff.is_integer() else diff:g}"
        else:
            gap_phrase = "差分0"
    else:
        line1 = "現状値（目標・実績）を数値で確定する（未入力があれば先に補完）"
        gap_phrase = "差分"

    kpi_axis = classify_kpi_axis(kpi_name, lever)
    line2 = fact_check_sentence(kpi_axis, lever)
    line3 = decompose_sentence(kpi_axis, lever, gap_phrase)

    return "\n".join([line1, line2, line3])


# =========================================================
# メイン
# =========================================================
def main():
    if not os.path.exists(INPUT_XLSX):
        raise FileNotFoundError(f"入力ファイルが見つかりません: {INPUT_XLSX}")

    wb = load_workbook(INPUT_XLSX)
    ws = pick_sheet(wb)

    for r in range(START_ROW, END_ROW + 1):
        # 想定：KPI名=E列、目標=F列、実績=G列、Step4=S列、Step5翻訳=T列
        kpi_name = ws[f"E{r}"].value
        target = ws[f"F{r}"].value
        actual = ws[f"G{r}"].value
        row_s = ws[f"S{r}"].value
        row_t = ws[f"T{r}"].value

        lever = parse_priority_lever(row_t, row_s)
        text = build_v_text(kpi_name, target, actual, lever)

        cell = ws[f"{COL_V}{r}"]
        cell.value = text
        # 行高は触らない。表示のみ最小限
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"完了：V列（{START_ROW}〜{END_ROW}行）を『分析に使える自然文』へ更新しました。\n出力: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
