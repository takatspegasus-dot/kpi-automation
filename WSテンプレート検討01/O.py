# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
import math
from openpyxl import load_workbook
from openpyxl.styles import Alignment


# ===== 入出力 =====
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_WSシート15.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_WSシート15_W列6月目標値自動生成.xlsx"


# ===== 対象シート候補（環境差の吸収）=====
SHEET_CANDIDATES = ["00_WS_サンプル回答", "00_WS_サンプル回答シート", "振り返りシート"]


def pick_sheet(wb):
    for name in SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    # それでも無ければ先頭
    return wb[wb.sheetnames[0]]


def parse_lever(step4_text: str | None) -> str:
    """S列（Step4記述）から優先レバーを抽出"""
    if not step4_text:
        return ""
    text = str(step4_text)

    m = re.search(r"【優先レバー】\s*\n([^\n]+)", text)
    if m:
        return m.group(1).strip().rstrip("。")

    m = re.search(r"優先レバー[:：]\s*([^\n]+)", text)
    return m.group(1).strip() if m else ""


def parse_judgement(step4_text: str | None) -> str:
    """S列（Step4記述）から判断を抽出"""
    if not step4_text:
        return ""
    text = str(step4_text)

    m = re.search(r"【判断】\s*\n([^\n]+)", text)
    return m.group(1).strip().rstrip("。") if m else ""


def calc_june_goal(target, actual, judgement: str) -> int | None:
    """目標(F)・現状(G)・判断から、6月目標値を算出（シンプルで破綻しないルール）"""
    if (target is None or target == "") and (actual is None or actual == ""):
        return None

    a = float(actual) if actual not in (None, "") else 0.0
    t = float(target) if target not in (None, "") else a

    j = judgement or ""

    # 判断によって“伸ばし方”を変える（やり過ぎない程度）
    if "大きく見直" in j or "修正" in j:
        factor = 1.30
    elif "改善" in j:
        factor = 1.20
    elif "一部見直" in j or "部分的" in j:
        factor = 1.15
    else:  # 継続
        factor = 1.10

    # 未達ならギャップを埋める＋少し上積み、達成済なら現状から少し伸ばす
    if a >= t:
        goal = a * factor
    else:
        goal = a + (t - a) * factor

    return int(math.ceil(goal))


def build_rationale(lever: str) -> str:
    """優先レバーに合わせて、目標設定の根拠文を自然な日本語で出し分け"""
    lever_clean = lever if lever else "（未記載）"

    if "対象" in lever_clean:
        return f"重点対象（優先順位／重点先）を再定義し、当て方（配分）を変えることで、次ステップ確定が起きる確度を上げる（優先レバー：{lever_clean}）"
    if "行動" in lever_clean and "フォロー" in lever_clean:
        return f"訪問後の次ステップ確定とフォロー間隔を標準化し、案件が止まるポイントを減らして積み上げる（優先レバー：{lever_clean}）"
    if "行動" in lever_clean:
        return f"活動の配分（誰に・いつ・何をするか）を再設計し、成果に直結する行動へ寄せて積み上げる（優先レバー：{lever_clean}）"
    if "伝え方" in lever_clean:
        return f"相手の判断論点に合わせて「判断条件・次アクション」を明確に伝え、次ステップ確定率を上げて積み上げる（優先レバー：{lever_clean}）"
    if "タイミング" in lever_clean:
        return f"接点を作る時期と追いかけ間隔を整え、案件化のタイミングロスを減らして積み上げる（優先レバー：{lever_clean}）"

    return f"Step4で決めた優先レバーを動かし、成果が出る構造に合わせて積み上げる（優先レバー：{lever_clean}）"


def build_w_text(step4_text, target, actual) -> str:
    judgement = parse_judgement(step4_text)
    lever = parse_lever(step4_text)

    june = calc_june_goal(target, actual, judgement)
    if june is None:
        return ""

    a = int(actual) if actual not in (None, "") else 0
    diff = june - a

    # 3月末/4月末/5月末/6月末の“累計”中間目標（単純で読める形）
    steps = 4
    if diff <= 0:
        mar = apr = may = june
    else:
        inc = int(math.ceil(diff / steps))
        mar = min(june, a + inc)
        apr = min(june, mar + inc)
        may = min(june, apr + inc)

    rationale = build_rationale(lever)

    lines = [
        f"6月目標値：{june}（現状{a} → 6月までに{diff:+d}）",
        f"中間目標（累計）：3月末{mar}／4月末{apr}／5月末{may}／6月末{june}",
        f"設定根拠：{rationale}",
        "運用ポイント：毎週、重点対象の母数・接触数・次ステップ確定数を確認し、未達の要因（対象／工程／伝え方）をその場で打ち手に落とす",
    ]
    return "\n".join(lines)


def main():
    if not os.path.exists(INPUT_XLSX):
        raise FileNotFoundError(f"入力ファイルが見つかりません: {INPUT_XLSX}")

    wb = load_workbook(INPUT_XLSX)
    ws = pick_sheet(wb)

    # W列のみ更新（7〜73）
    align = Alignment(wrap_text=True, vertical="top")

    for r in range(7, 74):
        kpi_name = ws[f"E{r}"].value
        if kpi_name is None or str(kpi_name).strip() == "":
            continue

        step4 = ws[f"S{r}"].value  # Step4の文章
        target = ws[f"F{r}"].value # 目標
        actual = ws[f"G{r}"].value # 実績

        ws[f"W{r}"].value = build_w_text(step4, target, actual)
        ws[f"W{r}"].alignment = align

    wb.save(OUTPUT_XLSX)
    print("完了：W列（7〜73）に「6月までの目標値」を自動生成しました。")
    print("出力:", OUTPUT_XLSX)


if __name__ == "__main__":
    main()
