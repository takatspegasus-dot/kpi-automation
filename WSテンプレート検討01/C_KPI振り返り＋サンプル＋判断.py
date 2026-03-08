# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# ===== 入出力 =====
XLSX_PATH = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_研修用_母艦＋サンプル_20260208.xlsx"
OUT_DIR = os.path.dirname(XLSX_PATH)

# ===== 対象シート =====
SHEET_TEMPLATE = "00_全課_①②③検討"
SHEET_SAMPLE   = "00_全課_①②③検討_サンプル"

# ===== 行・列 =====
DATA_START_ROW = 7
COL_ACHIEVE = 8        # H列：KPI達成率
COL_STRUCT_LABEL = 14  # N列：②-1 構造タイプ（ラベル）
COL_JUDGE = 17         # Q列：判断区分

LIST_SHEET = "_lists"  # 隠しリストシート

JUDGE_OPTIONS = ["継続", "変更", "削除", "新設"]

# ===== 正規化 =====
_fullwidth_map = str.maketrans({
    "０":"0","１":"1","２":"2","３":"3","４":"4","５":"5","６":"6","７":"7","８":"8","９":"9",
    "％":"%", "ー":"-", "−":"-",
})

def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).translate(_fullwidth_map)
    s = s.replace("　", "").replace(" ", "").strip()
    return s

def parse_rate(v) -> Optional[float]:
    """
    H列のKPI達成率を float に揃える。
    - 0.308 / 30.8% / "30.8%" / 30.8(%) などを許容
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v > 3:  # 30.8 のように入っていれば % とみなす
            return float(v) / 100.0
        return float(v)

    s = norm(v).replace("%", "")
    if not s:
        return None
    try:
        x = float(s)
        if x > 3:
            x = x / 100.0
        return x
    except:
        return None

# ===== 推奨ロジック（サンプル用の初期値）=====
def suggest_judge(struct_label: str, achieve: Optional[float]) -> str:
    lbl = norm(struct_label)

    # ラベル優先（営業推進部がレビューしやすい“保守的”推奨）
    if "因果成立" in lbl or "成功型" in lbl:
        return "継続"
    if "選択と集中" in lbl:
        return "継続"
    if "仕込み成功" in lbl and "転換未設計" in lbl:
        return "変更"
    if "プロセス完成" in lbl and "決定理由不足" in lbl:
        return "変更"
    if "KPI未達" in lbl and ("成果" in lbl or "発生" in lbl):
        return "変更"
    if "活動主義" in lbl:
        return "削除"
    if "戦略仮説" in lbl and "ミスマッチ" in lbl:
        return "新設"

    # ラベルがその他/空などの場合は達成率で補助（削除は強いのでデフォは変更）
    if achieve is None:
        return "変更"
    if achieve >= 0.8:
        return "継続"
    if achieve >= 0.5:
        return "変更"
    return "変更"

# ===== リストシート作成（DVの参照元）=====
def ensure_list_sheet(wb):
    if LIST_SHEET in wb.sheetnames:
        ws = wb[LIST_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(LIST_SHEET)

    ws["A1"].value = "判断区分"
    for i, opt in enumerate(JUDGE_OPTIONS, start=2):
        ws[f"A{i}"].value = opt

    ws.sheet_state = "hidden"
    return ws

def add_dropdown(ws):
    max_row = ws.max_row
    dv = DataValidation(
        type="list",
        formula1=f"={LIST_SHEET}!$A$2:$A${len(JUDGE_OPTIONS)+1}",
        allow_blank=True,
        showDropDown=True
    )
    ws.add_data_validation(dv)
    dv.add(f"Q{DATA_START_ROW}:Q{max_row}")

def fill_sample(ws):
    """
    サンプルシートのみ：Qが空欄の行に推奨値を入れる（上書き可）
    """
    filled = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        q = ws.cell(r, COL_JUDGE)
        if q.value not in (None, ""):
            continue

        struct_label = ws.cell(r, COL_STRUCT_LABEL).value
        achieve = parse_rate(ws.cell(r, COL_ACHIEVE).value)

        # ②-1も達成率も無ければ、無理に入れない
        if norm(struct_label) == "" and achieve is None:
            continue

        q.value = suggest_judge(struct_label, achieve)
        filled += 1
    return filled

def main():
    if not os.path.exists(XLSX_PATH):
        raise FileNotFoundError(XLSX_PATH)

    wb = load_workbook(XLSX_PATH)

    # シート存在チェック
    if SHEET_TEMPLATE not in wb.sheetnames:
        raise RuntimeError(f"対象シートがありません: {SHEET_TEMPLATE}")
    if SHEET_SAMPLE not in wb.sheetnames:
        raise RuntimeError(f"対象シートがありません: {SHEET_SAMPLE}")

    # リストシート作成
    ensure_list_sheet(wb)

    # ①テンプレ：プルダウンのみ
    ws_t = wb[SHEET_TEMPLATE]
    add_dropdown(ws_t)

    # ②サンプル：プルダウン＋推奨値入力
    ws_s = wb[SHEET_SAMPLE]
    add_dropdown(ws_s)
    filled = fill_sample(ws_s)

    ts = datetime.now().strftime("%Y%m%d")
    out_path = os.path.join(OUT_DIR, f"KPI振り返り_①事実②構造③判断_母艦_判断区分半自動_{ts}.xlsx")
    wb.save(out_path)

    print("保存完了:", out_path)
    print("サンプルシート：推奨入力したQ列セル数 =", filled)
    print("テンプレシート：プルダウンのみ（自動入力なし）")

if __name__ == "__main__":
    main()
