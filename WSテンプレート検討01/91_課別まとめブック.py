# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from copy import copy
from typing import Dict, List, Tuple, Set

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell


# =========================================================
# 入力フォルダ / 出力ファイル
# =========================================================
INPUT_DIR = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\09配布_課別ファイル\課別配布_研修テンプレ"
OUTPUT_XLSX = os.path.join(INPUT_DIR, "【研修用】課別まとめブック_Index付.xlsx")

# =========================================================
# 表示順（指定順）
# =========================================================
DEPT_ORDER = [
    "札幌", "仙台",
    "東京1", "東京2", "東京3", "東京4", "東京5",
    "名古屋1", "名古屋2",
    "大阪1", "大阪2",
    "広島",
    "福岡1", "福岡2",
]

DEPT_PATTERN = re.compile(r"(札幌|仙台|東京[1-5]|名古屋[12]|大阪[12]|広島|福岡[12])")


def list_input_files(folder: str) -> List[str]:
    files = []
    for f in os.listdir(folder):
        if f.startswith("~$"):
            continue
        fl = f.lower()
        if not (fl.endswith(".xlsx") or fl.endswith(".xlsm")):
            continue
        full = os.path.join(folder, f)
        if os.path.abspath(full) == os.path.abspath(OUTPUT_XLSX):
            continue
        files.append(full)
    return sorted(files)


def dept_from_filename(path: str) -> str:
    base = os.path.splitext(os.path.basename(path))[0]
    base = base.translate(str.maketrans({
        "０":"0","１":"1","２":"2","３":"3","４":"4",
        "５":"5","６":"6","７":"7","８":"8","９":"9",
        "　":" "
    })).strip()
    base = base.replace("東京１","東京1").replace("東京２","東京2").replace("東京３","東京3").replace("東京４","東京4").replace("東京５","東京5")
    base = base.replace("名古屋１","名古屋1").replace("名古屋２","名古屋2")
    base = base.replace("大阪１","大阪1").replace("大阪２","大阪2")
    base = base.replace("福岡１","福岡1").replace("福岡２","福岡2")

    m = DEPT_PATTERN.search(base)
    return m.group(1) if m else ""


def build_dept_to_file(files: List[str]) -> Dict[str, str]:
    dept_to_file: Dict[str, str] = {}
    for fp in files:
        d = dept_from_filename(fp)
        if d and d not in dept_to_file:
            dept_to_file[d] = fp
    return dept_to_file


def safe_sheetname(name: str) -> str:
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)[:31]
    return name if name else "Sheet"


def merged_top_left_map(ws: Worksheet) -> Tuple[Set[Tuple[int, int]], Set[Tuple[int, int]]]:
    """
    merged_ranges を走査して、
    - merged_all: 結合範囲に含まれる全セル座標
    - merged_topleft: 結合範囲の左上セル座標
    を返す
    """
    merged_all: Set[Tuple[int, int]] = set()
    merged_topleft: Set[Tuple[int, int]] = set()

    for rng in ws.merged_cells.ranges:
        min_row = rng.min_row
        min_col = rng.min_col
        max_row = rng.max_row
        max_col = rng.max_col

        merged_topleft.add((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_all.add((r, c))

    return merged_all, merged_topleft


def copy_sheet_contents(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    """
    1枚目シートを「見た目ごと」コピーする（openpyxlで可能な範囲で最大限）
    - 値/スタイル/罫線/表示形式/結合/列幅/行高/表示倍率/フリーズ等
    - 結合セルは「左上セルだけ値を入れる」ことで MergedCell の read-only を回避
    """
    # シート表示
    try:
        dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
    except Exception:
        pass
    try:
        dst_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass

    # 列幅
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_letter]
        dst_dim.width = dim.width
        dst_dim.hidden = dim.hidden
        dst_dim.outlineLevel = dim.outlineLevel

    # 行高
    for row_idx, dim in src_ws.row_dimensions.items():
        dst_dim = dst_ws.row_dimensions[row_idx]
        dst_dim.height = dim.height
        dst_dim.hidden = dim.hidden
        dst_dim.outlineLevel = dim.outlineLevel

    # 結合セル（先に結合を作る）
    for merged in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(merged))

    # 結合セル情報
    merged_all, merged_topleft = merged_top_left_map(src_ws)

    max_row = src_ws.max_row
    max_col = src_ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(row=r, column=c)
            dst_cell = dst_ws.cell(row=r, column=c)

            in_merged = (r, c) in merged_all
            is_topleft = (r, c) in merged_topleft

            # 値の書き込み：結合セルは左上だけ。非結合は通常通り。
            if (not in_merged) or is_topleft:
                # dst_cell が MergedCell になることはない想定だが念のため
                if not isinstance(dst_cell, MergedCell):
                    dst_cell.value = src_cell.value

            # スタイル：結合セルの非左上は触らない（安全）
            # 左上 or 非結合のみスタイルコピー
            if (not in_merged) or is_topleft:
                if src_cell.has_style and not isinstance(dst_cell, MergedCell):
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy(src_cell.protection)
                    dst_cell.alignment = copy(src_cell.alignment)

    # 条件付き書式（可能な範囲で）
    try:
        dst_ws.conditional_formatting = copy(src_ws.conditional_formatting)
    except Exception:
        pass

    # データ検証（環境によっては警告あり）
    try:
        for dv in src_ws.data_validations.dataValidation:
            dst_ws.add_data_validation(copy(dv))
    except Exception:
        pass


def place_backlink_safely(ws: Worksheet) -> None:
    """
    レイアウトを壊さないため、空いているセルにだけ戻りリンク
    候補：A1→A2→B1→B2→H2
    """
    candidates = ["A1", "A2", "B1", "B2", "H2"]
    for addr in candidates:
        cell = ws[addr]
        if cell.value in (None, ""):
            cell.value = "← Indexに戻る"
            cell.hyperlink = "#Index!A1"
            try:
                cell.font = cell.font.copy(bold=True, underline="single", color="0000FF")
            except Exception:
                pass
            return


def create_index_sheet(wb: Workbook, dept_to_sheetname: Dict[str, str]) -> None:
    idx = wb.create_sheet("Index", 0)
    idx["A1"].value = "表示順"
    idx["B1"].value = "課"

    row = 2
    order_no = 1
    for dept in DEPT_ORDER:
        if dept not in dept_to_sheetname:
            continue
        sh = dept_to_sheetname[dept]
        idx[f"A{row}"].value = order_no
        idx[f"B{row}"].value = dept
        idx[f"B{row}"].hyperlink = f"#{sh}!A1"
        row += 1
        order_no += 1

    idx.column_dimensions["A"].width = 10
    idx.column_dimensions["B"].width = 20


def main():
    if not os.path.isdir(INPUT_DIR):
        raise FileNotFoundError(f"入力フォルダが見つかりません: {INPUT_DIR}")

    files = list_input_files(INPUT_DIR)
    if not files:
        raise FileNotFoundError(f"フォルダ内に .xlsx/.xlsm がありません: {INPUT_DIR}")

    dept_to_file = build_dept_to_file(files)

    print("=== 認識した 課→ファイル ===")
    for d in DEPT_ORDER:
        if d in dept_to_file:
            print(f"  {d}: {os.path.basename(dept_to_file[d])}")
        else:
            print(f"  {d}: (見つからず)")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    dept_to_sheetname: Dict[str, str] = {}

    for dept in DEPT_ORDER:
        if dept not in dept_to_file:
            continue

        src_path = dept_to_file[dept]
        wb_src = load_workbook(src_path, data_only=False)
        src_ws = wb_src.worksheets[0]

        sheet_name = safe_sheetname(dept)
        dst_ws = wb_out.create_sheet(sheet_name)

        copy_sheet_contents(src_ws, dst_ws)

        dept_to_sheetname[dept] = sheet_name
        print(f"[OK] 取り込み: {dept} ← {os.path.basename(src_path)}")

        wb_src.close()

    if not dept_to_sheetname:
        raise RuntimeError("課別シートを1枚も取り込めませんでした。")

    create_index_sheet(wb_out, dept_to_sheetname)

    for dept in DEPT_ORDER:
        if dept not in dept_to_sheetname:
            continue
        place_backlink_safely(wb_out[dept_to_sheetname[dept]])

    if os.path.exists(OUTPUT_XLSX):
        os.remove(OUTPUT_XLSX)

    wb_out.save(OUTPUT_XLSX)
    print("✅ 完了:", OUTPUT_XLSX)


if __name__ == "__main__":
    main()
