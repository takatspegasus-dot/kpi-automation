# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Dict, Tuple, Set

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ====== 入力 ======
BASE_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\KPI振り返り_①事実②構造③判断_母艦_20260208.xlsx"
SRC_XLSX  = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI振り返りシート\20研修_KPI振り返りシート_最終案\WSテンプレート\サンプルシート_列項目整理.xlsx"

TEMPLATE_SHEET = "00_全課_①②③検討"
SRC_SHEET = "サンプル列整理"
SAMPLE_SHEET = "00_全課_①②③検討_サンプル"
OUT_DIR = os.path.dirname(BASE_XLSX)

# ====== 行・列（テンプレ側）=====
DATA_START_ROW = 7
COL_DEPT = 2      # ★B列が「課」

COL_FACT = 13     # M
COL_LABEL = 14    # N
COL_STRUCT = 15   # O
COL_ISSUE = 16    # P
COL_Q = 17        # Q
COL_R = 18        # R
COL_S = 19        # S

# 折り返し＋上寄せ
wrap_top = Alignment(wrap_text=True, vertical="top")

_fullwidth_map = str.maketrans({
    "０":"0","１":"1","２":"2","３":"3","４":"4","５":"5","６":"6","７":"7","８":"8","９":"9",
})

def normalize_dept(v) -> str:
    if v is None:
        return ""
    s = str(v).translate(_fullwidth_map)
    s = s.replace("\n", "").replace("\r", "")
    s = s.replace("　", "").replace(" ", "")
    s = re.sub(r"課$", "", s)  # 末尾の「課」を除去
    return s.strip()

def guess_label(subtitle: str) -> str:
    t = (subtitle or "").strip()
    if "転換" in t or "未完成" in t:
        return "仕込み成功・転換未設計"
    if "決定理由" in t:
        return "プロセス完成・決定理由不足"
    if ("未達" in t or "KPI未達" in t) and ("成果" in t or "発生" in t):
        return "KPI未達だが成果発生"
    if "活動主義" in t:
        return "活動主義"
    if "選択と集中" in t:
        return "選択と集中型"
    if "因果" in t or ("成功" in t and "噛み合" in t):
        return "因果成立（成功型）"
    if "仮説" in t or "ミスマッチ" in t:
        return "戦略仮説ミスマッチ"
    return "その他"

def load_sample_map() -> Dict[str, Tuple[str, str, str, str]]:
    """
    列項目整理側（サンプル列整理）：
    A=課, B=見出し, C=①, D=②-2, E=③ を想定
    """
    wb = load_workbook(SRC_XLSX, data_only=True)
    if SRC_SHEET not in wb.sheetnames:
        raise RuntimeError(f"SRCシートが見つかりません: {SRC_SHEET} / {wb.sheetnames}")

    ws = wb[SRC_SHEET]
    d: Dict[str, Tuple[str, str, str, str]] = {}

    for r in range(2, ws.max_row + 1):
        dept = normalize_dept(ws.cell(r, 1).value)   # A列：課
        if not dept:
            continue
        subtitle = ws.cell(r, 2).value or ""         # B列：見出し
        fact = ws.cell(r, 3).value or ""             # C列：①
        struct = ws.cell(r, 4).value or ""           # D列：②-2
        issue = ws.cell(r, 5).value or ""            # E列：③
        d[dept] = (str(subtitle), str(fact), str(struct), str(issue))

    if not d:
        raise RuntimeError("列項目整理から1件も取得できません（列位置/シート名を要確認）")
    return d

def forward_fill_dept(ws) -> None:
    """テンプレ側の課(B列)が先頭行だけ入っている場合に、下の行を上の値で埋める"""
    last = ""
    for r in range(DATA_START_ROW, ws.max_row + 1):
        v = normalize_dept(ws.cell(r, COL_DEPT).value)
        if v:
            last = v
        else:
            if last:
                ws.cell(r, COL_DEPT).value = last

def estimate_row_height(*texts: str, base: float = 15.0, max_h: float = 320.0) -> float:
    """
    openpyxl にはExcel同等のオート行高がないため推定で設定。
    - 改行数（\n）を最優先
    - 文字数も加味して「だいたい見える」高さにする
    """
    score = 1
    for t in texts:
        if not t:
            continue
        s = str(t)
        # 改行 + 文字数（概算）
        lines = s.count("\n") + 1
        chars = len(s)
        score = max(score, lines + int(chars / 60))  # 60文字で1行換算（概算）
    h = base * score
    return min(max_h, max(base, h))

def main() -> None:
    if not os.path.exists(BASE_XLSX):
        raise FileNotFoundError(BASE_XLSX)
    if not os.path.exists(SRC_XLSX):
        raise FileNotFoundError(SRC_XLSX)

    sample_map = load_sample_map()

    wb = load_workbook(BASE_XLSX)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise RuntimeError(f"テンプレがありません: {TEMPLATE_SHEET} / {wb.sheetnames}")

    # サンプルシートを作り直す
    if SAMPLE_SHEET in wb.sheetnames:
        del wb[SAMPLE_SHEET]

    ws_t = wb[TEMPLATE_SHEET]
    ws_s = wb.copy_worksheet(ws_t)
    ws_s.title = SAMPLE_SHEET

    # 重要：B列の課を全行に展開
    forward_fill_dept(ws_s)

    # 貼り付け
    filled_rows = 0
    missing_depts: Set[str] = set()

    for r in range(DATA_START_ROW, ws_s.max_row + 1):
        dept = normalize_dept(ws_s.cell(r, COL_DEPT).value)  # ★B列
        if not dept:
            continue

        if dept not in sample_map:
            missing_depts.add(dept)
            continue

        subtitle, fact, struct, issue = sample_map[dept]
        label = guess_label(subtitle)

        ws_s.cell(r, COL_FACT).value = fact
        ws_s.cell(r, COL_LABEL).value = label
        ws_s.cell(r, COL_STRUCT).value = struct
        ws_s.cell(r, COL_ISSUE).value = issue

        # 判断列（研修用なので空）
        ws_s.cell(r, COL_Q).value = None
        ws_s.cell(r, COL_R).value = None
        ws_s.cell(r, COL_S).value = None

        # 折り返し＋上寄せ
        for c in (COL_FACT, COL_LABEL, COL_STRUCT, COL_ISSUE, COL_Q, COL_R, COL_S):
            ws_s.cell(r, c).alignment = wrap_top

        # ★行高を調整（サンプルシートだけ）
        h = estimate_row_height(fact, struct, issue)
        ws_s.row_dimensions[r].height = h

        filled_rows += 1

    ts = datetime.now().strftime("%Y%m%d")
    out_path = os.path.join(OUT_DIR, f"KPI振り返り_研修用_母艦＋サンプル_{ts}.xlsx")
    wb.save(out_path)

    print("保存完了:", out_path)
    print("反映できた行数:", filled_rows)
    print("未一致の課:", sorted(missing_depts))

if __name__ == "__main__":
    main()
