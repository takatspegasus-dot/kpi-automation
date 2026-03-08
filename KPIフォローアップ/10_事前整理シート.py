import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_followup_prep_workbook(out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "03_27事前整理"

    # ---- Column widths ----
    col_widths = {"A": 4, "B": 34, "C": 62, "D": 10, "E": 46}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    # ---- Page setup ----
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ---- Freeze panes ----
    ws.freeze_panes = "A6"

    # ---- Styles ----
    font_title = Font(bold=True, size=14)
    font_section = Font(bold=True, size=11)
    font_header = Font(bold=True, size=10)
    font_normal = Font(size=10)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_section = PatternFill("solid", fgColor="EDEDED")
    fill_header = PatternFill("solid", fgColor="F5F5F5")
    fill_label = PatternFill("solid", fgColor="F5F5F5")

    thin = Side(style="thin", color="9E9E9E")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_border_range(rng: str):
        for row in ws[rng]:
            for cell in row:
                cell.border = border_thin

    def merge_and_set(r1, c1, r2, c2, value="", font=font_normal, fill=None, alignment=align_left_top):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        cell.font = font
        cell.alignment = alignment
        if fill:
            cell.fill = fill
        return cell

    def set_cell(r, c, value="", font=font_normal, fill=None, alignment=align_left_center):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = font
        cell.alignment = alignment
        if fill:
            cell.fill = fill
        return cell

    # ---- Header ----
    merge_and_set(1, 1, 1, 5, "KPI再設計モデル研修 Part2 フォローアップ｜事前整理シート（3/27）", font=font_title, alignment=align_center)
    merge_and_set(2, 1, 2, 5, "※事実ベースで記入ください（評価ではありません）。空欄OK／分かる範囲で構いません。", font=Font(size=9), alignment=align_left_center)

    # ---- Basic info ----
    set_cell(4, 1, "課名", font=font_header, fill=fill_label, alignment=align_center)
    merge_and_set(4, 2, 4, 3, "", alignment=align_left_center)
    set_cell(4, 4, "記入者", font=font_header, fill=fill_label, alignment=align_center)
    set_cell(4, 5, "", alignment=align_left_center)

    set_cell(5, 1, "記入日", font=font_header, fill=fill_label, alignment=align_center)
    merge_and_set(5, 2, 5, 3, "", alignment=align_left_center)
    set_cell(5, 4, "提出先", font=font_header, fill=fill_label, alignment=align_center)
    set_cell(5, 5, "営業推進部", alignment=align_left_center)
    set_border_range("A4:E5")

    def section_row(row, title):
        merge_and_set(row, 1, row, 5, title, font=font_section, fill=fill_section, alignment=align_left_center)
        set_border_range(f"A{row}:E{row}")

    def header_row(row):
        set_cell(row, 1, "No", font=font_header, fill=fill_header, alignment=align_center)
        set_cell(row, 2, "設問", font=font_header, fill=fill_header, alignment=align_center)
        set_cell(row, 3, "回答（事実／具体）", font=font_header, fill=fill_header, alignment=align_center)
        set_cell(row, 4, "自己評価", font=font_header, fill=fill_header, alignment=align_center)
        set_cell(row, 5, "具体例・補足（任意）", font=font_header, fill=fill_header, alignment=align_center)
        set_border_range(f"A{row}:E{row}")

    def question_block(start_row, no, question_text):
        end_row = start_row + 2
        merge_and_set(start_row, 1, end_row, 1, str(no), alignment=align_center)
        merge_and_set(start_row, 2, end_row, 2, question_text, alignment=align_left_top)
        merge_and_set(start_row, 3, end_row, 3, "", alignment=align_left_top)
        merge_and_set(start_row, 4, end_row, 4, "", alignment=align_center)
        merge_and_set(start_row, 5, end_row, 5, "", alignment=align_left_top)
        set_border_range(f"A{start_row}:E{end_row}")
        return start_row, end_row

    # ---- Step1 ----
    q_ranges = []  # DataValidation 登録用に (start_row, end_row) を収集
    section_row(7, "Step1｜運用実施確認（事実）")
    header_row(8)
    q_ranges.append(question_block(9,  1, "戦略一文化は実際の判断に使いましたか？（どの場面で使ったか）"))
    q_ranges.append(question_block(12, 2, "主指標（AP列）で月次を見ましたか？（他の数字に逃げた場面は？）"))
    q_ranges.append(question_block(15, 3, "未達時にAQ列の順番で切り分けましたか？（実際に見た順）"))

    # ---- Step2 ----
    section_row(19, "Step2｜対話の質の変化")
    q_ranges.append(question_block(20, 4, "会議の議論はどう変わりましたか？（活動量→状態、など）"))
    q_ranges.append(question_block(23, 5, "営業推進部とのコミュニケーションの質は変わりましたか？"))
    q_ranges.append(question_block(26, 6, "未達時の議論はどう変わりましたか？（責任論→構造論）"))

    # ---- Step3 ----
    section_row(30, "Step3｜行動変容")
    q_ranges.append(question_block(31, 7, "現場行動で変わったこと（変えたこと）"))
    q_ranges.append(question_block(34, 8, "判断スピード／迷いは変わりましたか？"))

    # ---- Step4 ----
    section_row(38, "Step4｜定着度セルフスコア（5段階）")
    set_cell(39, 2, "項目", font=font_header, fill=fill_header, alignment=align_center)
    set_cell(39, 4, "スコア", font=font_header, fill=fill_header, alignment=align_center)
    set_border_range("A39:E39")

    score_items = [
        "戦略一文化を使った",
        "主指標1つで判断した",
        "構造で切り分けた",
        "営業推進部との会話が構造化された",
        "感情論が減った",
    ]
    score_rows = []  # DataValidation 登録用に行番号を収集
    r = 40
    for item in score_items:
        merge_and_set(r, 2, r, 3, item, alignment=align_left_center)
        set_cell(r, 4, "", alignment=align_center)
        set_cell(r, 5, "", alignment=align_left_center)
        set_border_range(f"A{r}:E{r}")
        score_rows.append(r)
        r += 1

    # ---- Step5 ----
    section_row(46, "Step5｜正直な振り返り（重要）")

    ws.merge_cells("A47:A49")
    ws["A47"].value = "9"
    ws["A47"].alignment = align_center
    ws.merge_cells("B47:B49")
    ws["B47"].value = "うまく回らなかった理由（遠慮なく）"
    ws["B47"].alignment = align_left_top
    ws.merge_cells("C47:E49")
    ws["C47"].alignment = align_left_top
    set_border_range("A47:E49")

    ws.merge_cells("A50:A52")
    ws["A50"].value = "10"
    ws["A50"].alignment = align_center
    ws.merge_cells("B50:B52")
    ws["B50"].value = "次四半期で“固定したいこと”（1つに絞る）"
    ws["B50"].alignment = align_left_top
    ws.merge_cells("C50:E52")
    ws["C50"].alignment = align_left_top
    set_border_range("A50:E52")

    # ---- Data Validation (1~5) ----
    dv_1to5 = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dv_1to5)

    # 設問ブロック（D列）：question_block の戻り値から自動登録
    for start, end in q_ranges:
        dv_1to5.add(f"D{start}:D{end}")

    # Step4 スコア行（D列）：score_rows から自動登録
    for rr in score_rows:
        dv_1to5.add(f"D{rr}:D{rr}")

    wb.save(out_path)

if __name__ == "__main__":
    out_dir = os.path.join(os.path.dirname(__file__), "output")
    out_path = os.path.join(out_dir, "03_27事前整理シート.xlsx")

    # フォルダが存在しない場合に備えて作成（既にあれば何もしない）
    os.makedirs(out_dir, exist_ok=True)

    create_followup_prep_workbook(out_path)
    print(f"Created: {out_path}")