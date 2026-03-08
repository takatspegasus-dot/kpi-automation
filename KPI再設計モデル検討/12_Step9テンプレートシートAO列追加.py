# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import MergedCell


# =========================================================
# 入出力
# =========================================================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step9_Ver1.0.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step9_Ver1.0_AO追加_書式強化.xlsx"

SHEET_NAME = "WS_サンプル回答"

ROW_START = 7
ROW_END = 73

# Step9既存列（参照元書式）
STYLE_SRC_COL = "AN"   # ←この列の書式を AO にコピーして統一
AO_COL = "AO"


# =========================================================
# ユーティリティ
# =========================================================
def s(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n").strip()


def first_line(text: str) -> str:
    t = s(text)
    return t.split("\n", 1)[0].strip() if t else ""


def is_mergedcell(cell) -> bool:
    return isinstance(cell, MergedCell)


def copy_cell_style(src, dst) -> None:
    """
    openpyxlセル書式を可能な限りコピー（値・コメントはコピーしない）
    ※ dst が MergedCell（結合セルの左上以外）の場合は書き込み不可なので何もしない
    """
    if src is None or dst is None:
        return
    if is_mergedcell(dst):
        return

    # _style をコピー（ここが一番効く）
    try:
        dst._style = copy(src._style)
    except Exception:
        pass

    # 個別に安全コピー
    try:
        dst.font = copy(src.font)
    except Exception:
        pass
    try:
        dst.fill = copy(src.fill)
    except Exception:
        pass
    try:
        dst.border = copy(src.border)
    except Exception:
        pass
    try:
        dst.alignment = copy(src.alignment)
    except Exception:
        pass
    try:
        dst.number_format = src.number_format
    except Exception:
        pass
    try:
        dst.protection = copy(src.protection)
    except Exception:
        pass

    # コメントは無理に触らない（MergedCellやread-only対策）
    try:
        dst.comment = None
    except Exception:
        pass


def ensure_wrap_top(cell) -> None:
    """
    折返し＋上寄せを安全に適用（openpyxlバージョン非依存）
    ※ MergedCell は書き込み不可なので何もしない
    """
    if is_mergedcell(cell):
        return

    if cell.alignment:
        a = copy(cell.alignment)
        a.wrap_text = True
        a.vertical = "top"
        cell.alignment = a
    else:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# =========================================================
# 戦略一文化生成（AO列）
# =========================================================
def build_strategy_sentence(ws, r: int) -> str:
    """
    AO列用：戦略一文化（固定文）を生成
    参照：
      AD = 最優先状態
      AE = 主軸レバー
      AN = KPI示唆（先頭行）
      X  = 新KPI（ANが空の時の保険）
    """
    state = s(ws[f"AD{r}"].value)
    lever = s(ws[f"AE{r}"].value)

    kpi_hint = first_line(ws[f"AN{r}"].value)
    if not kpi_hint:
        kpi_hint = first_line(ws[f"X{r}"].value)
    if not kpi_hint:
        kpi_hint = "（測定指標未定）"

    if not state:
        state = "（最優先状態未記入）"
    if not lever:
        lever = "（レバー未選択）"

    return (
        "今フェーズは、\n"
        f"「{state}を、{lever}レバーで進展させ、\n"
        f"{kpi_hint}で測る」"
    )


# =========================================================
# メイン
# =========================================================
def main() -> None:
    in_path = Path(INPUT_XLSX)
    out_path = Path(OUTPUT_XLSX)

    if not in_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {in_path}")

    wb = load_workbook(in_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SHEET_NAME} / sheets={wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------
    # 1) 列幅：AN → AO へコピー（Noneなら標準幅）
    # -----------------------------------------------------
    src_dim = ws.column_dimensions.get(STYLE_SRC_COL)
    src_w = getattr(src_dim, "width", None)
    if src_w is None:
        src_w = 20
    ws.column_dimensions[AO_COL].width = src_w

    # -----------------------------------------------------
    # 2) AO5:AO6 を結合してタイトルセット
    #    ※結合セルは左上(AO5)だけ操作する
    # -----------------------------------------------------
    try:
        ws.merge_cells(f"{AO_COL}5:{AO_COL}6")
    except Exception:
        pass

    # 見出し書式：AN5 を AO5 へコピーして寄せる
    src_h1 = ws[f"{STYLE_SRC_COL}5"]
    dst_h1 = ws[f"{AO_COL}5"]
    copy_cell_style(src_h1, dst_h1)

    # 見出し文（AO5のみ設定）
    dst_h1.value = (
        "戦略一文化（固定）\n"
        "今フェーズは、\n"
        "「◯◯な状態を、△△レバーで進展させ、□□で測る」"
    )
    ensure_wrap_top(dst_h1)

    # 追加強調：太字＋薄グレー（既存塗りがある場合は尊重）
    try:
        base_font = dst_h1.font
        dst_h1.font = Font(
            name=base_font.name,
            size=base_font.size,
            bold=True,
            italic=base_font.italic,
            underline=base_font.underline,
            color=base_font.color,
        )
    except Exception:
        pass

    try:
        fill = dst_h1.fill
        if fill is None or getattr(fill, "patternType", None) is None:
            dst_h1.fill = PatternFill(patternType="solid", fgColor="F2F2F2")
    except Exception:
        pass

    # -----------------------------------------------------
    # 3) AO7〜AO73：AN列の書式を行ごとにコピーし、値を生成
    # -----------------------------------------------------
    for r in range(ROW_START, ROW_END + 1):
        src_cell = ws[f"{STYLE_SRC_COL}{r}"]
        dst_cell = ws[f"{AO_COL}{r}"]

        copy_cell_style(src_cell, dst_cell)

        dst_cell.value = build_strategy_sentence(ws, r)
        ensure_wrap_top(dst_cell)

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()