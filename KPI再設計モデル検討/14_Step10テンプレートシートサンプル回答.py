# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from typing import Optional, Tuple, List, Dict

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# =========================================================
# 入出力（あなたの実在パス）
# =========================================================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step10追加04.xlsx"

# ★ユーザー指定：別名保存先
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step10追加04_AR_Step7-8連動.xlsx"

SHEET_NAME = "WS_サンプル回答"
ROW_START = 7
ROW_END = 73

# 母艦の固定列（今回の運用ルール）
COL_STEP7 = "W"    # Step7
COL_STEP8 = "AA"   # Step8
COL_AP = "AP"      # Step10 主指標（固定）
COL_AQ = "AQ"      # Step10 未達時の切り分け（一次）
COL_AR = "AR"      # Step10 次フェーズ方針（一文化・宣言）


# =========================================================
# 表示維持（手動調整済みを壊さない）
# =========================================================
def keep_wrap_top(cell):
    al = cell.alignment
    cell.alignment = Alignment(
        horizontal=al.horizontal,
        vertical="top",
        wrap_text=True,
        text_rotation=al.text_rotation,
        shrink_to_fit=al.shrink_to_fit,
        indent=al.indent,
        relativeIndent=al.relativeIndent,
        justifyLastLine=al.justifyLastLine,
        readingOrder=al.readingOrder,
    )


def norm(s: str) -> str:
    return s.replace("\r\n", "\n").replace("\r", "\n").strip()


# =========================================================
# Step7：重点領域 抽出（短文化）
# =========================================================
def extract_step7_focus(step7_text: str) -> str:
    t = norm(step7_text)

    # 先頭行が一番“宣言”に近いことが多い
    first = (t.split("\n")[0] if t else "").strip()

    # よくある選択肢を優先的に拾う（複数あれば統合）
    candidates = ["定着", "継続処方", "競合対抗", "拡大型採用"]
    hits = [c for c in candidates if c in t]
    if hits:
        # 例：定着／継続処方
        return "／".join(hits)

    # 先頭が短いなら採用
    if first and len(first) <= 20:
        return first

    # “最優先”以降を拾う
    m = re.search(r"(最優先|重点|強化).*?[:：]?\s*(.+)", t)
    if m:
        s = m.group(2).strip()
        return s[:24] if len(s) > 24 else s

    return first[:24] if first else "重点領域"


# =========================================================
# Step10 AP：主指標名 抽出＆“指標名として短文化”
# =========================================================
def extract_ap_raw_indicator(ap_text: str) -> str:
    """
    APセルから「主指標の候補」を拾う（箇条書き先頭を優先）
    """
    t = norm(ap_text)

    # 箇条書きの先頭（・）
    m = re.search(r"・\s*([^\n]+)", t)
    if m:
        return m.group(1).strip()

    # コロン以降
    m = re.search(r"[:：]\s*([^\n]+)", t)
    if m:
        return m.group(1).strip()

    # 先頭行
    return (t.split("\n")[0].strip() if t else "状態KPI")


def compact_indicator_name(raw: str) -> str:
    """
    “文章KPI” を “指標名” に圧縮する（ここが高橋モデルの肝）
    ※元の意味はAPに残るので、ARでは短い名称にする
    """
    s = raw.strip()

    # 典型パターン：案件割合→案件率
    s = s.replace("案件割合", "案件率")
    s = s.replace("割合", "率")

    # よくある冗長語を削る（意味は落とさない範囲）
    s = re.sub(r"(が|を|に|で|として|となっている|している|進んでいる|できている|している状態)", "", s)
    s = s.replace("、", "・")
    s = s.replace("／", "・")

    # 強めの圧縮ルール（ドメイン寄り）
    patterns: List[Tuple[str, str]] = [
        # “継続発注 + 追加拡大”系
        (r"継続発注.*追加拡大.*案件率", "継続拡大案件率"),
        (r"継続発注.*拡大.*案件率", "継続拡大案件率"),
        (r"継続.*拡大.*案件率", "継続拡大案件率"),
        # “継続定着”系
        (r"継続.*定着.*率", "継続定着率"),
        # “採用”系
        (r"採用.*継続.*率", "採用継続率"),
        (r"採用.*拡大.*率", "採用拡大率"),
        # “転換”系
        (r"転換.*継続.*率", "転換継続率"),
        (r"転換.*率", "転換率"),
        # “継続率”が含まれるなら単独でOK
        (r".*継続率.*", "継続率"),
        (r".*定着率.*", "定着率"),
    ]

    for pat, rep in patterns:
        if re.fullmatch(pat, s) or re.search(pat, s):
            return rep

    # 長すぎる場合の最終圧縮
    if len(s) > 18:
        # キーワードだけ残す簡易圧縮
        keywords = []
        for k in ["継続", "定着", "拡大", "転換", "採用", "発注", "処方", "案件", "率", "母数", "件数"]:
            if k in s and k not in keywords:
                keywords.append(k)
        if "率" not in keywords:
            keywords.append("率")
        return "".join(keywords)[:18]

    return s


# =========================================================
# Step10 AQ：②「まずここを見る」抽出（具体の起点）
# =========================================================
def extract_aq_first_check(aq_text: str) -> Optional[str]:
    t = norm(aq_text)
    if not t:
        return None

    lines = [ln.strip() for ln in t.split("\n") if ln.strip()]

    # ②行を最優先
    for ln in lines:
        if ln.startswith("②"):
            s = re.sub(r"^②\s*", "", ln).strip()
            if "：" in s:
                s = s.split("：", 1)[1].strip()
            return s

    # fallback：キーワード
    for ln in lines:
        if "まず" in ln or "最初" in ln or "一次切り分け" in ln:
            s = ln.replace("②", "").strip()
            if "：" in s:
                s = s.split("：", 1)[1].strip()
            return s

    return None


# =========================================================
# Step8 AA：具体（対象KPI/閾値/分子/分母/注意）抽出
# =========================================================
def extract_step8_concrete(aa_text: str) -> Dict[str, Optional[str]]:
    t = norm(aa_text)

    target = None
    thresh = None
    numerator = None
    denom = None
    note = None

    m = re.search(r"・対象KPI：\s*([^\n]+)", t)
    if m:
        target = m.group(1).strip()

    m = re.search(r"（\s*警戒<[^>]+>\s*／\s*未達<[^>]+>\s*）", t)
    if m:
        thresh = m.group(0).strip()

    m = re.search(r"・分子：\s*([^\n]+)", t)
    if m:
        numerator = m.group(1).strip()

    m = re.search(r"・分母：\s*([^\n]+)", t)
    if m:
        denom = m.group(1).strip()

    # “判定/集計/定義/ブレ”系の注意を拾う
    m = re.search(r"（[^）]*(判定|集計|定義|ブレ|カウント|漏れ)[^）]*）", t)
    if m:
        note = m.group(0).strip()

    return {
        "target": target,
        "thresh": thresh,
        "numerator": numerator,
        "denom": denom,
        "note": note,
    }


# =========================================================
# 行ごとの「構造タイプ」を決める（高橋モデルの分岐）
# =========================================================
def classify_structure_type(focus: str, indicator_raw: str, aq_first: Optional[str], step8: Dict[str, Optional[str]]) -> str:
    """
    upstream  : 上流（母数/機会/導入）滞留型
    midstream : 中流（転換/定着/継続）停滞型
    downstream: 下流（拡大/追加/伸長）阻害型
    """
    text = " ".join([
        focus or "",
        indicator_raw or "",
        aq_first or "",
        step8.get("numerator") or "",
        step8.get("denom") or "",
        step8.get("target") or "",
    ])

    # 下流
    if any(k in text for k in ["拡大", "追加", "増加", "伸長", "展開", "拡張"]):
        return "downstream"

    # 上流
    if any(k in text for k in ["母数", "件数", "接触", "提案", "導入", "機会", "新規", "採用開始", "処方開始"]):
        return "upstream"

    # 中流（デフォルト）
    return "midstream"


# =========================================================
# 構造宣言（タイプ×重点領域で“文の骨格”を変える）
# =========================================================
def build_structure_sentence(struct_type: str, focus: str) -> str:
    """
    「次フェーズは、◯◯構造を固定する。」の◯◯を、行ごとに変える
    """
    # 重点領域のニュアンスを追加
    focus_mod = ""
    if "競合対抗" in focus:
        focus_mod = "（競合に押し戻されない）"
    elif "拡大型採用" in focus:
        focus_mod = "（追加採用まで進む）"
    elif "継続処方" in focus:
        focus_mod = "（継続に落ちる）"
    elif "定着" in focus:
        focus_mod = "（定着が途切れない）"

    if struct_type == "upstream":
        return f"次フェーズは、「母数を滞留させない構造」{focus_mod}を固定する。"
    if struct_type == "downstream":
        return f"次フェーズは、「拡大が止まらない構造」{focus_mod}を固定する。"
    # midstream
    return f"次フェーズは、「転換・継続が止まらない構造」{focus_mod}を固定する."


# =========================================================
# 未達時ブロック：具体を拾って“読みやすい”番号付きに整形
# =========================================================
def build_unmet_block(aq_first: Optional[str], step8: Dict[str, Optional[str]]) -> List[str]:
    """
    ここで「抽象3分類」ではなく、AA/AQから“具体”を拾ってARに載せる
    """
    target = step8.get("target")
    thresh = step8.get("thresh")
    numerator = step8.get("numerator")
    denom = step8.get("denom")
    note = step8.get("note")

    lines: List[str] = []
    lines.append("未達時は、以下の順で切り分ける。")
    lines.append("")

    if aq_first:
        lines.append("① まず確認：")
        lines.append(f"   {aq_first}")
        lines.append("")

    if target:
        lines.append("② KPI判定（未達ラインの統一）：")
        if thresh:
            lines.append(f"   対象KPI「{target}」{thresh}")
        else:
            lines.append(f"   対象KPI「{target}」")
        lines.append("")

    if numerator:
        lines.append("③ 分子（成果側）確認：")
        lines.append(f"   {numerator}")
        lines.append("")

    if denom:
        lines.append("④ 分母（母集団側）確認：")
        lines.append(f"   {denom}")
        lines.append("")

    if note:
        lines.append("⑤ 判定・集計ブレの除去：")
        lines.append(f"   {note}")
        lines.append("")

    if len(lines) <= 2:
        # 具体が拾えなかった場合の保険
        lines = [
            "未達時は、KPIの定義（分子/分母）と未達判定を先に揃え、詰まり箇所を一次切り分けする。"
        ]

    return lines


# =========================================================
# AR生成（行ごとに構造が変わる：高橋モデル）
# =========================================================
def build_ar_text(focus: str, indicator_short: str, struct_sentence: str, unmet_lines: List[str]) -> str:
    lines: List[str] = []
    lines.append("【次フェーズ方針（一文化・宣言）】")
    lines.append("")
    lines.append(struct_sentence)
    lines.append(f"月次は「{indicator_short}」で状態を固定管理する。")
    lines.append("")
    lines.extend(unmet_lines)
    return "\n".join(lines)


# =========================================================
# Main
# =========================================================
def main():
    wb = load_workbook(INPUT_XLSX)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_NAME}")

    ws = wb[SHEET_NAME]

    for r in range(ROW_START, ROW_END + 1):
        v7 = ws[f"{COL_STEP7}{r}"].value
        v8 = ws[f"{COL_STEP8}{r}"].value
        vap = ws[f"{COL_AP}{r}"].value
        vaq = ws[f"{COL_AQ}{r}"].value

        if not (isinstance(v7, str) and v7.strip()):
            continue
        if not (isinstance(v8, str) and v8.strip()):
            continue
        if not (isinstance(vap, str) and vap.strip()):
            # APが無い行は、ARを触らない（構造維持）
            continue

        focus = extract_step7_focus(v7)

        indicator_raw = extract_ap_raw_indicator(vap)
        indicator_short = compact_indicator_name(indicator_raw)

        aq_first = extract_aq_first_check(vaq) if isinstance(vaq, str) and vaq.strip() else None
        step8 = extract_step8_concrete(v8)

        struct_type = classify_structure_type(focus, indicator_raw, aq_first, step8)
        struct_sentence = build_structure_sentence(struct_type, focus)

        unmet_lines = build_unmet_block(aq_first, step8)

        new_ar = build_ar_text(
            focus=focus,
            indicator_short=indicator_short,
            struct_sentence=struct_sentence,
            unmet_lines=unmet_lines,
        )

        cell = ws[f"{COL_AR}{r}"]
        cell.value = new_ar
        keep_wrap_top(cell)

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)


if __name__ == "__main__":
    main()