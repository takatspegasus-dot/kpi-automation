# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from copy import copy
from typing import Any, Optional

from openpyxl import load_workbook


# =========================================================
# 入出力（母艦）
# =========================================================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シートVer1.0.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シートVer1.0_再設計モデル.xlsx"

SRC_SHEET_NAME = "振り返りシート"
OUT_SHEET_NAME = "再設計モデル"


# =========================================================
# 列定義（既存：B〜K想定）
# =========================================================
COL = {
    "課": 2,        # B
    "区分": 3,      # C
    "期間": 4,      # D
    "KPI": 5,       # E
    "目標": 6,      # F
    "実績": 7,      # G
    "達成率": 8,    # H
    "採用軒数": 9,  # I
    "納入本数": 10, # J
    "全体比率": 11, # K
}


# =========================================================
# 表示形式（ここは好みに合わせて調整）
# =========================================================
FMT_PERCENT = "0.0%"
FMT_DECIMAL1 = "0.0"
FMT_INT = "0"


# =========================================================
# ユーティリティ
# =========================================================
def _to_number(v: Any) -> Optional[float]:
    """None/空文字はNone、'200%'等は0.02系に変換、カンマ除去"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s == "":
            return None
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _safe_div(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None or b == 0:
        return None
    return a / b


def find_data_start(ws) -> int:
    """B列に最初に課名が出る行を探す（'課'見出しは除外）"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, COL["課"]).value
        if v is None:
            continue
        s = str(v).strip()
        if s == "" or s == "課":
            continue
        return r
    raise RuntimeError("データ開始行が見つかりません（B列=課の値がありません）。")


def clone_style(dst_cell, src_cell) -> None:
    """セルの見た目（フォント/塗り/罫線/配置/表示形式/保護）をコピー"""
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell._style = copy(src_cell._style)


# =========================================================
# メイン処理
# =========================================================
def build_model() -> None:
    wb = load_workbook(INPUT_XLSX)

    if SRC_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"シート '{SRC_SHEET_NAME}' が見つかりません: {wb.sheetnames}")

    ws_src = wb[SRC_SHEET_NAME]

    # 既存の再設計モデルがある場合は削除
    if OUT_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[OUT_SHEET_NAME])

    # ★シート丸ごと複製：既存の罫線・色・列幅・行高などはそのまま継承
    ws_out = wb.copy_worksheet(ws_src)
    ws_out.title = OUT_SHEET_NAME

    # 追加列の開始位置（既存最終列の右）
    # ※「振り返りシート」の最終列（K）以降を増やす
    base = ws_out.max_column + 1

    # スタイルの“ひな型列”を決める：通常は既存の最終列（例：K列）のスタイルを横展開
    style_col = ws_out.max_column  # 複製直後の最終列（通常K）
    # 追加する列数
    add_cols = 4

    # 追加列の列幅も既存最終列に合わせる（課別の見た目を揃える）
    style_width = ws_out.column_dimensions[ws_out.cell(6, style_col).column_letter].width
    for i in range(add_cols):
        col_letter = ws_out.cell(6, base + i).column_letter
        ws_out.column_dimensions[col_letter].width = style_width

    # ヘッダを置く行（既存の体裁を壊しにくい行）
    header_row = 6

    # ① 追加列ヘッダ（表示名）
    headers = [
        "採用転換率(採用/実績)",
        "本数効率(本数/採用)",
        "活動効率(本数/実績)",
        "目標差分(実績-目標)",
    ]
    for i, title in enumerate(headers):
        dst = ws_out.cell(header_row, base + i)
        src = ws_out.cell(header_row, style_col)  # 既存最終列のヘッダの見た目をコピー
        clone_style(dst, src)
        dst.value = title

    # ② データ開始行
    data_start = find_data_start(ws_out)
    if data_start < 7:
        data_start = 7

    # ③ データ行：値を書き込み + 行ごとにスタイルをコピー（課別の色帯・罫線を維持）
    for r in range(data_start, ws_out.max_row + 1):
        dept = ws_out.cell(r, COL["課"]).value
        if dept is None or str(dept).strip() == "":
            break

        target = _to_number(ws_out.cell(r, COL["目標"]).value)
        actual = _to_number(ws_out.cell(r, COL["実績"]).value)
        adopt = _to_number(ws_out.cell(r, COL["採用軒数"]).value)
        units = _to_number(ws_out.cell(r, COL["納入本数"]).value)

        conv = _safe_div(adopt, actual)            # 採用転換率
        units_per_adopt = _safe_div(units, adopt)  # 本数効率
        units_per_act = _safe_div(units, actual)   # 活動効率
        diff = None if (actual is None or target is None) else (actual - target)

        values = [conv, units_per_adopt, units_per_act, diff]
        formats = [FMT_PERCENT, FMT_DECIMAL1, FMT_DECIMAL1, FMT_INT]

        # 行スタイルの“元”は既存最終列（style_col）の同じ行
        src_style_cell = ws_out.cell(r, style_col)

        for i in range(add_cols):
            dst = ws_out.cell(r, base + i)
            clone_style(dst, src_style_cell)        # 罫線/塗り/フォント等を課別帯のまま延長
            dst.number_format = formats[i]          # 表示形式だけ上書き
            dst.value = values[i]

    # 保存
    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print("OK:", OUTPUT_XLSX)


if __name__ == "__main__":
    build_model()