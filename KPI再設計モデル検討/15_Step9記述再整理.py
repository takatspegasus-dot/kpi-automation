# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from typing import Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# =========================================================
# 入出力（★ここだけ直せば動きます）
# =========================================================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step10追加05_out1.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step10追加05_out2.xlsx"

# Step9の記述対象（行）
ROW_START = 7
ROW_END = 73

# Step9の列（通常はAD）
COL_STEP9 = "AD"

# 参照列（現状フェーズ/最優先領域/根拠文章）
COL_PHASE = "AC"   # 例: 「発売前 / ～1月末｜滞留:初回導入」
COL_LEVER = "AE"   # 例: 「定着／継続処方／競合対抗／拡大型採用」
COL_DETAIL1 = "AF" # 例: 「対象KPI/停滞要因/打ち手 ...」
COL_DETAIL2 = "AG" # 例: 「狙う状態/意向変化 ...」


# =========================================================
# ユーティリティ
# =========================================================
def _s(v) -> str:
    return "" if v is None else str(v).strip()

def extract_after(label: str, text: str) -> str:
    """
    複数行テキストから「label：...」の ... 部分を抽出（次のラベル/行頭っぽい区切りまで）
    """
    if not text:
        return ""
    # label の後ろから、次の「xxx：」っぽい行までを取る
    pat = re.compile(rf"{re.escape(label)}\s*(.+?)(?=\n[^\n：]{1,15}：|\Z)", re.DOTALL)
    m = pat.search(text)
    if not m:
        return ""
    return m.group(1).strip().replace("\r\n", "\n")

def shorten_one_line(text: str, max_len: int = 60) -> str:
    t = " ".join(text.replace("\n", " ").split())
    if len(t) <= max_len:
        return t
    return t[: max_len - 1] + "…"

def build_step9_text(phase: str, lever: str, detail1: str, detail2: str) -> str:
    """
    課長会議で「発表できる」口調・構造に統一
    ①現状 ②到達 ③最優先（宣言）
    """
    phase = _s(phase)
    lever = _s(lever)
    d1 = _s(detail1)
    d2 = _s(detail2)

    # 現状：まず「意向変化」を優先（会議で説明しやすい）
    current = extract_after("意向変化：", d2)
    if not current:
        current = extract_after("停滞要因：", d1)
    if not current and phase:
        current = f"{phase}の段階で、狙いが十分に形になりきっていない状況です。"

    # 到達：狙う状態
    target = extract_after("狙う状態：", d2)
    if not target:
        # d2の先頭1文を採用（なければ空）
        first = d2.splitlines()[0].strip() if d2 else ""
        target = first

    # 文章を“発表向け”に整形（長すぎると読みにくいので軽く圧縮）
    current = shorten_one_line(current, 70) if current else "（現状整理中）"
    target = shorten_one_line(target, 70) if target else "（到達状態を明確化中）"
    lever_txt = lever if lever else "（未選択）"

    # 口調：断定しすぎず、宣言は強めに
    return (
        f"① 現状：{current}\n"
        f"② 到達：{target}\n"
        f"③ 最優先：{lever_txt}（ここをまず固めます）"
    )

def is_step9_sheet(ws) -> bool:
    """
    Step9列の見出しがあるシートだけ対象にする。
    （母艦の形式を崩さないためのガード）
    """
    v = _s(ws[f"{COL_STEP9}4"].value)  # 見出しは4行目にある前提
    return "Step9" in v or "最優先状態" in v


# =========================================================
# メイン
# =========================================================
def main() -> None:
    in_path = INPUT_XLSX
    out_path = OUTPUT_XLSX

    if not os.path.exists(in_path):
        raise FileNotFoundError(f"入力ファイルが見つかりません: {in_path}")

    wb = load_workbook(in_path)

    changed = 0
    for ws in wb.worksheets:
        if not is_step9_sheet(ws):
            continue

        for r in range(ROW_START, ROW_END + 1):
            # 行が空ならスキップ（B列=課などが空のケース）
            if ws[f"B{r}"].value is None and ws[f"C{r}"].value is None:
                continue

            phase = ws[f"{COL_PHASE}{r}"].value
            lever = ws[f"{COL_LEVER}{r}"].value
            d1 = ws[f"{COL_DETAIL1}{r}"].value
            d2 = ws[f"{COL_DETAIL2}{r}"].value

            new_text = build_step9_text(phase, lever, _s(d1), _s(d2))

            cell = ws[f"{COL_STEP9}{r}"]
            # 値更新
            if _s(cell.value) != _s(new_text):
                cell.value = new_text
                changed += 1

            # 見やすさ（折り返し・上詰め）を強制
            # 既存の揃え設定はなるべく維持しつつ、wrapとverticalだけ確実にする
            al = cell.alignment if cell.alignment else Alignment()
            cell.alignment = Alignment(
                horizontal=al.horizontal,
                vertical="top",
                text_rotation=al.text_rotation,
                wrap_text=True,
                shrink_to_fit=al.shrink_to_fit,
                indent=al.indent,
            )

    wb.save(out_path)
    print(f"保存しました: {out_path}")
    print(f"更新セル数（{COL_STEP9}{ROW_START}:{COL_STEP9}{ROW_END} 相当）: {changed}")


if __name__ == "__main__":
    main()