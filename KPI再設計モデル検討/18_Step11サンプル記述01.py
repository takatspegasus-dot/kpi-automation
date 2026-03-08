# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# =========================================================
# 入出力
# =========================================================
FILE_PATH = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シートPart2_02.xlsx"

SHEET_CANDIDATES = ["サンプル回答", "WS_サンプル回答", "WS_サンプル回答シート"]

# 対象行
ROW_START = 7
ROW_END = 73

# Step11（出力先）
OUT_COLS = {
    "structure": "AS",
    "change": "AT",
    "kpi": "AU",
    "kind": "AV",
    "denom": "AW",
    "connect": "AX",
}

# 参照列（Step10まで）
COL_STAGE = "P"   # Step3：停滞段階（認知／提案／検討／初回導入／継続）
COL_STATE = "V"   # Step6：6月末で進展させたい状態（状態で書く）
COL_LEVER = "W"   # Step7：重点領域（定着／継続処方／競合対抗／拡大型採用）
COL_NEWKPI = "X"  # Step8：新KPI
COL_STEP10_STATEKPI = "AP"  # Step10：月次で見る状態KPI（固定）


# =========================================================
# 行特性→KPIテンプレ（課長会議で話せる定番）
#  ※AT/AUの文面生成に使う（AS/AV/AWは従来推定）
# =========================================================
KPI_TEMPLATES = {
    "定着": {
        # ---- AT（測りたい変化：3ブロック）----
        "current_summary": "初回導入は実現しているが、継続が安定しているとはまだ言い切れない段階にある。",
        "target_state_fallback": "運用上の懸念が解消され、継続が合理的な選択肢として定着している状態。",
        "check_method": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいる。",

        # ---- AU（KPI候補：新KPI～定義まで）----
        "new_kpi_name": "継続発注が定着し、追加拡大検討が進んでいる案件割合",
        "reverse_kpi": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいることが確認できる案件割合",
        "progress_definition": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいる。",
        "numerator": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいることが確認できる案件数",
        "denominator": "当月の対象案件総数（例：重点案件／検討進行案件）",
        "unit": "割合（%）",
    },

    # 参考として他レバーも “同じ構造” で埋める（必要なら文言調整可能）
    "継続処方": {
        "current_summary": "初回導入は進みつつあるが、再処方・継続処方への移行が安定していない段階にある。",
        "target_state_fallback": "再処方が合理的に選択され、継続処方へ自然に移行している状態。",
        "check_method": "再処方が一定割合で発生し、継続処方へ移行する流れが安定している。",

        "new_kpi_name": "再処方が安定し、継続処方へ移行している医師割合",
        "reverse_kpi": "初回採用後に再処方が発生し、継続処方へ移行していることが確認できる医師割合",
        "progress_definition": "初回採用後に再処方が発生し、継続処方へ移行している。",
        "numerator": "初回採用後に再処方が発生し、継続処方へ移行していることが確認できる医師数",
        "denominator": "当月の初回採用医師総数（例：重点医師／フォロー対象医師）",
        "unit": "割合（%）",
    },

    "競合対抗": {
        "current_summary": "競合が選ばれている状況が残り、切替・奪還が安定して起きているとは言い切れない段階にある。",
        "target_state_fallback": "競合に対する優位性が整理され、切替・奪還が合理的に進む状態。",
        "check_method": "競合からの切替・奪還が発生し、逆戻りが抑えられている。",

        "new_kpi_name": "競合からの切替・奪還が進み、逆戻りが抑えられている医師割合",
        "reverse_kpi": "競合からの切替・奪還が確認でき、かつ逆戻りが抑えられている医師割合",
        "progress_definition": "競合からの切替・奪還が確認でき、逆戻りが抑えられている。",
        "numerator": "競合からの切替・奪還が確認でき、かつ逆戻りが抑えられている医師数",
        "denominator": "当月の競合処方医師総数（例：重点競合医師／対抗対象医師）",
        "unit": "割合（%）",
    },

    "拡大型採用": {
        "current_summary": "採用の芽はあるが、採用が面で広がり、次の再処方につながるところまで安定していない段階にある。",
        "target_state_fallback": "採用が面で増え、採用後の再処方へつながる流れが形成されている状態。",
        "check_method": "新規採用が増え、採用後の再処方が一定割合で確認できる。",

        "new_kpi_name": "新規採用が増え、採用後の再処方が確認できる医師割合",
        "reverse_kpi": "新規採用後に再処方が確認できる医師割合",
        "progress_definition": "新規採用が増え、採用後の再処方が確認できる。",
        "numerator": "新規採用後に再処方が確認できる医師数",
        "denominator": "当月の新規採用医師総数（例：重点ターゲット医師／新規提案医師）",
        "unit": "割合（%）",
    },
}

ALLOWED_LEVERS = set(KPI_TEMPLATES.keys())


# =========================================================
# ユーティリティ
# =========================================================
def pick_sheet_name(sheetnames: list[str]) -> str:
    for c in SHEET_CANDIDATES:
        if c in sheetnames:
            return c
    # 保険：部分一致
    for n in sheetnames:
        if "サンプル" in n and "回答" in n:
            return n
    raise ValueError(f"対象シートが見つかりません。現在のシート一覧: {sheetnames}")


def norm(s) -> str:
    return ("" if s is None else str(s)).replace("\u3000", " ").strip()


def normalize_placeholders(text: str) -> str:
    """
    課長会議で詰まりやすい表現を、可能な範囲で“固定値”へ。
    - ○日/◯日/○日以内 -> 30日以内
    - 直近○日 -> 直近30日
    """
    t = text or ""

    # ○日/◯日 -> 30日（以内/まで 等の語尾は残す）
    t = re.sub(r"[○◯]\s*日\s*以内", "30日以内", t)
    t = re.sub(r"[○◯]\s*日", "30日", t)
    t = re.sub(r"直近\s*[○◯]\s*日", "直近30日", t)

    # 余計な空白整理（改行は維持したいので行ごとに処理）
    lines = []
    for line in t.splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()
        lines.append(line)
    return "\n".join(lines).strip()


def infer_lever(v_lever: str, fallback_text: str) -> str:
    lv = norm(v_lever)
    if lv in ALLOWED_LEVERS:
        return lv
    t = fallback_text
    if any(k in t for k in ["競合", "奪還", "切替"]):
        return "競合対抗"
    if any(k in t for k in ["新規", "拡大", "採用施設", "面"]):
        return "拡大型採用"
    if "定着" in t:
        return "定着"
    return "継続処方"


def infer_kind_and_structure(kpi_text: str) -> tuple[str, str]:
    """
    KPI文言から「種類(状態/レバー/分母/監視)」と「構造起点」をざっくり推定
    ※AUが長文になるので、先頭の【新KPI】行だけで判定する
    """
    k = norm(kpi_text)
    # AUが長文の場合は、最初のKPI名だけ抽出して判定
    first = k.splitlines()
    # 「【新KPI】」の次行をKPI名として扱う
    if len(first) >= 2 and "【新KPI】" in first[0]:
        kpi_name = first[1].strip()
    else:
        kpi_name = k.strip()

    if any(w in kpi_name for w in ["実施率", "回収率", "提示率", "合意率", "接触率", "確認率", "取得率"]):
        return ("レバーKPI", "レバー")
    if any(w in kpi_name for w in ["中央値", "日数", "滞留", "未分類", "適合率", "測定可能率", "一致率", "数（課別）", "分散度"]):
        return ("監視KPI", "監視")
    if any(w in kpi_name for w in ["母集団", "捕捉率", "リスト整備率", "更新頻度", "特定率"]):
        return ("分母KPI", "分母")
    return ("状態KPI", "状態")


def infer_denom(lever: str, stage: str, kpi_text: str) -> str:
    """
    分母定義をざっくり推定（AWセル用：短文）
    ※AUが長文なので、分母はテンプレのdenominatorを短く言い換え
    """
    if lever == "定着":
        return "対象案件（重点/検討進行）"
    if lever == "競合対抗":
        return "競合処方医師"
    if lever == "拡大型採用":
        return "新規採用医師"
    return "初回採用医師"


# =========================================================
# AT/AU 生成（今回の修正の中核）
# =========================================================
def build_at_text(lever: str, stage: str, state_text: str) -> str:
    t = KPI_TEMPLATES[lever]
    current_summary = t["current_summary"]

    # 6月末到達状態：V列があれば優先（あなたの指示通り）
    target_state = norm(state_text) or t["target_state_fallback"]

    check_method = t["check_method"]

    at = (
        "【現状整理】\n"
        f"{current_summary}\n\n"
        "【6月末で到達させたい状態】\n"
        f"{target_state}\n\n"
        "【進展の確認方法】\n"
        f"{check_method}"
    )
    return normalize_placeholders(at)


def build_au_text(lever: str, state_text: str) -> str:
    t = KPI_TEMPLATES[lever]
    target_state = norm(state_text) or t["target_state_fallback"]

    au = (
        "【新KPI】\n"
        f"{t['new_kpi_name']}\n\n"
        f"到達状態：{target_state}\n\n"
        f"戦略レバー：{lever} 逆算KPI：\n"
        f"{t['reverse_kpi']}\n\n"
        "【どうなっていれば進んだと言えるか】\n"
        f"{t['progress_definition']}\n\n"
        "【KPI定義】\n"
        f"分子：{t['numerator']}\n"
        f"分母：{t['denominator']} 単位：{t['unit']}"
    )
    return normalize_placeholders(au)


# =========================================================
# メイン
# =========================================================
def main() -> None:
    wb = load_workbook(FILE_PATH)
    sheet_name = pick_sheet_name(wb.sheetnames)
    ws = wb[sheet_name]

    # 出力列index
    out_idx = {k: column_index_from_string(v) for k, v in OUT_COLS.items()}

    # 参照列index
    idx_stage = column_index_from_string(COL_STAGE)
    idx_state = column_index_from_string(COL_STATE)
    idx_lever = column_index_from_string(COL_LEVER)
    idx_newkpi = column_index_from_string(COL_NEWKPI)
    idx_step10 = column_index_from_string(COL_STEP10_STATEKPI)

    # Step11を一度クリア（値のみ）
    for r in range(ROW_START, ROW_END + 1):
        for k in OUT_COLS:
            ws.cell(row=r, column=out_idx[k]).value = None

    # 行別に作成
    for r in range(ROW_START, ROW_END + 1):
        stage = norm(ws.cell(row=r, column=idx_stage).value)
        state = norm(ws.cell(row=r, column=idx_state).value)
        lever_raw = norm(ws.cell(row=r, column=idx_lever).value)
        newkpi = norm(ws.cell(row=r, column=idx_newkpi).value)
        step10_statekpi = norm(ws.cell(row=r, column=idx_step10).value)

        # 「行特性」がほぼ無い（全部空）の場合は空欄のままにする
        if not any([stage, state, lever_raw, newkpi, step10_statekpi]):
            continue

        # 行文脈（推定用）
        context = " ".join([stage, state, lever_raw, newkpi, step10_statekpi])
        lever = infer_lever(lever_raw, context)
        if lever not in KPI_TEMPLATES:
            lever = "継続処方"

        # -------------------------------
        # ここが修正ポイント：
        # AT/AU を “あなたの指定フォーマット”で生成
        # -------------------------------
        change_text = build_at_text(lever=lever, stage=stage, state_text=state)
        kpi_candidate = build_au_text(lever=lever, state_text=state)

        # 種類（AV）と構造起点（AS）
        kind, structure = infer_kind_and_structure(kpi_candidate)

        # 分母（AW）
        denom = infer_denom(lever, stage, kpi_candidate)

        # 戦略接続（AX）
        connect = "○ 接続している" if lever else "△ 弱い"

        # 書き込み（値のみ）
        ws.cell(row=r, column=out_idx["structure"]).value = structure
        ws.cell(row=r, column=out_idx["change"]).value = change_text
        ws.cell(row=r, column=out_idx["kpi"]).value = kpi_candidate
        ws.cell(row=r, column=out_idx["kind"]).value = kind
        ws.cell(row=r, column=out_idx["denom"]).value = denom
        ws.cell(row=r, column=out_idx["connect"]).value = connect

    # 上書き保存ではなく、念のため別名で保存（母艦保護）
    base, ext = os.path.splitext(FILE_PATH)
    out_path = base + "_Step11行別反映" + ext
    wb.save(out_path)
    print(f"完了：Step11（AS7:AX73）を行特性に合わせて生成しました。\n保存先: {out_path}\n元シート: {sheet_name}")


if __name__ == "__main__":
    main()