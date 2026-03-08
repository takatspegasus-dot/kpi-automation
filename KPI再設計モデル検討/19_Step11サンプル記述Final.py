# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# =========================================================
# 入出力
# =========================================================
FILE_PATH = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シートPart2_02.xlsx"

SHEET_CANDIDATES = ["サンプル回答", "WS_サンプル回答", "WS_サンプル回答シート"]

ROW_START = 7
ROW_END = 73

OUT_COLS = {
    "structure": "AS",
    "change": "AT",
    "kpi": "AU",
    "kind": "AV",
    "denom": "AW",
    "connect": "AX",
}

COL_STAGE = "P"
COL_STATE = "V"
COL_LEVER = "W"
COL_NEWKPI = "X"
COL_STEP10_STATEKPI = "AP"


# =========================================================
# テンプレ（★定着の分母を案Aに統一）
# =========================================================
KPI_TEMPLATES = {
    "定着": {
        "target_state_fallback": "運用上の懸念が解消され、継続が合理的な選択肢として定着している状態。",
        "current_summary": "初回導入は実現しているが、継続が安定しているとはまだ言い切れない段階にある。",
        "check_method": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいる。",

        "new_kpi_name": "継続発注が定着し、追加拡大検討が進んでいる案件割合",
        "reverse_kpi": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいることが確認できる案件割合",
        "progress_definition": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいる。",
        "numerator": "発注が継続的に行われ、追加対象の拡大についても前向きな検討が進んでいることが確認できる案件数",

        # ★変更（AU側：分母文）
        "denominator": "当月のフォロー対象施設総数（重点施設リスト）",

        "unit": "割合（%）",

        # ★変更（AW側：短い分母定義）
        "denom_short": "当月のフォロー対象施設（重点施設）",
    },

    "継続処方": {
        "target_state_fallback": "再処方が合理的に選択され、継続処方へ自然に移行している状態。",
        "current_summary": "初回導入は進みつつあるが、再処方・継続処方への移行が安定していない段階にある。",
        "check_method": "再処方が一定割合で発生し、継続処方へ移行する流れが安定している。",

        "new_kpi_name": "再処方が安定し、継続処方へ移行している医師割合",
        "reverse_kpi": "初回採用後に再処方が発生し、継続処方へ移行していることが確認できる医師割合",
        "progress_definition": "初回採用後に再処方が発生し、継続処方へ移行している。",
        "numerator": "初回採用後に再処方が発生し、継続処方へ移行していることが確認できる医師数",
        "denominator": "当月の初回採用医師総数（例：重点医師／フォロー対象医師）",
        "unit": "割合（%）",
        "denom_short": "初回採用医師",
    },

    "競合対抗": {
        "target_state_fallback": "競合に対する優位性が整理され、切替・奪還が合理的に進む状態。",
        "current_summary": "競合が選ばれている状況が残り、切替・奪還が安定して起きているとは言い切れない段階にある。",
        "check_method": "競合からの切替・奪還が発生し、逆戻りが抑えられている。",

        "new_kpi_name": "競合からの切替・奪還が進み、逆戻りが抑えられている医師割合",
        "reverse_kpi": "競合からの切替・奪還が確認でき、かつ逆戻りが抑えられている医師割合",
        "progress_definition": "競合からの切替・奪還が確認でき、逆戻りが抑えられている。",
        "numerator": "競合からの切替・奪還が確認でき、かつ逆戻りが抑えられている医師数",
        "denominator": "当月の競合処方医師総数（例：重点競合医師／対抗対象医師）",
        "unit": "割合（%）",
        "denom_short": "競合処方医師",
    },

    "拡大型採用": {
        "target_state_fallback": "採用が面で増え、院内採用プロセスが滞りなく進み、採用が拡がっている状態。",
        "current_summary": "採用の芽はあるが、採用が面で広がり、院内で採用プロセスが安定して進んでいるとは言い切れない段階にある。",
        "check_method": "未採用施設で採用申請・提案が進み、採用決裁・運用に向けた検討が前進している。",

        "new_kpi_name": "採用プロセスが前進し、採用決裁・運用検討が進んでいる施設割合",
        "reverse_kpi": "採用申請・提案が進み、採用決裁・運用検討が前進していることが確認できる施設割合",
        "progress_definition": "採用申請・提案が進み、採用決裁・運用検討が前進している。",
        "numerator": "採用申請・提案が進み、採用決裁・運用検討が前進していることが確認できる施設数",
        "denominator": "当月の対象施設総数（例：未採用施設／重点施設）",
        "unit": "割合（%）",
        "denom_short": "対象施設（未採用/重点）",
    },
}

ALLOWED_LEVERS = set(KPI_TEMPLATES.keys())


# =========================================================
# Utility
# =========================================================
def pick_sheet_name(sheetnames: list[str]) -> str:
    for c in SHEET_CANDIDATES:
        if c in sheetnames:
            return c
    for n in sheetnames:
        if "サンプル" in n and "回答" in n:
            return n
    raise ValueError(f"対象シートが見つかりません。現在のシート一覧: {sheetnames}")


def norm(s) -> str:
    return ("" if s is None else str(s)).replace("\u3000", " ").strip()


def normalize_placeholders(text: str) -> str:
    t = text or ""
    t = re.sub(r"[○◯]\s*日\s*以内", "30日以内", t)
    t = re.sub(r"[○◯]\s*日", "30日", t)
    t = re.sub(r"直近\s*[○◯]\s*日", "直近30日", t)

    lines = []
    for line in t.splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()
        lines.append(line)
    return "\n".join(lines).strip()


def normalize_lever(raw: str) -> str:
    s = norm(raw)
    if not s:
        return ""

    s = re.sub(r"（.*?）", "", s)
    s = re.sub(r"\(.*?\)", "", s)
    s = s.strip()

    if "競合" in s:
        return "競合対抗"
    if "拡大" in s or "新規" in s:
        return "拡大型採用"
    if "定着" in s:
        return "定着"
    if "継続" in s:
        return "継続処方"
    if s in ALLOWED_LEVERS:
        return s
    return ""


def infer_lever_from_context(context: str) -> str:
    t = context
    if any(k in t for k in ["発注", "案件", "追加対象", "拡大検討"]):
        return "定着"
    if any(k in t for k in ["採用申請", "未採用施設", "院内", "薬剤部", "施設"]):
        return "拡大型採用"
    if any(k in t for k in ["競合", "奪還", "切替"]):
        return "競合対抗"
    return "継続処方"


def infer_kind_and_structure(kpi_text: str) -> tuple[str, str]:
    k = norm(kpi_text)
    lines = k.splitlines()
    kpi_name = lines[1].strip() if len(lines) >= 2 and "【新KPI】" in lines[0] else k.strip()

    if any(w in kpi_name for w in ["実施率", "回収率", "提示率", "合意率", "接触率", "確認率", "取得率"]):
        return ("レバーKPI", "レバー")
    if any(w in kpi_name for w in ["中央値", "日数", "滞留", "未分類", "適合率", "測定可能率", "一致率", "数（課別）", "分散度"]):
        return ("監視KPI", "監視")
    if any(w in kpi_name for w in ["母集団", "捕捉率", "リスト整備率", "更新頻度", "特定率"]):
        return ("分母KPI", "分母")
    return ("状態KPI", "状態")


def build_at_text(lever: str, state_text: str) -> str:
    t = KPI_TEMPLATES[lever]
    target_state = norm(state_text) or t["target_state_fallback"]
    at = (
        "【現状整理】\n"
        f"{t['current_summary']}\n\n"
        "【6月末で到達させたい状態】\n"
        f"{target_state}\n\n"
        "【進展の確認方法】\n"
        f"{t['check_method']}"
    )
    return normalize_placeholders(at)


def build_au_text(lever: str, state_text: str) -> str:
    t = KPI_TEMPLATES[lever]
    target_state = norm(state_text) or t["target_state_fallback"]
    au = (
        "【新KPI】\n"
        f"{t['new_kpi_name']}\n\n"
        f"到達状態：{target_state}\n\n"
        f"戦略レバー：{lever} 逆算KPI：\n"
        f"{t['reverse_kpi']}\n\n"
        "【どうなっていれば進んだと言えるか】\n"
        f"{t['progress_definition']}\n\n"
        "【KPI定義】\n"
        f"分子：{t['numerator']}\n"
        f"分母：{t['denominator']} 単位：{t['unit']}"
    )
    return normalize_placeholders(au)


# =========================================================
# Main
# =========================================================
def main() -> None:
    wb = load_workbook(FILE_PATH)
    sheet_name = pick_sheet_name(wb.sheetnames)
    ws = wb[sheet_name]

    out_idx = {k: column_index_from_string(v) for k, v in OUT_COLS.items()}

    idx_stage = column_index_from_string(COL_STAGE)
    idx_state = column_index_from_string(COL_STATE)
    idx_lever = column_index_from_string(COL_LEVER)
    idx_newkpi = column_index_from_string(COL_NEWKPI)
    idx_step10 = column_index_from_string(COL_STEP10_STATEKPI)

    # Step11クリア（値のみ）
    for r in range(ROW_START, ROW_END + 1):
        for k in OUT_COLS:
            ws.cell(row=r, column=out_idx[k]).value = None

    lever_count = {k: 0 for k in ALLOWED_LEVERS}

    for r in range(ROW_START, ROW_END + 1):
        stage = norm(ws.cell(row=r, column=idx_stage).value)
        state = norm(ws.cell(row=r, column=idx_state).value)
        lever_raw = norm(ws.cell(row=r, column=idx_lever).value)
        newkpi = norm(ws.cell(row=r, column=idx_newkpi).value)
        step10_statekpi = norm(ws.cell(row=r, column=idx_step10).value)

        if not any([stage, state, lever_raw, newkpi, step10_statekpi]):
            continue

        lever = normalize_lever(lever_raw)
        if not lever:
            context = " ".join([stage, state, newkpi, step10_statekpi])
            lever = infer_lever_from_context(context)

        if lever not in KPI_TEMPLATES:
            lever = "継続処方"

        lever_count[lever] = lever_count.get(lever, 0) + 1

        change_text = build_at_text(lever=lever, state_text=state)
        kpi_candidate = build_au_text(lever=lever, state_text=state)

        kind, structure = infer_kind_and_structure(kpi_candidate)

        # ★AWは案Aに統一された denom_short が入る
        denom_short = KPI_TEMPLATES[lever]["denom_short"]
        connect = "○ 接続している"

        ws.cell(row=r, column=out_idx["structure"]).value = structure
        ws.cell(row=r, column=out_idx["change"]).value = change_text
        ws.cell(row=r, column=out_idx["kpi"]).value = kpi_candidate
        ws.cell(row=r, column=out_idx["kind"]).value = kind
        ws.cell(row=r, column=out_idx["denom"]).value = denom_short
        ws.cell(row=r, column=out_idx["connect"]).value = connect

    base, ext = os.path.splitext(FILE_PATH)
    out_path = base + "_Step11行別反映" + ext
    wb.save(out_path)

    print("完了：Step11（AS7:AX73）を生成しました（定着の分母は案A）。")
    print(f"保存先: {out_path}")
    print("レバー判定数:", lever_count)


if __name__ == "__main__":
    main()