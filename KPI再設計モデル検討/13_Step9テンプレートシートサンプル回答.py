# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, Dict

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# =========================================================
# 入出力（★ユーザー指定の母艦）
# =========================================================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step9_Ver1.0_AO追加01.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シート_Step9_Ver1.0_AO追加01_Step9サンプル記述_v4.xlsx"

SHEET_NAME = "WS_サンプル回答"
ROW_START = 7
ROW_END = 73

# 値だけ更新（書式は触らない）
OUT_COLS = ["AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO"]


# =========================================================
# ユーティリティ
# =========================================================
def s(v) -> str:
    if v is None:
        return ""
    t = str(v).replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"\n{3,}", "\n\n", t).strip()
    return t


def first_line(text: str) -> str:
    t = s(text)
    return t.split("\n", 1)[0].strip() if t else ""


def safe_int(v) -> Optional[int]:
    try:
        if v is None or isinstance(v, bool):
            return None
        return int(v)
    except Exception:
        return None


def pick_after(text: str, marker: str) -> str:
    """
    marker 行以降〜次の「【...】」が来るまでを抽出
    """
    t = s(text)
    if not t:
        return ""
    idx = t.find(marker)
    if idx < 0:
        return ""
    rest = t[idx + len(marker):].lstrip()
    m = re.search(r"\n【[^】]+】", rest)
    if m:
        rest = rest[: m.start()]
    return rest.strip()


def pick_section(text: str, header: str) -> str:
    return pick_after(text, header)


def summarize_one_sentence(text: str, fallback: str = "") -> str:
    """
    ざっくり1文に寄せる（句点 or 改行で切る）
    """
    t = s(text)
    if not t:
        return fallback
    t = re.sub(r"^[・\-–—\u2022]+\s*", "", t)
    if "。" in t:
        return t.split("。", 1)[0].strip() + "。"
    if "\n" in t:
        return t.split("\n", 1)[0].strip()
    return t.strip()


def normalize_choice(text: str, choices: list[str]) -> str:
    t = s(text)
    for c in choices:
        if c and c in t:
            return c
    return first_line(t)


def safe_set(ws, addr: str, value: str) -> None:
    """
    結合セルの左上以外(MergedCell)は書き込み不可なのでスキップ。
    書式は触らず値だけセット。
    """
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def extract_new_kpi_name(x_text: str) -> str:
    """
    X列（Step8）から新KPI名を抽出
    """
    t = s(x_text)
    body = pick_section(t, "【新KPI】")
    if body:
        return first_line(body)
    # フォールバック：最初の非空行
    for line in t.split("\n"):
        line = line.strip()
        if line and "【" not in line:
            return line
    return ""


def extract_target_short(z_text: str) -> str:
    """
    Z列（目標値）から短い目標（例：6月末：60%）を抽出
    """
    t = s(z_text)
    body = pick_section(t, "【目標値（結論）】")
    if body:
        return first_line(body)
    for line in t.split("\n"):
        line = line.strip()
        if "6月末" in line and (("%" in line) or ("件" in line) or ("本" in line)):
            return line
    return first_line(t)


# =========================================================
# Step9 生成（読みやすい改行＋AN/AOを強化）
# =========================================================
def build_step9(ws, r: int) -> Dict[str, str]:
    # 参照（Step1〜Step8相当）
    c_phase = s(ws[f"C{r}"].value)
    d_term = s(ws[f"D{r}"].value)
    kpi_name = s(ws[f"E{r}"].value)

    f_target = s(ws[f"F{r}"].value)
    g_actual = s(ws[f"G{r}"].value)
    i_adopt = s(ws[f"I{r}"].value)
    j_units = s(ws[f"J{r}"].value)

    p_stuck = s(ws[f"P{r}"].value)
    q_fact = s(ws[f"Q{r}"].value)
    r_struct = s(ws[f"R{r}"].value)
    s_state = s(ws[f"S{r}"].value)
    t_judge = s(ws[f"T{r}"].value)
    u_logic = s(ws[f"U{r}"].value)

    v_priority = s(ws[f"V{r}"].value)
    w_lever = s(ws[f"W{r}"].value)

    x_kpi = s(ws[f"X{r}"].value)
    y_check = s(ws[f"Y{r}"].value)
    z_target = s(ws[f"Z{r}"].value)
    ab_pre = s(ws[f"AB{r}"].value)

    # AC：現在フェーズ
    cur_phase = " / ".join([p for p in [c_phase, d_term] if p])
    if p_stuck:
        cur_phase = f"{cur_phase}｜滞留:{p_stuck}" if cur_phase else f"滞留:{p_stuck}"

    # AD：最優先状態
    desired = pick_section(v_priority, "【6月末で到達させたい状態】")
    desired = summarize_one_sentence(desired, fallback=summarize_one_sentence(v_priority))
    if not desired:
        desired = "（6月末到達状態未記入）"

    # AE：主軸レバー
    lever = normalize_choice(w_lever, ["定着", "継続処方", "競合対抗", "拡大型採用"])
    if not lever:
        lever = "（レバー未選択）"

    # AF：打ち手（KPI名＋停滞要因＋次に動かすポイント）
    stop = pick_section(q_fact, "【停滞要因（具体）】")
    next_point = pick_section(q_fact, "【次に動かすポイント】")
    af_lines = [f"対象KPI：{first_line(kpi_name) or '（KPI未記入）'}"]
    if stop:
        af_lines.append(f"停滞要因：{summarize_one_sentence(stop)}")
    af_lines.append(f"打ち手：{summarize_one_sentence(next_point, fallback=summarize_one_sentence(q_fact, fallback='（打ち手未記入）'))}")
    af = "\n".join(af_lines)

    # AG：状態変化
    intent = pick_section(r_struct, "【採用・継続意向の変化】")
    ag_lines = [f"狙う状態：{desired}"]
    if intent:
        ag_lines.append(f"意向変化：{summarize_one_sentence(intent)}")
    ag = "\n".join(ag_lines)

    # AH：意思決定変化（判断軸＋競合）
    axis = pick_section(r_struct, "【判断軸の変化】")
    comp = pick_section(r_struct, "【競合状況の変化】")
    ah_lines = [f"判断軸：{summarize_one_sentence(axis, fallback=summarize_one_sentence(r_struct, fallback='（判断軸未記入）'))}"]
    if comp:
        ah_lines.append(f"競合：{summarize_one_sentence(comp)}")
    ah = "\n".join(ah_lines)

    # AI：成果（KPI名＋目標/実績/採用/納入）
    nums = []
    if f_target:
        nums.append(f"目標{f_target}")
    if g_actual:
        nums.append(f"実績{g_actual}")
    if i_adopt:
        nums.append(f"採用{i_adopt}")
    if j_units:
        nums.append(f"納入{j_units}")
    ai = "KPI：" + (first_line(kpi_name) or "（KPI未記入）") + "\n" + (" / ".join(nums) if nums else "（数値未入力）")

    # AJ：成立前提（2点）
    pre = pick_section(ab_pre, "【前提・留意点】") or ab_pre
    bullets = [b.strip() for b in re.split(r"\n[・\-–—\u2022]", "\n" + s(pre)) if b.strip()]
    if bullets:
        aj = "成立前提：\n・" + bullets[0]
        if len(bullets) > 1:
            aj += "\n・" + bullets[1]
    else:
        aj = "成立前提：（未記入）"

    # AK：滞留段階
    ak = p_stuck or "（滞留段階未選択）"

    # AL：ズレ（状態判定＋想定因果＋現場状態）
    mark = first_line(s_state)
    planned = pick_section(u_logic, "【因果（N列）】")
    state_line = pick_section(u_logic, "【状態（S列）】")
    al_lines = []
    if mark:
        al_lines.append(f"状態判定：{mark}")
    if planned:
        al_lines.append(f"想定因果：{summarize_one_sentence(planned)}")
    if state_line:
        al_lines.append(f"現場状態：{summarize_one_sentence(state_line)}")
    al = "\n".join(al_lines) if al_lines else summarize_one_sentence(s_state, fallback="（ズレ未記入）")

    # AM：仮説修正要否
    am = normalize_choice(t_judge, ["継続", "補強", "修正", "再構築"])
    if not am:
        am = "（要否未選択）"

    # AN：各課接続示唆（KPI設計への影響） ←★ここを強化
    new_kpi_name = extract_new_kpi_name(x_kpi) or "（新KPI未記入）"
    tgt_short = extract_target_short(z_target)
    check_body = pick_section(y_check, "【確認方法】") or y_check
    check_short = summarize_one_sentence(check_body, fallback="（確認方法未記入）")
    why_short = summarize_one_sentence(t_judge, fallback="")
    an_lines = [
        f"主軸レバー：{lever}",
        f"新KPI案：{new_kpi_name}" + (f"（{tgt_short}）" if tgt_short else ""),
        f"確認：{check_short}",
        f"設計への影響：{am}" + (f"（{why_short}）" if why_short else ""),
    ]
    an = "\n".join(an_lines)

    # AO：戦略一文化（自然な日本語・3行） ←★ここを改善
    phase_short = " / ".join([p for p in [c_phase, d_term] if p]) or "（フェーズ未記入）"
    ao = (
        f"現フェーズ：{phase_short}（滞留：{ak}）\n"
        f"6月末に目指す状態：{desired}\n"
        f"主軸は「{lever}」。測定は「{new_kpi_name}」"
        + (f"（{tgt_short}）" if tgt_short else "")
        + "で揃える。"
    )

    return {
        "AC": cur_phase,
        "AD": desired,
        "AE": lever,
        "AF": af,
        "AG": ag,
        "AH": ah,
        "AI": ai,
        "AJ": aj,
        "AK": ak,
        "AL": al,
        "AM": am,
        "AN": an,
        "AO": ao,
    }


def main() -> None:
    in_path = Path(INPUT_XLSX)
    out_path = Path(OUTPUT_XLSX)

    if not in_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {in_path}")

    wb = load_workbook(in_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SHEET_NAME} / sheets={wb.sheetnames}")

    ws = wb[SHEET_NAME]

    for r in range(ROW_START, ROW_END + 1):
        vals = build_step9(ws, r)
        for col in OUT_COLS:
            safe_set(ws, f"{col}{r}", vals.get(col, ""))

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()