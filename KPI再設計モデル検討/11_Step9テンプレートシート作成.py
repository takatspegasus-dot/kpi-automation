# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


# =========================================================
# 入出力
# =========================================================
INPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シートVer1.0.xlsx"
OUTPUT_XLSX = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シートVer1.0_Step9完成版_課枠実線_切替線実線.xlsx"

SHEET_NAME = "WS_サンプル回答"

# Step9はAC列から（AC〜AN）
STEP9_COL_START = 29  # AC
STEP9_COL_END = 40    # AN

# 行構造
TITLE_ROW = 4
HEADER_ROW_START = 5
HEADER_ROW_END = 6
DATA_START_ROW = 7
DATA_END_ROW = 73

# テーブル範囲（既存WSの表全体）
TABLE_COL_START = 2   # B
TABLE_COL_END = 40    # AN
TABLE_ROW_START = 4
TABLE_ROW_END = 73

# 課名：B列、フェーズ判定：C列
DEPT_COL = 2  # B
PHASE_COL = 3  # C

# 書式コピー元（Step9用）
TEMPLATE_COL = 29  # AC


# =========================================================
# 共通関数
# =========================================================
def copy_cell_style(src, dst) -> None:
    """セルのスタイルをコピー（値はコピーしない）"""
    dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def unmerge_overlapping(ws, min_row, max_row, min_col, max_col):
    """指定範囲と重なる結合セルを解除（MergedCell回避）"""
    targets = []
    for rng in list(ws.merged_cells.ranges):
        c1, r1, c2, r2 = rng.bounds
        if not (c2 < min_col or c1 > max_col or r2 < min_row or r1 > max_row):
            targets.append(str(rng))
    for a in targets:
        ws.unmerge_cells(a)


def add_list_validation(ws, addr_range: str, items: list[str]):
    """リスト型データ検証（プルダウン）を追加"""
    formula = '"' + ",".join(items) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv)
    dv.add(addr_range)


def set_border_side(cell, *, left=None, right=None, top=None, bottom=None):
    """既存Borderを壊さず、指定辺だけ差し替える"""
    b: Border = cell.border or Border()
    cell.border = Border(
        left=left if left is not None else b.left,
        right=right if right is not None else b.right,
        top=top if top is not None else b.top,
        bottom=bottom if bottom is not None else b.bottom,
        diagonal=b.diagonal,
        diagonal_direction=b.diagonal_direction,
        outline=b.outline,
        vertical=b.vertical,
        horizontal=b.horizontal,
    )


# =========================================================
# 1) 全体テーブル：外枠実線／内部点線（ベースを整える）
# =========================================================
def apply_table_borders(ws):
    """
    B4:AN73 をベース整形
      - 外枠：thin（実線）
      - 内部：dotted（点線）
      - 4行目下：thin（実線）
      - 6行目下：thin（実線）
    """
    dotted = Side(style="dotted")
    solid = Side(style="thin")

    for r in range(TABLE_ROW_START, TABLE_ROW_END + 1):
        for c in range(TABLE_COL_START, TABLE_COL_END + 1):
            left = dotted
            right = dotted
            top = dotted
            bottom = dotted

            # テーブル外枠（実線）
            if r == TABLE_ROW_START:
                top = solid
            if r == TABLE_ROW_END:
                bottom = solid
            if c == TABLE_COL_START:
                left = solid
            if c == TABLE_COL_END:
                right = solid

            # 4行目・6行目区切り（実線）
            if r == 4:
                bottom = solid
            if r == 6:
                bottom = solid

            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)


# =========================================================
# 2) 課ブロック検出（B列の連続値で区切る）
# =========================================================
def detect_dept_blocks(ws):
    """
    B列（課名）が連続している範囲を課ブロックとして抽出
    戻り値: [(dept_name, start_row, end_row), ...]
    """
    blocks = []
    current = None
    start = None

    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        v = ws.cell(r, DEPT_COL).value
        if current is None:
            current = v
            start = r
            continue
        if v != current:
            blocks.append((current, start, r - 1))
            current = v
            start = r

    blocks.append((current, start, DATA_END_ROW))
    return blocks


# =========================================================
# 3) 課ブロック外枠を実線に（B〜AN × start〜end）
# =========================================================
def apply_dept_outer_solid(ws):
    solid = Side(style="thin")
    blocks = detect_dept_blocks(ws)

    for (_dept, r1, r2) in blocks:
        c1 = TABLE_COL_START
        c2 = TABLE_COL_END

        # 上辺
        for c in range(c1, c2 + 1):
            set_border_side(ws.cell(r1, c), top=solid)
        # 下辺
        for c in range(c1, c2 + 1):
            set_border_side(ws.cell(r2, c), bottom=solid)
        # 左辺
        for r in range(r1, r2 + 1):
            set_border_side(ws.cell(r, c1), left=solid)
        # 右辺
        for r in range(r1, r2 + 1):
            set_border_side(ws.cell(r, c2), right=solid)


# =========================================================
# 4) 課別：発売前→発売後 切替線を実線に（C列判定）
# =========================================================
def apply_phase_separator_per_dept(ws):
    """
    各課ブロック内で、C列に「発売後」が初めて出る行を探し、
    その行の上罫線を B〜AN で実線化する
    """
    solid = Side(style="thin")
    blocks = detect_dept_blocks(ws)

    for (_dept, r1, r2) in blocks:
        release_row = None
        for r in range(r1, r2 + 1):
            v = ws.cell(r, PHASE_COL).value
            if isinstance(v, str) and "発売後" in v:
                release_row = r
                break

        if release_row is None:
            continue  # その課に発売後がない場合

        # 「発売後先頭行」の上線を実線に
        for c in range(TABLE_COL_START, TABLE_COL_END + 1):
            set_border_side(ws.cell(release_row, c), top=solid)


# =========================================================
# Step9 作成（AC〜AN）
# =========================================================
def build_step9(ws):
    # Step9領域（AC〜AN, 4〜73）にかぶる結合解除
    unmerge_overlapping(ws, TITLE_ROW, DATA_END_ROW, STEP9_COL_START, STEP9_COL_END)

    # 1) Step9タイトル（4行目：AC4:AN4 横結合）
    ws.merge_cells(start_row=TITLE_ROW, start_column=STEP9_COL_START,
                   end_row=TITLE_ROW, end_column=STEP9_COL_END)
    title_cell = ws.cell(TITLE_ROW, STEP9_COL_START)
    copy_cell_style(ws.cell(TITLE_ROW, TEMPLATE_COL), title_cell)
    title_cell.value = "Step9｜戦略接続モデル"
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 2) 各列タイトル（5〜6行結合：AC5:AC6 ... AN5:AN6）
    headers = [
        "現在フェーズ\n（今は何フェーズか）",                               # AC
        "最優先状態\n（6月末に到達すべき状態）",                             # AD
        "主軸レバー\n【選択】定着／継続処方／競合対抗／拡大型採用",            # AE
        "打ち手\n（戦略施策）",                                             # AF
        "状態変化\n（市場・顧客の状態）",                                   # AG
        "意思決定変化\n（決裁・判断の変化）",                               # AH
        "成果\n（採用・継続・シェアなど）",                                 # AI
        "成立前提\n（因果が成立する条件）",                                 # AJ
        "実際の滞留段階\n【選択】再設計モデル参照",                           # AK
        "戦略とのズレ\n（因果の乖離）",                                     # AL
        "仮説修正要否\n【選択】継続／補強／修正／再構築",                      # AM
        "各課接続示唆\n（KPI設計への影響）",                                 # AN
    ]

    for i, text in enumerate(headers):
        col = STEP9_COL_START + i
        ws.merge_cells(start_row=HEADER_ROW_START, start_column=col,
                       end_row=HEADER_ROW_END, end_column=col)
        cell = ws.cell(HEADER_ROW_START, col)
        copy_cell_style(ws.cell(HEADER_ROW_START, TEMPLATE_COL), cell)
        cell.value = text
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 3) 入力エリア（7〜73）：AC列同一行のスタイルを横展開
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        base = ws.cell(r, TEMPLATE_COL)
        for c in range(STEP9_COL_START, STEP9_COL_END + 1):
            dst = ws.cell(r, c)
            copy_cell_style(base, dst)
            dst.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 4) プルダウン
    add_list_validation(ws, f"AE{DATA_START_ROW}:AE{DATA_END_ROW}",
                        ["定着", "継続処方", "競合対抗", "拡大型採用"])
    add_list_validation(ws, f"AK{DATA_START_ROW}:AK{DATA_END_ROW}",
                        ["認知", "比較対象入り", "採用検討", "初回導入", "継続処方", "拡大採用"])
    add_list_validation(ws, f"AM{DATA_START_ROW}:AM{DATA_END_ROW}",
                        ["継続", "補強", "修正", "再構築"])


# =========================================================
# メイン
# =========================================================
def main():
    wb = load_workbook(INPUT_XLSX)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Worksheet '{SHEET_NAME}' does not exist. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # Step9生成
    build_step9(ws)

    # ① 全体ベース（外枠実線・内部点線）
    apply_table_borders(ws)

    # ② 課ブロック外枠を「課単位」で実線に（ここが今回の要件）
    apply_dept_outer_solid(ws)

    # ③ 課別：発売前→発売後 切替線を実線に（C列判定）
    apply_phase_separator_per_dept(ws)

    wb.save(OUTPUT_XLSX)
    print("出力完了:", OUTPUT_XLSX)


if __name__ == "__main__":
    main()