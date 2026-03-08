# -*- coding: utf-8 -*-
from __future__ import annotations

from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# =========================================================
# 入出力
# =========================================================
FILE_PATH = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\KPI再設計モデル検討\KPI管理シートPart2_01.xlsx"
SHEET_NAME = "WS_サンプル回答"  # ★実シート名（あなたのエラー表示に出ている名称）


# =========================================================
# Step11 仕様
# =========================================================
# 追加開始：AS列〜AX列（6列）
STEP11_START_COL = "AS"
STEP11_END_COL = "AX"

TITLE_ROW = 4
HEADER_TOP_ROW = 5
HEADER_BOTTOM_ROW = 6
DATA_ROW_START = 7
DATA_ROW_END = 73

TITLE_TEXT = "Step11｜KPI候補抽出\n戦略接続モデルから自然に導出されるKPIを構造別に整理する"

HEADERS = [
    ("AS", "構造起点"),
    ("AT", "測りたい変化"),
    ("AU", "KPI候補"),
    ("AV", "種類"),
    ("AW", "分母定義"),
    ("AX", "戦略接続○×"),
]

# ★Step10（AP〜AR）を「見た目テンプレ」として使う
# Step10は 3列(AP,AQ,AR)なので、Step11 6列は以下の対応で“型”をコピーする
# - 先頭列の境界(左線なし等)を引き継ぐため、ASはAPのスタイルを採用
# - 以降はAQ/ARのスタイルを使い回す
STYLE_TEMPLATE_MAP = {
    "AS": "AP",
    "AT": "AQ",
    "AU": "AR",
    "AV": "AQ",
    "AW": "AQ",
    "AX": "AQ",
}


def unmerge_overlaps(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """指定範囲と重なる結合セルを解除（結合衝突回避）"""
    overlaps = []
    for rng in list(ws.merged_cells.ranges):
        rmin, cmin, rmax, cmax = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        if not (rmax < min_row or rmin > max_row or cmax < min_col or cmin > max_col):
            overlaps.append(str(rng))
    for r in overlaps:
        ws.unmerge_cells(r)


def copy_cell_style(src, dst) -> None:
    """セルの見た目（スタイル類）を丸ごとコピー"""
    dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.comment = None  # コメントは不要ならコピーしない（必要なら変更してください）


def main() -> None:
    wb = load_workbook(FILE_PATH)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シート '{SHEET_NAME}' が見つかりません。現在のシート一覧: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    s_col = column_index_from_string(STEP11_START_COL)  # AS
    e_col = column_index_from_string(STEP11_END_COL)    # AX

    # ---------------------------------------------------------
    # 0) Step11範囲(AS〜AX, 4〜6行)に既存結合があれば解除
    # ---------------------------------------------------------
    unmerge_overlaps(ws, TITLE_ROW, HEADER_BOTTOM_ROW, s_col, e_col)

    # ---------------------------------------------------------
    # 1) Step10の「型」をStep11領域にコピー
    #    - 列幅
    #    - 行高（4〜6行）
    #    - セルスタイル（4〜73行）
    # ---------------------------------------------------------
    # 1-1) 列幅コピー（AP/AQ/AR -> AS〜AX）
    for tgt_col_letter, src_col_letter in STYLE_TEMPLATE_MAP.items():
        ws.column_dimensions[tgt_col_letter].width = ws.column_dimensions[src_col_letter].width

    # 1-2) 行高コピー（4〜6行はStep10と同じ「見出し帯」なのでそのまま）
    # ※ 参照元の行高は “行番号”でコピーします（Step10の行帯と同じ行なので、そのまま自行を維持してもOK）
    for r in range(TITLE_ROW, HEADER_BOTTOM_ROW + 1):
        if ws.row_dimensions[r].height is not None:
            ws.row_dimensions[r].height = ws.row_dimensions[r].height

    # 1-3) セルスタイルコピー（4〜73行）
    for r in range(TITLE_ROW, DATA_ROW_END + 1):
        for tgt_col_letter, src_col_letter in STYLE_TEMPLATE_MAP.items():
            tgt_col_idx = column_index_from_string(tgt_col_letter)
            src_col_idx = column_index_from_string(src_col_letter)

            src_cell = ws.cell(row=r, column=src_col_idx)
            dst_cell = ws.cell(row=r, column=tgt_col_idx)

            copy_cell_style(src_cell, dst_cell)

    # ---------------------------------------------------------
    # 2) Step11 タイトル（4行目：AS〜AX結合、文言設定）
    # ---------------------------------------------------------
    ws.merge_cells(start_row=TITLE_ROW, start_column=s_col, end_row=TITLE_ROW, end_column=e_col)
    c = ws.cell(row=TITLE_ROW, column=s_col)
    c.value = TITLE_TEXT

    # ---------------------------------------------------------
    # 3) Step11 見出し（5〜6行：各列ごと縦結合、文言設定）
    # ---------------------------------------------------------
    for col_letter, header in HEADERS:
        col_idx = column_index_from_string(col_letter)

        # 念のためその列の5-6行の結合を解除→再結合（衝突回避）
        unmerge_overlaps(ws, HEADER_TOP_ROW, HEADER_BOTTOM_ROW, col_idx, col_idx)
        ws.merge_cells(start_row=HEADER_TOP_ROW, start_column=col_idx, end_row=HEADER_BOTTOM_ROW, end_column=col_idx)

        ws.cell(row=HEADER_TOP_ROW, column=col_idx).value = header

    # ---------------------------------------------------------
    # 4) 保存
    # ---------------------------------------------------------
    wb.save(FILE_PATH)
    print(f"完了：Step10までの罫線・レイアウト（型）をStep11（AS〜AX）へ反映し、タイトル/見出しを設定しました。シート: {SHEET_NAME}")


if __name__ == "__main__":
    main()