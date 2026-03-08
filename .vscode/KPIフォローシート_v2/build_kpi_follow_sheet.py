"""
build_kpi_follow_sheet.py
KPIフォローシート v2 自動生成スクリプト（改善版）

改善点:
  - 課名を正式名称に変更
  - 目標値・達成率列を追加（達成率は自動計算・% 表示）
  - 達成率に条件付き書式（赤 <80% / 黄 80〜99% / 緑 100%+）
  - 課・戦略レバーのセル結合 + グループ別背景色
  - オートフィルター / A〜C列＋1行目のウィンドウ枠固定
  - 印刷設定（横向き・1ページ幅に収める）
  - 課別入力シート（1課1シート）＋ 営業推進部_一覧シート（課別参照集約）
  - 元データを KPIフォローシート.xlsx から自動読み込み
  - フォロー期間: FOLLOW_MONTHS か月分の実績列を横展開

Python 3.11 + openpyxl
出力: ./output/KPIフォローシート_v2.xlsx
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from utils_excel import header_map, norm

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# 入出力パス（ここだけ変更）
# ---------------------------------------------------------------------------

SRC_PATH = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\わかもと製薬_KPIフォローシート\KPIフォローシート\KPIフォローシート.xlsx"
SRC_SHEET = "月次KPIフォロー"

OUTPUT_PATH = (
    Path(__file__).parents[3]
    / "わかもと製薬_KPIフォローシート"
    / "KPIフォローシート_v2"
    / "KPIフォローシート_v2.xlsx"
)

# ---------------------------------------------------------------------------
# フォロー月数（ここを変えるだけで列構成が変わる）
# ---------------------------------------------------------------------------

FOLLOW_MONTHS = 3  # フォロー月数

# ---------------------------------------------------------------------------
# マスタデータ
# ---------------------------------------------------------------------------

MISS_REASONS = [
    "ターゲット設計",
    "医師判断軸",
    "競合優位",
    "MR活動量",
    "タイミング",
]

# 固定列（課・戦略レバー・KPI・目標値）
_FIXED_COLS: list[tuple] = [
    ("課",         16, "fixed", False),   # A
    ("戦略レバー", 22, "fixed", True),    # B
    ("KPI",        28, "fixed", True),    # C
    ("目標値",     12, "input", False),   # D
]

# 月次フォロー列（1か月分のテンプレート）
_MONTHLY_COLS: list[tuple] = [
    ("実績",           14, "input", False),
    ("達成率",         10, "calc",  False),
    ("結果コメント",   32, "input", True),
    ("未達要因",       20, "input", False),
    ("来月アクション", 32, "input", True),
    ("課長コメント",   32, "input", True),
    ("横展開事項",     32, "input", True),
]

TARGET_COL_IDX = 4                   # 目標値列（D列・固定）
N_FIXED        = len(_FIXED_COLS)    # 4
N_MONTHLY      = len(_MONTHLY_COLS)  # 7


def _month_cols(month: int) -> dict[str, int]:
    """月番号(1始まり)から列インデックス(1始まり)を返す"""
    base = N_FIXED + (month - 1) * N_MONTHLY
    return {
        "actual":   base + 1,  # 実績
        "achieve":  base + 2,  # 達成率
        "dropdown": base + 4,  # 未達要因
    }


def _make_field_columns(n_months: int) -> list[tuple]:
    cols = list(_FIXED_COLS)
    for m in range(1, n_months + 1):
        suffix = f"({m})" if n_months > 1 else ""
        for name, width, kind, wrap in _MONTHLY_COLS:
            cols.append((f"{name}{suffix}", width, kind, wrap))
    return cols


def _make_eiyo_columns(n_months: int) -> list[tuple]:
    cols = list(_FIXED_COLS)
    for m in range(1, n_months + 1):
        suffix = f"({m})" if n_months > 1 else ""
        for name, width, kind, wrap in _MONTHLY_COLS:
            ref_kind = "calc" if kind == "calc" else "ref"
            cols.append((f"{name}{suffix}", width, ref_kind, wrap))
    cols += [
        ("推進コメント", 32, "input", True),
        ("構造課題",     32, "input", True),
    ]
    return cols


FIELD_COLUMNS       = _make_field_columns(FOLLOW_MONTHS)
EIYO_SUISIN_COLUMNS = _make_eiyo_columns(FOLLOW_MONTHS)

# 課別シートの総列数（目標値 + 月次×FOLLOW_MONTHS）
KA_TOTAL_COLS = N_FIXED + FOLLOW_MONTHS * N_MONTHLY

BLANK_ROWS = 20  # データ行末尾に追加する空白入力行数

# ---------------------------------------------------------------------------
# カラー・スタイル定数
# ---------------------------------------------------------------------------

CLR_HEADER_BG = "1F3864"   # ヘッダー背景（濃紺）
CLR_HEADER_FG = "FFFFFF"   # ヘッダー文字（白）
CLR_INPUT_BG  = "FFFACD"   # 入力列（薄黄）
CLR_CALC_BG   = "E2EFDA"   # 達成率列（薄緑）
CLR_FIXED_BG  = "D9D9D9"   # 空白行の固定列（薄グレー）

# 課グループ別背景色（14課対応）
GROUP_COLORS = [
    "DAE3F3", "FCE4D6", "E2F0D9", "FFF2CC", "F2E0F7", "D5F5E3",
    "FDEBD0", "D6EAF8", "FDEDEC", "E8DAEF", "D5F5E3", "FDFEFE",
    "EBF5FB", "FEF9E7",
]

THIN       = Side(style="thin")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size, name="メイリオ")


def _align(h: str = "left", v: str = "top", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------------------------
# 元データ読み込み
# ---------------------------------------------------------------------------

def load_initial_data() -> list[tuple[str, str, str]]:
    """KPIフォローシート.xlsx の SRC_SHEET から (課, 戦略レバー, KPI) を読み込む。"""
    if not os.path.exists(SRC_PATH):
        raise FileNotFoundError(f"元データファイルが見つかりません: {SRC_PATH}")

    wb_src = openpyxl.load_workbook(SRC_PATH, data_only=True, read_only=True)
    if SRC_SHEET not in wb_src.sheetnames:
        raise ValueError(f"元シート '{SRC_SHEET}' が見つかりません")

    ws_src = wb_src[SRC_SHEET]
    cols   = header_map(ws_src)

    c_dept  = cols.get("課")
    c_lever = cols.get("主軸レバー(AE)") or cols.get("主軸レバー")
    c_kpi   = cols.get("KPI候補(AU)") or cols.get("KPI候補") or cols.get("KPI名")

    if c_dept is None or c_kpi is None:
        raise ValueError("元シートに '課' または 'KPI候補(AU)' が見つかりません（列名を確認してください）")

    rows: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str, str]] = set()

    for row in ws_src.iter_rows(min_row=2, values_only=True):
        dept  = row[c_dept  - 1] if c_dept  else None
        lever = row[c_lever - 1] if c_lever else None
        kpi   = row[c_kpi   - 1] if c_kpi   else None

        if dept in (None, "") or kpi in (None, ""):
            continue

        entry = (norm(dept), norm(lever), norm(kpi))
        if entry not in seen:
            seen.add(entry)
            rows.append(entry)

    wb_src.close()

    if not rows:
        raise ValueError(f"元シート '{SRC_SHEET}' にデータ行が見つかりませんでした")

    print(f"[INFO] 元データ読み込み完了: {len(rows)} 行 / {SRC_PATH}")
    return rows


# ---------------------------------------------------------------------------
# グループ情報の計算
# ---------------------------------------------------------------------------

def _ka_groups(data: list[tuple]) -> list[tuple[int, int, str, int]]:
    """課ごとの (start_row, end_row, 課名, group_idx) を返す（Excelの行番号）"""
    groups: list[tuple[int, int, str, int]] = []
    i, g = 0, 0
    while i < len(data):
        ka = data[i][0]
        j = i + 1
        while j < len(data) and data[j][0] == ka:
            j += 1
        # data[i] は Excel 行 i+2、data[j-1] は Excel 行 j+1
        groups.append((i + 2, j + 1, ka, g))
        i, g = j, g + 1
    return groups


def _lever_groups(data: list[tuple]) -> list[tuple[int, int, str]]:
    """戦略レバーごとの (start_row, end_row, レバー名) を返す（課またぎは分割）"""
    groups: list[tuple[int, int, str]] = []
    i = 0
    while i < len(data):
        ka, lever, _ = data[i]
        j = i + 1
        while j < len(data) and data[j][0] == ka and data[j][1] == lever:
            j += 1
        groups.append((i + 2, j + 1, lever))
        i = j
    return groups


# ---------------------------------------------------------------------------
# マスタシート
# ---------------------------------------------------------------------------

def build_master_sheet(wb: Workbook) -> None:
    """未達要因マスタシートを生成する（非表示）"""
    ws = wb.create_sheet("マスタ")
    ws["A1"] = "未達要因"
    ws["A1"].font = _font(bold=True)
    for i, reason in enumerate(MISS_REASONS, start=2):
        ws.cell(row=i, column=1, value=reason)
    ws.column_dimensions["A"].width = 20
    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# データシート共通処理
# ---------------------------------------------------------------------------

def _write_header(ws, columns: list[tuple]) -> None:
    for col_idx, (name, width, *_) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill      = _fill(CLR_HEADER_BG)
        cell.font      = _font(bold=True, color=CLR_HEADER_FG)
        cell.alignment = _align("center", "center")
        cell.border    = BORDER_ALL
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_dropdown(ws, col_letter: str, start_row: int, end_row: int) -> None:
    formula = f"マスタ!$A$2:$A${1 + len(MISS_REASONS)}"
    dv = DataValidation(type="list", formula1=formula,
                        allow_blank=True, showDropDown=False)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def _add_cf(ws, achieve_col_idx: int, total_rows: int) -> None:
    """達成率列に 赤/黄/緑 の条件付き書式を設定する"""
    c = get_column_letter(achieve_col_idx)
    cf_range = f"{c}2:{c}{total_rows}"

    def rule(formula_str: str, bg: str, font_color: str) -> Rule:
        return FormulaRule(
            formula=[formula_str],
            fill=PatternFill(fill_type="solid", fgColor=bg),
            font=Font(color=font_color, name="メイリオ", size=10),
        )

    # 優先度順（高 → 低）で追加
    ws.conditional_formatting.add(cf_range, rule(f"{c}2>=1",                "C6EFCE", "276221"))  # 緑
    ws.conditional_formatting.add(cf_range, rule(f"AND({c}2>=0.8,{c}2<1)",  "FFEB9C", "9C5700"))  # 黄
    ws.conditional_formatting.add(cf_range, rule(f'AND({c}2<>"",{c}2<0.8)', "FFC7CE", "9C0006"))  # 赤


def _build_data_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[tuple],
    data: list[tuple],
) -> None:
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, columns)

    ka_groups    = _ka_groups(data)
    lever_groups = _lever_groups(data)

    # 行番号 → グループ背景色
    row_to_color: dict[int, str] = {}
    for start, end, _, g_idx in ka_groups:
        color = GROUP_COLORS[g_idx % len(GROUP_COLORS)]
        for r in range(start, end + 1):
            row_to_color[r] = color

    data_end_row = 1 + len(data)
    total_rows   = data_end_row + BLANK_ROWS
    t_col        = get_column_letter(TARGET_COL_IDX)

    # ── データ行 ──────────────────────────────────────────────────────────
    for row_idx, (ka, lever, kpi) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=ka)
        ws.cell(row=row_idx, column=2, value=lever)
        ws.cell(row=row_idx, column=3, value=kpi)
        group_color = row_to_color.get(row_idx, CLR_FIXED_BG)

        for col_idx, col_def in enumerate(columns, start=1):
            kind, wrap = col_def[2], col_def[3]
            cell        = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_ALL
            cell.font   = _font()
            if kind == "fixed":
                cell.fill      = _fill(group_color)
                cell.alignment = _align("center", "top", wrap)
            elif kind == "calc":
                cell.fill          = _fill(CLR_CALC_BG)
                cell.alignment     = _align("center", "center")
                cell.number_format = "0%"
            else:
                cell.fill      = _fill(CLR_INPUT_BG)
                cell.alignment = _align("left", "top", wrap)

        # 各月の達成率数式
        for m in range(1, FOLLOW_MONTHS + 1):
            mc = _month_cols(m)
            a_col = get_column_letter(mc["actual"])
            ws.cell(row=row_idx, column=mc["achieve"]).value = (
                f'=IFERROR({a_col}{row_idx}/{t_col}{row_idx},"")'
            )

    # ── 空白入力行 ────────────────────────────────────────────────────────
    for row_idx in range(data_end_row + 1, total_rows + 1):
        for col_idx, col_def in enumerate(columns, start=1):
            kind, wrap = col_def[2], col_def[3]
            cell        = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_ALL
            cell.font   = _font()
            if kind == "fixed":
                cell.fill      = _fill(CLR_FIXED_BG)
                cell.alignment = _align("center", "top", wrap)
            elif kind == "calc":
                cell.fill          = _fill(CLR_CALC_BG)
                cell.alignment     = _align("center", "center")
                cell.number_format = "0%"
            else:
                cell.fill      = _fill(CLR_INPUT_BG)
                cell.alignment = _align("left", "top", wrap)

        for m in range(1, FOLLOW_MONTHS + 1):
            mc = _month_cols(m)
            a_col = get_column_letter(mc["actual"])
            ws.cell(row=row_idx, column=mc["achieve"]).value = (
                f'=IFERROR({a_col}{row_idx}/{t_col}{row_idx},"")'
            )

    # ── 課・戦略レバーのセル結合 ──────────────────────────────────────────
    a_wrap = columns[0][3]
    b_wrap = columns[1][3]

    for start, end, _, _ in ka_groups:
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
        ws.cell(row=start, column=1).alignment = _align("center", "top", a_wrap)

    for start, end, _ in lever_groups:
        if end > start:
            ws.merge_cells(f"B{start}:B{end}")
        ws.cell(row=start, column=2).alignment = _align("center", "top", b_wrap)

    # ── 条件付き書式・プルダウン（全月分） ───────────────────────────────
    for m in range(1, FOLLOW_MONTHS + 1):
        mc = _month_cols(m)
        _add_cf(ws, mc["achieve"], total_rows)
        _add_dropdown(ws, get_column_letter(mc["dropdown"]), 2, total_rows)

    # ── フィルター・枠固定 ────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{total_rows}"
    ws.freeze_panes    = "D2"   # A〜C列 + 1行目を固定

    # ── 行高 ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    for r in range(2, total_rows + 1):
        ws.row_dimensions[r].height = 30

    # ── 印刷設定 ──────────────────────────────────────────────────────────
    ws.print_title_rows         = "1:1"
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0
    ws.print_area               = f"A1:{get_column_letter(len(columns))}{total_rows}"


# ---------------------------------------------------------------------------
# 課別入力シート
# ---------------------------------------------------------------------------

def build_ka_input_sheets(wb: Workbook, data: list[tuple]) -> None:
    """課ごとの個別入力シートを生成する（1課1シート）"""
    ka_order: list[str] = []
    ka_rows: dict[str, list[tuple]] = {}
    for ka, lever, kpi in data:
        if ka not in ka_rows:
            ka_order.append(ka)
            ka_rows[ka] = []
        ka_rows[ka].append((ka, lever, kpi))

    for ka_name in ka_order:
        _build_data_sheet(
            wb, ka_name, FIELD_COLUMNS,
            data=ka_rows[ka_name],
        )


# ---------------------------------------------------------------------------
# 営業推進部_一覧シート（課別シートを参照して集約）
# ---------------------------------------------------------------------------

def build_eiyo_suisin_sheet(wb: Workbook, data: list[tuple]) -> None:
    """営業推進部_一覧シートを生成する。
    課別入力シートの月次列を参照し、課ごとに区分けして一覧表示する。
    推進コメント・構造課題は営業推進部が直接入力する。
    """
    ws = wb.create_sheet("営業推進部_一覧")
    _write_header(ws, EIYO_SUISIN_COLUMNS)

    # 課ごとにデータを整理（順序保持）
    ka_order: list[str] = []
    ka_rows: dict[str, list[tuple[str, str]]] = {}
    for ka, lever, kpi in data:
        if ka not in ka_rows:
            ka_order.append(ka)
            ka_rows[ka] = []
        ka_rows[ka].append((lever, kpi))

    current_row = 2
    ka_merge_ranges:    list[tuple[int, int, int]] = []  # (start, end, g_idx)
    lever_merge_ranges: list[tuple[int, int]]       = []  # (start, end)

    for g_idx, ka_name in enumerate(ka_order):
        rows       = ka_rows[ka_name]
        n_data     = len(rows)
        n_total    = n_data + BLANK_ROWS
        group_color = GROUP_COLORS[g_idx % len(GROUP_COLORS)]
        ka_start   = current_row

        # 戦略レバーの結合範囲（データ行のみ）
        i = 0
        while i < n_data:
            lever = rows[i][0]
            j = i + 1
            while j < n_data and rows[j][0] == lever:
                j += 1
            lever_merge_ranges.append((ka_start + i, ka_start + j - 1))
            i = j

        for row_offset in range(n_total):
            src_row = row_offset + 2   # 参照先・課別シートの行番号
            dest_row = current_row
            is_data  = row_offset < n_data

            lever_val = rows[row_offset][0] if is_data else ""
            kpi_val   = rows[row_offset][1] if is_data else ""

            for col_idx, col_def in enumerate(EIYO_SUISIN_COLUMNS, start=1):
                _, _, kind, wrap = col_def
                cell = ws.cell(row=dest_row, column=col_idx)
                cell.border = BORDER_ALL
                cell.font   = _font()

                if col_idx == 1:       # 課 (A) ← 構造から直接設定
                    cell.value     = ka_name
                    cell.fill      = _fill(group_color)
                    cell.alignment = _align("center", "center")

                elif col_idx == 2:     # 戦略レバー (B)
                    cell.value     = lever_val
                    cell.fill      = _fill(group_color)
                    cell.alignment = _align("center", "top", True)

                elif col_idx == 3:     # KPI (C)
                    cell.value     = kpi_val
                    cell.fill      = _fill(group_color)
                    cell.alignment = _align("center", "top", True)

                elif kind == "calc":   # 達成率 ← 課別!同列 を参照
                    safe = ka_name.replace("'", "''")
                    src_col = get_column_letter(col_idx)
                    cell.value         = f"='{safe}'!{src_col}{src_row}"
                    cell.fill          = _fill(CLR_CALC_BG)
                    cell.alignment     = _align("center", "center")
                    cell.number_format = "0%"

                elif kind == "ref":    # 目標値・月次入力列 ← 課別!同列 を参照
                    safe    = ka_name.replace("'", "''")
                    src_col = get_column_letter(col_idx)
                    cell.value     = f"='{safe}'!{src_col}{src_row}"
                    cell.fill      = _fill(CLR_INPUT_BG)
                    cell.alignment = _align("left", "top", wrap)

                else:                  # 推進コメント・構造課題 ← 直接入力
                    cell.fill      = _fill(CLR_INPUT_BG)
                    cell.alignment = _align("left", "top", wrap)

            current_row += 1

        ka_merge_ranges.append((ka_start, current_row - 1, g_idx))

    total_rows = current_row - 1

    # ── 課 (A列) のセル結合 ───────────────────────────────────────────────
    for start, end, _ in ka_merge_ranges:
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
        ws.cell(row=start, column=1).alignment = _align("center", "center")

    # ── 戦略レバー (B列) のセル結合（データ行のみ） ────────────────────────
    for start, end in lever_merge_ranges:
        if end > start:
            ws.merge_cells(f"B{start}:B{end}")
        ws.cell(row=start, column=2).alignment = _align("center", "top", True)

    # ── 条件付き書式・プルダウン（全月分） ───────────────────────────────
    for m in range(1, FOLLOW_MONTHS + 1):
        mc = _month_cols(m)
        _add_cf(ws, mc["achieve"], total_rows)
        _add_dropdown(ws, get_column_letter(mc["dropdown"]), 2, total_rows)

    # ── フィルター・枠固定 ────────────────────────────────────────────────
    n_cols = len(EIYO_SUISIN_COLUMNS)
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{total_rows}"
    ws.freeze_panes    = "D2"

    # ── 行高 ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    for r in range(2, total_rows + 1):
        ws.row_dimensions[r].height = 30

    # ── 印刷設定 ──────────────────────────────────────────────────────────
    ws.print_title_rows       = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area             = f"A1:{get_column_letter(n_cols)}{total_rows}"


# ---------------------------------------------------------------------------
# 公開ビルダー
# ---------------------------------------------------------------------------

def build() -> None:
    """ワークブックを組み立てて保存する"""
    data = load_initial_data()

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    build_ka_input_sheets(wb, data)    # 課別入力シート
    build_eiyo_suisin_sheet(wb, data)  # 営業推進部_一覧（課別シートを参照して集約）
    build_master_sheet(wb)             # 未達要因マスタ（非表示）

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] 生成完了: {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
