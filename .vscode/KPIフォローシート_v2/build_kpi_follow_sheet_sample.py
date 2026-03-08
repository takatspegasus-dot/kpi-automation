"""
build_kpi_follow_sheet_sample.py
サンプルデータを使ってKPIフォローシート v2 の表示を検証するスクリプト

元データファイル（KPIフォローシート.xlsx）が不要。
サンプルデータで即座にExcelを生成して表示確認できる。

対象期間: ４月・５月・６月（３か月）
出力: ./output/KPIフォローシート_v2_sample.xlsx
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

# メインスクリプトと同じユーティリティを共有
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# 出力パス
# ---------------------------------------------------------------------------

OUTPUT_PATH = (
    Path(__file__).parents[3]
    / "わかもと製薬_KPIフォローシート"
    / "KPIフォローシート_v2"
    / "KPIフォローシート_v2_sample.xlsx"
)

# ---------------------------------------------------------------------------
# サンプルデータ（課, 戦略レバー, KPI）
# ---------------------------------------------------------------------------

SAMPLE_DATA: list[tuple[str, str, str]] = [
    # 第一営業課
    ("第一営業課", "訪問活動強化",         "月次訪問件数（対象医師）"),
    ("第一営業課", "訪問活動強化",         "新規開拓医師数"),
    ("第一営業課", "訪問活動強化",         "ハイポテンシャル医師への月次面談率"),
    ("第一営業課", "製品プロモーション",   "製品説明会実施件数"),
    ("第一営業課", "製品プロモーション",   "パンフレット・資材配布数"),
    # 第二営業課
    ("第二営業課", "訪問活動強化",         "月次訪問件数（対象医師）"),
    ("第二営業課", "訪問活動強化",         "初回訪問率（新規医師）"),
    ("第二営業課", "売上貢献",             "担当エリア製品売上金額（万円）"),
    ("第二営業課", "売上貢献",             "前年同月比達成率"),
    ("第二営業課", "売上貢献",             "重点品目シェア率（対象病院）"),
    # 第三営業課
    ("第三営業課", "顧客関係強化",         "主要医師との面談回数（月次）"),
    ("第三営業課", "顧客関係強化",         "薬剤師向けフォロー訪問件数"),
    ("第三営業課", "デジタル活用",         "eディテール実施率（担当医師対比）"),
    ("第三営業課", "デジタル活用",         "Web講演会への医師招待・参加数"),
    # マーケティング課
    ("マーケティング課", "市場分析",       "競合動向レポート提出回数"),
    ("マーケティング課", "市場分析",       "処方データ分析レポートの定例共有回数"),
    ("マーケティング課", "販促支援",       "営業向けプロモーション資材作成件数"),
    ("マーケティング課", "販促支援",       "学術情報提供件数（講演・文献等）"),
    # 医薬情報課
    ("医薬情報課", "情報提供活動",         "医師・薬剤師への適正使用情報提供件数"),
    ("医薬情報課", "情報提供活動",         "副作用報告の収集・対応件数"),
    ("医薬情報課", "教育・研修",           "MR向け社内勉強会実施回数"),
]

# サンプル入力値（月別目標値・実績・コメントなどを仮置き）
SAMPLE_INPUT: dict[int, dict] = {
    0: {
        "apr_target": 80,  "apr_actual": 72,
        "may_target": 80,  "may_actual": 76,
        "jun_target": 80,  "jun_actual": 80,
        "comment": "月初の訪問計画未達。後半で挽回を試みたが届かず。",
        "reason": "MR活動量",
        "action": "週次訪問計画を細分化し、遅延を早期検知する体制を整備する。",
        "mgr": "計画精度向上を優先。週次レビューを導入してほしい。",
        "rollout": "",
    },
    1: {
        "apr_target": 5,   "apr_actual": 7,
        "may_target": 5,   "may_actual": 6,
        "jun_target": 5,   "jun_actual": 5,
        "comment": "新規医師へのアプローチが計画以上に進捗した。",
        "reason": "",
        "action": "引き続き新規開拓を継続。アプローチ先リストを更新する。",
        "mgr": "好調。成功パターンを他課にも展開してほしい。",
        "rollout": "新規開拓手法を第二・第三営業課へ共有",
    },
    2: {
        "apr_target": 0.7, "apr_actual": 0.65,
        "may_target": 0.7, "may_actual": 0.68,
        "jun_target": 0.7, "jun_actual": 0.70,
        "comment": "ハイポテンシャル医師のアポイント取得に苦戦。",
        "reason": "タイミング",
        "action": "学会・セミナー前後のアプローチ時期を最適化する。",
        "mgr": "学会カレンダーと連動した訪問計画を立てること。",
        "rollout": "",
    },
    3: {
        "apr_target": 4,   "apr_actual": 4,
        "may_target": 4,   "may_actual": 4,
        "jun_target": 4,   "jun_actual": 4,
        "comment": "予定通り4回実施。参加者平均8名。",
        "reason": "",
        "action": "次期は参加者数10名以上を目標に招待範囲を拡大する。",
        "mgr": "品質向上も意識してほしい。",
        "rollout": "",
    },
    4: {
        "apr_target": 200, "apr_actual": 185,
        "may_target": 200, "may_actual": 195,
        "jun_target": 200, "jun_actual": 200,
        "comment": "資材の在庫切れが一部で発生し配布数が減少した。",
        "reason": "ターゲット設計",
        "action": "資材の事前発注タイミングを2週間前倒しにする。",
        "mgr": "在庫管理フローを見直すこと。",
        "rollout": "",
    },
}

# ---------------------------------------------------------------------------
# 定数・ユーティリティ
# ---------------------------------------------------------------------------

MISS_REASONS = ["ターゲット設計", "医師判断軸", "競合優位", "MR活動量", "タイミング"]

# 列定義: (列名, 幅, 種別, 折り返し)
# 種別: "fixed"=固定, "input"=入力, "calc"=計算, "ref"=参照（営業推進シート用）
FIELD_COLUMNS = [
    ("課",             18, "fixed", False),   # A  col 1
    ("戦略レバー",     24, "fixed", True),    # B  col 2
    ("KPI",            36, "fixed", True),    # C  col 3
    ("４月目標",       12, "input", False),   # D  col 4
    ("４月実績",       12, "input", False),   # E  col 5
    ("４月達成率",     10, "calc",  False),   # F  col 6
    ("５月目標",       12, "input", False),   # G  col 7
    ("５月実績",       12, "input", False),   # H  col 8
    ("５月達成率",     10, "calc",  False),   # I  col 9
    ("６月目標",       12, "input", False),   # J  col 10
    ("６月実績",       12, "input", False),   # K  col 11
    ("６月達成率",     10, "calc",  False),   # L  col 12
    ("結果コメント",   38, "input", True),    # M  col 13
    ("未達要因",       22, "input", False),   # N  col 14
    ("次期アクション", 38, "input", True),    # O  col 15
    ("課長コメント",   38, "input", True),    # P  col 16
    ("横展開事項",     38, "input", True),    # Q  col 17
]

EIYO_SUISIN_COLUMNS = [
    ("課",             18, "fixed", False),   # col 1
    ("戦略レバー",     24, "fixed", True),    # col 2
    ("KPI",            36, "fixed", True),    # col 3
    ("４月目標",       12, "ref",   False),   # col 4
    ("４月実績",       12, "ref",   False),   # col 5
    ("４月達成率",     10, "calc",  False),   # col 6
    ("５月目標",       12, "ref",   False),   # col 7
    ("５月実績",       12, "ref",   False),   # col 8
    ("５月達成率",     10, "calc",  False),   # col 9
    ("６月目標",       12, "ref",   False),   # col 10
    ("６月実績",       12, "ref",   False),   # col 11
    ("６月達成率",     10, "calc",  False),   # col 12
    ("結果コメント",   38, "ref",   True),    # col 13
    ("未達要因",       22, "ref",   False),   # col 14
    ("次期アクション", 38, "ref",   True),    # col 15
    ("課長コメント",   38, "ref",   True),    # col 16
    ("横展開事項",     38, "ref",   True),    # col 17
    ("推進コメント",   38, "input", True),    # col 18
    ("構造課題",       38, "input", True),    # col 19
]

# 月ごとの（目標列, 実績列, 達成率列）インデックス（1-based）
MONTH_TRIPLETS = [(4, 5, 6), (7, 8, 9), (10, 11, 12)]
# 月ごとのサンプル入力キー
MONTH_INPUT_KEYS = [
    ("apr_target", "apr_actual"),
    ("may_target", "may_actual"),
    ("jun_target", "jun_actual"),
]

BLANK_ROWS = 5  # サンプルでは5行に短縮（確認しやすく）

CLR_HEADER_BG = "1F3864"
CLR_HEADER_FG = "FFFFFF"
CLR_INPUT_BG  = "FFFACD"
CLR_CALC_BG   = "E2EFDA"
CLR_FIXED_BG  = "D9D9D9"

GROUP_COLORS = [
    "DAE3F3", "FCE4D6", "E2F0D9", "FFF2CC", "F2E0F7", "D5F5E3",
    "FDEBD0", "D6EAF8", "FDEDEC", "E8DAEF", "D5F5E3", "FDFEFE",
    "EBF5FB", "FEF9E7",
]

THIN       = Side(style="thin")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size, name="メイリオ")


def _align(h: str = "left", v: str = "top", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _ka_groups(data: list[tuple]) -> list[tuple[int, int, str, int]]:
    groups = []
    i, g = 0, 0
    while i < len(data):
        ka = data[i][0]
        j = i + 1
        while j < len(data) and data[j][0] == ka:
            j += 1
        groups.append((i + 2, j + 1, ka, g))
        i, g = j, g + 1
    return groups


def _lever_groups(data: list[tuple]) -> list[tuple[int, int, str]]:
    groups = []
    i = 0
    while i < len(data):
        ka, lever, _ = data[i]
        j = i + 1
        while j < len(data) and data[j][0] == ka and data[j][1] == lever:
            j += 1
        groups.append((i + 2, j + 1, lever))
        i = j
    return groups


def _write_header(ws, columns: list[tuple]) -> None:
    for col_idx, (name, width, *_) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill      = _fill(CLR_HEADER_BG)
        cell.font      = _font(bold=True, color=CLR_HEADER_FG)
        cell.alignment = _align("center", "center")
        cell.border    = BORDER_ALL
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_dropdown(ws, col_letter: str, start_row: int, end_row: int) -> None:
    formula = f"マスタ!$A$2:$A${1 + len(MISS_REASONS)}"
    dv = DataValidation(type="list", formula1=formula,
                        allow_blank=True, showDropDown=False)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def _add_cf(ws, achieve_col_indices: list[int], total_rows: int) -> None:
    def rule(formula_str, bg, font_color):
        return FormulaRule(
            formula=[formula_str],
            fill=PatternFill(fill_type="solid", fgColor=bg),
            font=Font(color=font_color, name="メイリオ", size=10),
        )

    for achieve_col_idx in achieve_col_indices:
        c = get_column_letter(achieve_col_idx)
        cf_range = f"{c}2:{c}{total_rows}"
        ws.conditional_formatting.add(cf_range, rule(f"{c}2>=1",                "C6EFCE", "276221"))
        ws.conditional_formatting.add(cf_range, rule(f"AND({c}2>=0.8,{c}2<1)",  "FFEB9C", "9C5700"))
        ws.conditional_formatting.add(cf_range, rule(f'AND({c}2<>"",{c}2<0.8)', "FFC7CE", "9C0006"))


def _build_data_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[tuple],
    dropdown_col: str,
    month_triplets: list[tuple[int, int, int]],
    data: list[tuple],
    sample_input: dict[int, dict] | None = None,
    data_offset: int = 0,
) -> None:
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, columns)

    ka_groups    = _ka_groups(data)
    lever_groups = _lever_groups(data)

    row_to_color: dict[int, str] = {}
    for start, end, _, g_idx in ka_groups:
        color = GROUP_COLORS[g_idx % len(GROUP_COLORS)]
        for r in range(start, end + 1):
            row_to_color[r] = color

    data_end_row = 1 + len(data)
    total_rows   = data_end_row + BLANK_ROWS

    for row_idx, (ka, lever, kpi) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=ka)
        ws.cell(row=row_idx, column=2, value=lever)
        ws.cell(row=row_idx, column=3, value=kpi)
        group_color = row_to_color.get(row_idx, CLR_FIXED_BG)

        global_idx = data_offset + (row_idx - 2)
        si = (sample_input or {}).get(global_idx, {})

        for col_idx, col_def in enumerate(columns, start=1):
            kind, wrap = col_def[2], col_def[3]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_ALL
            cell.font   = _font()
            if kind == "fixed":
                cell.fill      = _fill(group_color)
                cell.alignment = _align("center", "top", wrap)
            elif kind == "calc":
                cell.fill          = _fill(CLR_CALC_BG)
                cell.alignment     = _align("center", "center")
                cell.number_format = "0%"
            else:
                cell.fill      = _fill(CLR_INPUT_BG)
                cell.alignment = _align("left", "top", wrap)

        # 月別達成率数式
        for t_col, a_col, ach_col in month_triplets:
            t_ltr = get_column_letter(t_col)
            a_ltr = get_column_letter(a_col)
            ws.cell(row=row_idx, column=ach_col).value = (
                f'=IFERROR({a_ltr}{row_idx}/{t_ltr}{row_idx},"")'
            )

        # サンプルデータの入力値を設定
        if si:
            for i, (t_col, a_col, _) in enumerate(month_triplets):
                tk, ak = MONTH_INPUT_KEYS[i]
                if tk in si:
                    ws.cell(row=row_idx, column=t_col).value = si[tk]
                if ak in si:
                    ws.cell(row=row_idx, column=a_col).value = si[ak]
            # M: 結果コメント
            if "comment" in si and len(columns) >= 13:
                ws.cell(row=row_idx, column=13).value = si["comment"]
            # N: 未達要因
            if "reason" in si and len(columns) >= 14:
                ws.cell(row=row_idx, column=14).value = si["reason"]
            # O: 次期アクション
            if "action" in si and len(columns) >= 15:
                ws.cell(row=row_idx, column=15).value = si["action"]
            # P: 課長コメント
            if "mgr" in si and len(columns) >= 16:
                ws.cell(row=row_idx, column=16).value = si["mgr"]
            # Q: 横展開事項
            if "rollout" in si and len(columns) >= 17:
                ws.cell(row=row_idx, column=17).value = si["rollout"]

    # 空白入力行
    for row_idx in range(data_end_row + 1, total_rows + 1):
        for col_idx, col_def in enumerate(columns, start=1):
            kind, wrap = col_def[2], col_def[3]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_ALL
            cell.font   = _font()
            if kind == "fixed":
                cell.fill      = _fill(CLR_FIXED_BG)
                cell.alignment = _align("center", "top", wrap)
            elif kind == "calc":
                cell.fill          = _fill(CLR_CALC_BG)
                cell.alignment     = _align("center", "center")
                cell.number_format = "0%"
            else:
                cell.fill      = _fill(CLR_INPUT_BG)
                cell.alignment = _align("left", "top", wrap)
        for t_col, a_col, ach_col in month_triplets:
            t_ltr = get_column_letter(t_col)
            a_ltr = get_column_letter(a_col)
            ws.cell(row=row_idx, column=ach_col).value = (
                f'=IFERROR({a_ltr}{row_idx}/{t_ltr}{row_idx},"")'
            )

    # 課・戦略レバーのセル結合
    a_wrap = columns[0][3]
    b_wrap = columns[1][3]
    for start, end, _, _ in ka_groups:
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
        ws.cell(row=start, column=1).alignment = _align("center", "top", a_wrap)
    for start, end, _ in lever_groups:
        if end > start:
            ws.merge_cells(f"B{start}:B{end}")
        ws.cell(row=start, column=2).alignment = _align("center", "top", b_wrap)

    _add_cf(ws, [ach_col for _, _, ach_col in month_triplets], total_rows)
    _add_dropdown(ws, dropdown_col, 2, total_rows)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{total_rows}"
    ws.freeze_panes    = "D2"

    ws.row_dimensions[1].height = 22
    for r in range(2, total_rows + 1):
        ws.row_dimensions[r].height = 42

    ws.print_title_rows       = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{get_column_letter(len(columns))}{total_rows}"


def build_master_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("マスタ")
    ws["A1"] = "未達要因"
    ws["A1"].font = _font(bold=True)
    for i, reason in enumerate(MISS_REASONS, start=2):
        ws.cell(row=i, column=1, value=reason)
    ws.column_dimensions["A"].width = 20
    ws.sheet_state = "hidden"


def build_ka_input_sheets(wb: Workbook, data: list[tuple]) -> None:
    ka_order: list[str] = []
    ka_rows: dict[str, list[tuple]] = {}
    ka_offsets: dict[str, int] = {}
    offset = 0
    for ka, lever, kpi in data:
        if ka not in ka_rows:
            ka_order.append(ka)
            ka_rows[ka] = []
            ka_offsets[ka] = offset
        ka_rows[ka].append((ka, lever, kpi))
        offset += 1

    for ka_name in ka_order:
        _build_data_sheet(
            wb, ka_name, FIELD_COLUMNS, "N",
            month_triplets=MONTH_TRIPLETS,
            data=ka_rows[ka_name],
            sample_input=SAMPLE_INPUT,
            data_offset=ka_offsets[ka_name],
        )


def build_eiyo_suisin_sheet(wb: Workbook, data: list[tuple]) -> None:
    ws = wb.create_sheet("営業推進部_一覧")
    _write_header(ws, EIYO_SUISIN_COLUMNS)

    ka_order: list[str] = []
    ka_rows: dict[str, list[tuple[str, str]]] = {}
    for ka, lever, kpi in data:
        if ka not in ka_rows:
            ka_order.append(ka)
            ka_rows[ka] = []
        ka_rows[ka].append((lever, kpi))

    current_row = 2
    ka_merge_ranges:    list[tuple[int, int, int]] = []
    lever_merge_ranges: list[tuple[int, int]]       = []

    for g_idx, ka_name in enumerate(ka_order):
        rows        = ka_rows[ka_name]
        n_data      = len(rows)
        n_total     = n_data + BLANK_ROWS
        group_color = GROUP_COLORS[g_idx % len(GROUP_COLORS)]
        ka_start    = current_row

        i = 0
        while i < n_data:
            lever = rows[i][0]
            j = i + 1
            while j < n_data and rows[j][0] == lever:
                j += 1
            lever_merge_ranges.append((ka_start + i, ka_start + j - 1))
            i = j

        for row_offset in range(n_total):
            src_row  = row_offset + 2
            dest_row = current_row
            is_data  = row_offset < n_data

            lever_val = rows[row_offset][0] if is_data else ""
            kpi_val   = rows[row_offset][1] if is_data else ""
            safe      = ka_name.replace("'", "''")

            for col_idx, col_def in enumerate(EIYO_SUISIN_COLUMNS, start=1):
                _, _, kind, wrap = col_def
                cell = ws.cell(row=dest_row, column=col_idx)
                cell.border = BORDER_ALL
                cell.font   = _font()

                if col_idx == 1:
                    cell.value     = ka_name
                    cell.fill      = _fill(group_color)
                    cell.alignment = _align("center", "center")
                elif col_idx == 2:
                    cell.value     = lever_val
                    cell.fill      = _fill(group_color)
                    cell.alignment = _align("center", "top", True)
                elif col_idx == 3:
                    cell.value     = kpi_val
                    cell.fill      = _fill(group_color)
                    cell.alignment = _align("center", "top", True)
                elif kind == "calc":
                    src_col            = get_column_letter(col_idx)
                    cell.value         = f"='{safe}'!{src_col}{src_row}"
                    cell.fill          = _fill(CLR_CALC_BG)
                    cell.alignment     = _align("center", "center")
                    cell.number_format = "0%"
                elif kind == "ref":
                    src_col        = get_column_letter(col_idx)
                    cell.value     = f"='{safe}'!{src_col}{src_row}"
                    cell.fill      = _fill(CLR_INPUT_BG)
                    cell.alignment = _align("left", "top", wrap)
                else:  # input
                    cell.fill      = _fill(CLR_INPUT_BG)
                    cell.alignment = _align("left", "top", wrap)

            current_row += 1

        ka_merge_ranges.append((ka_start, current_row - 1, g_idx))

    total_rows = current_row - 1

    for start, end, _ in ka_merge_ranges:
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
        ws.cell(row=start, column=1).alignment = _align("center", "center")

    for start, end in lever_merge_ranges:
        if end > start:
            ws.merge_cells(f"B{start}:B{end}")
        ws.cell(row=start, column=2).alignment = _align("center", "top", True)

    achieve_cols = [col_idx for col_idx, col_def in enumerate(EIYO_SUISIN_COLUMNS, start=1)
                    if col_def[2] == "calc"]
    _add_cf(ws, achieve_cols, total_rows)
    _add_dropdown(ws, "N", 2, total_rows)
    n_cols = len(EIYO_SUISIN_COLUMNS)
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{total_rows}"
    ws.freeze_panes    = "D2"

    ws.row_dimensions[1].height = 22
    for r in range(2, total_rows + 1):
        ws.row_dimensions[r].height = 42

    ws.print_title_rows       = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{get_column_letter(n_cols)}{total_rows}"


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------

def build() -> None:
    data = SAMPLE_DATA
    print(f"[INFO] サンプルデータ使用: {len(data)} 行（対象期間: ４月・５月・６月）")

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    build_ka_input_sheets(wb, data)
    build_eiyo_suisin_sheet(wb, data)
    build_master_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] サンプル生成完了: {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
