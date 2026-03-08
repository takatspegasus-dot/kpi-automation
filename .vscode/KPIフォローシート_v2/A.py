"""
build_kpi_follow_sheet_from_master.py

元ファイル:
KPIフォローシート.xlsx

元ファイルの「月次KPIフォロー」シートを読み込み、
課ごとに別Excelファイルを出力する。

元ファイル:
C:/Users/SujiT/OneDrive/ドキュメント/HiproBiz わかもと製薬/わかもと製薬_KPIフォローシート/KPIフォローシート/KPIフォローシート.xlsx

出力先:
C:/Users/SujiT/OneDrive/ドキュメント/HiproBiz わかもと製薬/わかもと製薬_KPIフォローシート/KPIフォローシート_v2/KPIフォローシート_課別
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# パス
# ---------------------------------------------------------------------------

SOURCE_FILE = Path(
    "C:/Users/SujiT/OneDrive/ドキュメント/HiproBiz わかもと製薬/わかもと製薬_KPIフォローシート/KPIフォローシート/KPIフォローシート.xlsx"
)

OUTPUT_DIR = Path(
    "C:/Users/SujiT/OneDrive/ドキュメント/HiproBiz わかもと製薬/わかもと製薬_KPIフォローシート/KPIフォローシート_v2/KPIフォローシート_課別"
)

SOURCE_SHEET_NAME = "月次KPIフォロー"
PUSHIN_SHEET_NAME = "推進部_月次管理"

KA_LIST = [
    "札幌", "仙台",
    "東京1", "東京2", "東京3", "東京4", "東京5",
    "名古屋1", "名古屋2",
    "大阪1", "大阪2",
    "広島",
    "福岡1", "福岡2",
]

AUX_SHEETS = ["_lists", "_dv_lists", "_tpl_lists"]

# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def safe_filename(name: str) -> str:
    for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
        name = name.replace(ch, "_")
    return name


def copy_cell(src_cell, dst_cell) -> None:
    dst_cell.value = src_cell.value

    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)

    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)

    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)

def copy_column_widths(src_ws, dst_ws, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in src_ws.column_dimensions:
            src_dim = src_ws.column_dimensions[col_letter]
            dst_dim = dst_ws.column_dimensions[col_letter]

            # 安全にコピーできる属性だけを移す
            dst_dim.width = src_dim.width
            dst_dim.hidden = src_dim.hidden
            dst_dim.bestFit = src_dim.bestFit
            dst_dim.outlineLevel = src_dim.outlineLevel
            dst_dim.collapsed = src_dim.collapsed
            dst_dim.min = src_dim.min
            dst_dim.max = src_dim.max


def copy_row_heights(src_ws, dst_ws, row_map: dict[int, int] | None = None) -> None:
    """
    row_map がある場合は {元行: 新行} に従ってコピー
    row_map がない場合は全行を同じ行番号でコピー
    """
    if row_map:
        for src_row, dst_row in row_map.items():
            if src_row in src_ws.row_dimensions:
                src_dim = src_ws.row_dimensions[src_row]
                dst_dim = dst_ws.row_dimensions[dst_row]
                if src_dim.height is not None:
                    dst_dim.height = src_dim.height
                dst_dim.hidden = src_dim.hidden
    else:
        for row_idx, src_dim in src_ws.row_dimensions.items():
            dst_dim = dst_ws.row_dimensions[row_idx]
            if src_dim.height is not None:
                dst_dim.height = src_dim.height
            dst_dim.hidden = src_dim.hidden


def copy_sheet_views_and_print_settings(src_ws, dst_ws) -> None:
    dst_ws.freeze_panes = src_ws.freeze_panes

    if src_ws.auto_filter and src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref

    dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines

    dst_ws.page_setup.orientation = src_ws.page_setup.orientation
    dst_ws.page_setup.paperSize = src_ws.page_setup.paperSize
    dst_ws.page_setup.fitToWidth = src_ws.page_setup.fitToWidth
    dst_ws.page_setup.fitToHeight = src_ws.page_setup.fitToHeight
    dst_ws.page_margins = copy(src_ws.page_margins)
    dst_ws.print_options = copy(src_ws.print_options)

    try:
        dst_ws.print_title_rows = src_ws.print_title_rows
    except Exception:
        pass

    try:
        dst_ws.print_title_cols = src_ws.print_title_cols
    except Exception:
        pass

    try:
        dst_ws.print_area = src_ws.print_area
    except Exception:
        pass


def copy_merged_cells_subset(src_ws, dst_ws, row_map: dict[int, int], max_col: int) -> None:
    """
    元シートの結合セルのうち、対象行だけで完結しているものをコピー
    """
    for merged_range in src_ws.merged_cells.ranges:
        min_col, min_row, max_col_rng, max_row = merged_range.bounds

        if max_col_rng > max_col:
            continue

        source_rows = list(range(min_row, max_row + 1))
        if all(r in row_map for r in source_rows):
            new_min_row = row_map[min_row]
            new_max_row = row_map[max_row]
            start_cell = f"{get_column_letter(min_col)}{new_min_row}"
            end_cell = f"{get_column_letter(max_col_rng)}{new_max_row}"
            dst_ws.merge_cells(f"{start_cell}:{end_cell}")


def extract_rows_by_ka(src_ws, ka_name: str) -> list[int]:
    """
    B列=課 で対象行を取得
    1行目はヘッダーなので除外
    """
    target_rows: list[int] = []
    for row_idx in range(2, src_ws.max_row + 1):
        ka_val = src_ws.cell(row=row_idx, column=2).value
        if ka_val == ka_name:
            target_rows.append(row_idx)
    return target_rows


def copy_header_row(src_ws, dst_ws, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        copy_cell(src_ws.cell(row=1, column=col_idx), dst_ws.cell(row=1, column=col_idx))


def copy_selected_rows(
    src_ws,
    dst_ws,
    src_rows: list[int],
    max_col: int,
) -> dict[int, int]:
    """
    指定行だけをコピーし、{元行: 新行} を返す
    """
    row_map: dict[int, int] = {}
    dst_row = 2

    for src_row in src_rows:
        row_map[src_row] = dst_row
        for col_idx in range(1, max_col + 1):
            copy_cell(src_ws.cell(row=src_row, column=col_idx), dst_ws.cell(row=dst_row, column=col_idx))
        dst_row += 1

    return row_map


def rewrite_formula_for_row_map(formula: str, row_map: dict[int, int], sheet_name: str) -> str:
    """
    =月次KPIフォロー!A25 のような単純参照を、
    新しい行番号に置き換える
    """
    pattern = rf"(='{sheet_name}'!|={sheet_name}!)([A-Z]+)(\d+)$"
    m = re.match(pattern, formula)
    if not m:
        return formula

    prefix = m.group(1)
    col_part = m.group(2)
    old_row = int(m.group(3))

    if old_row not in row_map:
        return formula

    new_row = row_map[old_row]
    return f"{prefix}{col_part}{new_row}"


def build_ka_workbook(master_wb: Workbook, ka_name: str) -> Path | None:
    if SOURCE_SHEET_NAME not in master_wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SOURCE_SHEET_NAME}")

    src_ws = master_wb[SOURCE_SHEET_NAME]
    pushin_ws = master_wb[PUSHIN_SHEET_NAME] if PUSHIN_SHEET_NAME in master_wb.sheetnames else None

    target_rows = extract_rows_by_ka(src_ws, ka_name)
    if not target_rows:
        print(f"[SKIP] {ka_name}: 対象行なし")
        return None

    new_wb = Workbook()
    default_ws = new_wb.active
    new_wb.remove(default_ws)

    # -----------------------------------------------------------------------
    # 1) 月次KPIフォロー（課別）
    # -----------------------------------------------------------------------
    new_ws = new_wb.create_sheet(SOURCE_SHEET_NAME)

    copy_header_row(src_ws, new_ws, src_ws.max_column)
    row_map = copy_selected_rows(src_ws, new_ws, target_rows, src_ws.max_column)

    copy_column_widths(src_ws, new_ws, src_ws.max_column)
    copy_row_heights(src_ws, new_ws, row_map={1: 1, **row_map})
    copy_sheet_views_and_print_settings(src_ws, new_ws)
    copy_merged_cells_subset(src_ws, new_ws, row_map, src_ws.max_column)

    new_ws.auto_filter.ref = f"A1:{get_column_letter(src_ws.max_column)}{1 + len(target_rows)}"

    # -----------------------------------------------------------------------
    # 2) 推進部_月次管理（課別）
    # -----------------------------------------------------------------------
    if pushin_ws is not None:
        new_pushin = new_wb.create_sheet(PUSHIN_SHEET_NAME)

        copy_header_row(pushin_ws, new_pushin, pushin_ws.max_column)

        dst_row = 2
        pushin_row_map: dict[int, int] = {}

        for src_row in target_rows:
            pushin_row_map[src_row] = dst_row

            for col_idx in range(1, pushin_ws.max_column + 1):
                src_cell = pushin_ws.cell(row=src_row, column=col_idx)
                dst_cell = new_pushin.cell(row=dst_row, column=col_idx)
                copy_cell(src_cell, dst_cell)

                if isinstance(src_cell.value, str) and "月次KPIフォロー!" in src_cell.value:
                    dst_cell.value = rewrite_formula_for_row_map(
                        src_cell.value,
                        row_map=row_map,
                        sheet_name=SOURCE_SHEET_NAME,
                    )

            dst_row += 1

        copy_column_widths(pushin_ws, new_pushin, pushin_ws.max_column)
        copy_row_heights(pushin_ws, new_pushin, row_map={1: 1, **pushin_row_map})
        copy_sheet_views_and_print_settings(pushin_ws, new_pushin)
        copy_merged_cells_subset(pushin_ws, new_pushin, pushin_row_map, pushin_ws.max_column)

        new_pushin.auto_filter.ref = f"A1:{get_column_letter(pushin_ws.max_column)}{1 + len(target_rows)}"

    # -----------------------------------------------------------------------
    # 3) 補助シートコピー
    # -----------------------------------------------------------------------
    for aux_name in AUX_SHEETS:
        if aux_name in master_wb.sheetnames:
            src_aux = master_wb[aux_name]
            dst_aux = new_wb.create_sheet(aux_name)

            for r in range(1, src_aux.max_row + 1):
                for c in range(1, src_aux.max_column + 1):
                    copy_cell(src_aux.cell(r, c), dst_aux.cell(r, c))

            copy_column_widths(src_aux, dst_aux, src_aux.max_column)
            copy_row_heights(src_aux, dst_aux)
            copy_sheet_views_and_print_settings(src_aux, dst_aux)

            try:
                for merged_range in src_aux.merged_cells.ranges:
                    dst_aux.merge_cells(str(merged_range))
            except Exception:
                pass

            dst_aux.sheet_state = src_aux.sheet_state

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"KPIフォローシート_{safe_filename(ka_name)}.xlsx"
    new_wb.save(out_path)
    print(f"[OK] {ka_name}: {out_path}")
    return out_path


def main() -> None:
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"元ファイルが見つかりません: {SOURCE_FILE}")

    master_wb = openpyxl.load_workbook(SOURCE_FILE)

    print(f"[INFO] 元ファイル: {SOURCE_FILE}")
    print(f"[INFO] 出力先: {OUTPUT_DIR}")

    for ka_name in KA_LIST:
        build_ka_workbook(master_wb, ka_name)


if __name__ == "__main__":
    main()