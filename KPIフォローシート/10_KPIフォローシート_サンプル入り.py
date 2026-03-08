import openpyxl
import os
import random
import re
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils_excel import header_map, norm

# ===== 入出力パス =====
# スクリプト位置: VSCode/KPIフォローシート/
# 2階層上 (HiproBiz わかもと製薬/) を基点に入力ファイルを解決する
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(os.path.dirname(_SCRIPT_DIR))   # HiproBiz わかもと製薬/

_DEFAULT_SRC = os.path.join(
    _BASE_DIR,
    "わかもと製薬_KPIフォローシート",
    "KPIフォローシート",
    "KPIフォローシート.xlsx",
)

# コマンドライン第1引数でパスを上書き可: python script.py "別ファイル.xlsx"
src_path = sys.argv[1] if len(sys.argv) > 1 else _DEFAULT_SRC

# 出力は入力と同フォルダ・同ベース名に "_サンプル回答入り" を付与
_stem    = os.path.splitext(os.path.basename(src_path))[0]
out_path = os.path.join(os.path.dirname(src_path), _stem + "_サンプル回答入り.xlsx")
# =====================

def to_ratio(v):
    """'25%' / '0.25' / 0.25 → 0.25 に統一"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    m = re.fullmatch(r"([0-9]+(?:\.[0-9]+)?)\s*%", s)
    if m:
        return float(m.group(1)) / 100.0
    try:
        return float(s)
    except:
        return None

def detect_focus(*texts):
    """主軸レバー(AE) / 次フェーズ方針(AR) / 未達時の疑い軸(AQ) から重点領域を推定"""
    s = "\n".join([str(t) for t in texts if t is not None])
    for k in ["定着", "継続処方", "競合対抗", "拡大型採用"]:
        if k in s:
            return k
    if "継続" in s:
        return "継続処方"
    if "競合" in s:
        return "競合対抗"
    if "拡大" in s:
        return "拡大型採用"
    if "定着" in s:
        return "定着"
    return ""


def main():
    wb = openpyxl.load_workbook(src_path)

    # Excelで開いたときの再計算（KPI値/達成率が空に見える対策）
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    if "月次KPIフォロー" not in wb.sheetnames:
        raise ValueError("月次KPIフォロー シートが見つかりません")

    ws = wb["月次KPIフォロー"]
    col = header_map(ws)

    def C(name: str) -> int:
        if name not in col:
            raise ValueError(f"月次KPIフォローにヘッダ '{name}' が見つかりません。列名が変わっている可能性があります。")
        return col[name]

    # 対象列名（この名前が月次KPIフォローのヘッダに存在する前提）
    H_MONTH = "月(YYYY-MM)"
    H_DEPT = "課"
    H_LEVER = "主軸レバー(AE)"
    H_POLICY = "次フェーズ方針(AR)"
    H_DOUBT = "未達時の疑い軸(AQ)"
    H_KPI_NAME = "KPI候補(AU)"

    H_DEN = "分母(実績)"
    H_NUM = "分子(実績)"
    H_KPI = "KPI値(自動)"
    H_TGT = "目標(入力/比率)"
    H_ACH = "達成率(自動)"

    H_ACTION = "今月強化した具体行動"
    H_VOLUME = "実施量"
    H_REASON = "KPI増減理由(分子がなぜ動いたか)"
    H_NEXT = "来月の修正(何を変えるか)"

    # ★推進部コメント列（無ければ末尾に追加）
    H_COMMENT = "推進部コメント"
    if H_COMMENT not in col:
        new_col = ws.max_column + 1
        ws.cell(1, new_col).value = H_COMMENT
        col[H_COMMENT] = new_col

    # -------------------------
    # 1) 月次KPIフォロー：サンプル投入
    # -------------------------
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, C(H_MONTH)).value in (None, ""):
            continue

        lever = ws.cell(r, C(H_LEVER)).value
        policy = ws.cell(r, C(H_POLICY)).value
        doubt = ws.cell(r, C(H_DOUBT)).value
        kpi_name = ws.cell(r, C(H_KPI_NAME)).value or ""

        focus = detect_focus(lever, policy, doubt)

        # 目標（空なら0.30）
        tgt = to_ratio(ws.cell(r, C(H_TGT)).value)
        if tgt is None:
            tgt = 0.30
        ws.cell(r, C(H_TGT)).value = tgt
        ws.cell(r, C(H_TGT)).number_format = "0.0%"

        # 分母・分子（サンプル生成）
        den = random.randint(20, 60)
        num = int(max(0, min(den, round(den * tgt * random.uniform(0.7, 1.1)))))
        ws.cell(r, C(H_DEN)).value = den
        ws.cell(r, C(H_NUM)).value = num

        # KPI値・達成率（数式）
        den_addr = ws.cell(r, C(H_DEN)).coordinate
        num_addr = ws.cell(r, C(H_NUM)).coordinate

        kpi_cell = ws.cell(r, C(H_KPI))
        kpi_cell.value = f'=IFERROR({num_addr}/{den_addr},"")'
        kpi_cell.number_format = "0.0%"

        tgt_addr = ws.cell(r, C(H_TGT)).coordinate
        ach_cell = ws.cell(r, C(H_ACH))
        ach_cell.value = f'=IFERROR({kpi_cell.coordinate}/{tgt_addr},"")'
        ach_cell.number_format = "0.0%"

        # 重点領域別：行動・理由・来月修正（サンプル文章）＋ 推進部コメント
        if focus == "定着":
            action = "初回採用医師への早期フォロー強化（2週間以内再訪）"
            reason = f"フォロー実施群で継続が安定し、『{kpi_name}』の分子増加に寄与。"
            nxt = "成功パターン（対象・タイミング・トーク）を型化し対象拡大。"
            comment = "（推進部）分母内の優先群が妥当か、フォロー条件（タイミング/頻度）も併記してください。"
        elif focus == "継続処方":
            action = "中断理由ヒアリング → 理由別の再提案（次アクション）を実施"
            reason = f"中断理由を把握できた医師で再処方が発生し、『{kpi_name}』が改善。"
            nxt = "理由別の標準トーク／資料を整理し、優先群へ再展開。"
            comment = "（推進部）中断理由の分類（3〜5類型）と、理由別の打ち手をセットで残してください。"
        elif focus == "競合対抗":
            action = "競合使用医師へ症例提示型の切替提案（比較軸を明確化）"
            reason = f"比較軸が刺さった医師で切替が進み、『{kpi_name}』が改善傾向。"
            nxt = "刺さった症例・比較軸を標準化し、同タイプ医師へ横展開。"
            comment = "（推進部）医師タイプ別に『刺さる比較軸』を明確化し、次月の当たり先優先順位も書いてください。"
        elif focus == "拡大型採用":
            action = "未採用ターゲットへの重点アプローチ（採用の入口づくり）"
            reason = f"新規ターゲット掘り起こしが奏功し、『{kpi_name}』の分子が増加。"
            nxt = "採用に至ったターゲット像を定義し、見込み高群へ集中展開。"
            comment = "（推進部）採用に至ったターゲット像（条件/院内事情）を言語化し、横展開条件を整理してください。"
        else:
            action = "重点ターゲットへ戦略訪問（分母内優先で介入）"
            reason = f"行動が分子へ一定寄与し、『{kpi_name}』が改善傾向。"
            nxt = "対象優先順位を再整理し、介入内容（比較軸/フォロー）を組み替える。"
            comment = "（推進部）分母・分子の定義がブレていないか、未達時の疑い軸と接続して確認してください。"

        ws.cell(r, C(H_ACTION)).value = action
        ws.cell(r, C(H_VOLUME)).value = f"訪問{random.randint(10, 25)}件／重点{random.randint(5, 15)}名"
        ws.cell(r, C(H_REASON)).value = reason
        ws.cell(r, C(H_NEXT)).value = nxt

        # 月次側 推進部コメント：空なら必ず入れる
        ccell = ws.cell(r, C(H_COMMENT))
        if ccell.value in (None, ""):
            ccell.value = comment

    # -------------------------
    # 2) 推進部_月次管理：J列（推進部コメント）を必ず埋める（他列は触らない）
    # -------------------------
    if "推進部_月次管理" in wb.sheetnames:
        ws_s = wb["推進部_月次管理"]

        # J1ヘッダを強制
        ws_s.cell(1, 10).value = "推進部コメント"

        # 推進部側ヘッダマップ
        col_s = header_map(ws_s)

        def pick(cands, default):
            for n in cands:
                if n in col_s:
                    return col_s[n]
            return default

        # 推進部側の月/課/KPI列：拾えなければA/B/Dにフォールバック
        s_month = pick(["月(YYYY-MM)", "月", "年月"], default=1)
        s_dept  = pick(["課", "部署", "担当課"], default=2)
        s_kpi   = pick(["KPI候補(AU)", "KPI候補", "KPI"], default=4)

        # 月次側：キーで索引化
        idx_full = {}  # (月,課,KPI候補) -> 行
        idx_md = {}    # (月,課) -> 先頭行
        for r in range(2, ws.max_row + 1):
            m = ws.cell(r, C(H_MONTH)).value
            d = ws.cell(r, C(H_DEPT)).value
            k = ws.cell(r, C(H_KPI_NAME)).value
            if m in (None, "") or d in (None, ""):
                continue
            m2 = norm(m)
            d2 = norm(d)
            k2 = norm(k)
            idx_full[(m2, d2, k2)] = r
            idx_md.setdefault((m2, d2), r)

        comment_col_f = col[H_COMMENT]

        # 推進部側の各行：J列に値コピー（必ず埋める）
        for r in range(2, ws_s.max_row + 1):
            m = ws_s.cell(r, s_month).value
            if m in (None, ""):
                continue
            d = ws_s.cell(r, s_dept).value
            k = ws_s.cell(r, s_kpi).value

            m2 = norm(m)
            d2 = norm(d)
            k2 = norm(k)

            src_row = idx_full.get((m2, d2, k2))
            if src_row is None:
                src_row = idx_md.get((m2, d2))

            # 最終保険：同じ行番号（空回避）
            if src_row is None:
                src_row = min(r, ws.max_row)

            ws_s.cell(r, 10).value = ws.cell(src_row, comment_col_f).value

    # 保存
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print("✅ 完了：サンプル回答投入＋推進部_月次管理(J列コメント)埋めまで保存しました:", out_path)

if __name__ == "__main__":
    main()