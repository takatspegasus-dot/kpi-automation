import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils_excel import header_map, norm

# ===== 入出力パス（ここだけ変更）=====
src_path = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\わかもと製薬_KPIフォローシート\KPIフォローシート\KPIフォローシート.xlsx"
out_path = r"C:\Users\SujiT\OneDrive\ドキュメント\HiproBiz わかもと製薬\わかもと製薬_KPIフォローシート\KPIフォローシート\KPIフォローシート_現場入力テンプレ全課_計算式入り.xlsx"
# ====================================

SRC_SHEET = "月次KPIフォロー"
NEW_SHEET = "月次KPIフォロー_現場入力"   # 全課縦並びテンプレ
LIST_SHEET = "_tpl_lists"               # DV用（無ければ作成して非表示）

LEVER_LIST = ["定着", "継続処方", "競合対抗", "拡大型採用"]
DOUBT_LIST = ["分母定義", "対象選定", "行動量", "行動質", "競合", "その他"]

# 入力枠：3〜203行（200行）
FORCE_START = 3
FORCE_END = 203
MAX_ROWS = FORCE_END - FORCE_START + 1  # 200

# テンプレ列（A〜P）
TPL_COLS = [
    "月(YYYY-MM)", "課", "主軸レバー", "KPI名", "分母定義",
    "分母(実績)", "分子(実績)", "KPI値(自動)", "目標(入力/比率)", "達成率(自動)",
    "①今月動かした対象（分母内のどこ？）",
    "②実施した具体行動（何をどう変えた？）",
    "③分子はどう動いた？（事実のみ）",
    "④なぜ動いた？（仮説）",
    "未達時の疑い軸（選択）",
    "来月の修正（具体）",
]


def ensure_list_sheet(wb, dept_list):
    if LIST_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(LIST_SHEET)
    else:
        ws = wb[LIST_SHEET]
    ws.sheet_state = "hidden"

    ws.delete_rows(1, ws.max_row)

    ws["A1"].value = "課"
    if not dept_list:
        dept_list = ["（課リストをここに入力）"]
    for i, d in enumerate(dept_list, start=2):
        ws.cell(i, 1).value = d

    ws["B1"].value = "主軸レバー"
    for i, x in enumerate(LEVER_LIST, start=2):
        ws.cell(i, 2).value = x

    ws["C1"].value = "未達時の疑い軸"
    for i, x in enumerate(DOUBT_LIST, start=2):
        ws.cell(i, 3).value = x

    dept_range  = f"'{LIST_SHEET}'!$A$2:$A${1+len(dept_list)}"
    lever_range = f"'{LIST_SHEET}'!$B$2:$B${1+len(LEVER_LIST)}"
    doubt_range = f"'{LIST_SHEET}'!$C$2:$C${1+len(DOUBT_LIST)}"
    return dept_range, lever_range, doubt_range

def build_template(wb):
    if SRC_SHEET not in wb.sheetnames:
        raise ValueError(f"元シート '{SRC_SHEET}' が見つかりません")

    ws_src = wb[SRC_SHEET]
    src_cols = header_map(ws_src)

    # 元シートから拾う列（目標は「取らない」設計に変更）
    c_month = src_cols.get("月(YYYY-MM)")
    c_dept  = src_cols.get("課")
    c_lever = src_cols.get("主軸レバー(AE)") or src_cols.get("主軸レバー")
    c_kpi   = src_cols.get("KPI候補(AU)") or src_cols.get("KPI候補") or src_cols.get("KPI名")
    c_den_def = src_cols.get("分母定義(AW)") or src_cols.get("分母定義")

    if c_dept is None or c_kpi is None:
        raise ValueError("元シートに '課' または 'KPI候補(AU)' が見つかりません（列名が変わっている可能性）")

    # 元シートから（課, KPI）行を抽出
    base_rows = []
    for r in range(2, ws_src.max_row + 1):
        dept = ws_src.cell(r, c_dept).value
        kpi  = ws_src.cell(r, c_kpi).value
        if dept in (None, "") or kpi in (None, ""):
            continue

        month = ws_src.cell(r, c_month).value if c_month else ""
        lever = ws_src.cell(r, c_lever).value if c_lever else ""
        den_def = ws_src.cell(r, c_den_def).value if c_den_def else ""

        base_rows.append({
            "month": month if month is not None else "",
            "dept": norm(dept),
            "lever": norm(lever),
            "kpi": norm(kpi),
            "den_def": norm(den_def),
            # ★目標は毎月入力するので、テンプレ作成時は空欄にする
            "tgt": "",
        })

    # 課リスト（DV用）
    dept_list = []
    seen = set()
    for x in base_rows:
        d = x["dept"]
        if d and d not in seen:
            seen.add(d)
            dept_list.append(d)

    dept_range, lever_range, doubt_range = ensure_list_sheet(wb, dept_list)

    # テンプレシート作成（既にあれば作り直し）
    if NEW_SHEET in wb.sheetnames:
        del wb[NEW_SHEET]
    ws = wb.create_sheet(NEW_SHEET, 0)

    # 書式
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    sample_fill = PatternFill("solid", fgColor="F2F2F2")
    input_fill  = PatternFill("solid", fgColor="FFF2CC")
    auto_fill   = PatternFill("solid", fgColor="E7E6E6")

    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1行目：ヘッダ
    for c, name in enumerate(TPL_COLS, start=1):
        cell = ws.cell(1, c, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # テンプレ側ヘッダ→列番号
    tpl_cols = header_map(ws)
    def T(name): return tpl_cols[name]

    # 2行目：記入例（目標は例として入れるが、実運用は毎月入力）
    examples = {
        "月(YYYY-MM)": "2026-04",
        "課": "東京1課",
        "主軸レバー": "定着",
        "KPI名": "例）重点対象の継続処方率",
        "分母定義": "例）初回採用医師（当月フォロー対象）",
        "分母(実績)": 40,
        "分子(実績)": 12,
        "目標(入力/比率)": 0.30,
        "①今月動かした対象（分母内のどこ？）": "例）分母40名のうち未フォロー12名",
        "②実施した具体行動（何をどう変えた？）": "例）2週間以内再訪を徹底（対象12名に実施）",
        "③分子はどう動いた？（事実のみ）": "例）再処方4件発生",
        "④なぜ動いた？（仮説）": "例）切替前の早期介入が奏功",
        "未達時の疑い軸（選択）": "行動量",
        "来月の修正（具体）": "例）対象を優先群に絞り、再訪頻度を増やす",
    }
    for c in range(1, len(TPL_COLS) + 1):
        h = ws.cell(1, c).value
        v = examples.get(h, None)
        cell = ws.cell(2, c, v)
        cell.fill = sample_fill
        cell.alignment = wrap_top
        cell.border = border

    # 記入例行：計算式
    ws.cell(2, T("KPI値(自動)")).value = f'=IFERROR({get_column_letter(T("分子(実績)"))}2/{get_column_letter(T("分母(実績)"))}2,"")'
    ws.cell(2, T("達成率(自動)")).value = f'=IFERROR({get_column_letter(T("KPI値(自動)"))}2/{get_column_letter(T("目標(入力/比率)"))}2,"")'
    ws.cell(2, T("目標(入力/比率)")).number_format = "0.0%"
    ws.cell(2, T("KPI値(自動)")).number_format = "0.0%"
    ws.cell(2, T("達成率(自動)")).number_format = "0.0%"

    # 3〜203行：固定で作る（200行）
    for i in range(MAX_ROWS):
        r = FORCE_START + i

        # 枠の書式
        for c in range(1, len(TPL_COLS) + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = wrap_top

            header = ws.cell(1, c).value
            if header in ("KPI値(自動)", "達成率(自動)"):
                cell.fill = auto_fill
            else:
                if header in (
                    "月(YYYY-MM)", "課", "主軸レバー", "KPI名", "分母定義",
                    "分母(実績)", "分子(実績)", "目標(入力/比率)",
                    "①今月動かした対象（分母内のどこ？）",
                    "②実施した具体行動（何をどう変えた？）",
                    "③分子はどう動いた？（事実のみ）",
                    "④なぜ動いた？（仮説）",
                    "未達時の疑い軸（選択）",
                    "来月の修正（具体）",
                ):
                    cell.fill = input_fill

        # 値流し込み（ベースがある分だけ）
        if i < len(base_rows):
            x = base_rows[i]
            ws.cell(r, T("月(YYYY-MM)")).value = x["month"]
            ws.cell(r, T("課")).value = x["dept"]
            ws.cell(r, T("主軸レバー")).value = x["lever"]
            ws.cell(r, T("KPI名")).value = x["kpi"]
            ws.cell(r, T("分母定義")).value = x["den_def"]
            # ★目標は毎月入力：ここでは入れない（空欄のまま）

        # ★H〜Jを必ず設定（3〜203）
        ws.cell(r, T("KPI値(自動)")).value = (
            f'=IFERROR({get_column_letter(T("分子(実績)"))}{r}/'
            f'{get_column_letter(T("分母(実績)"))}{r},"")'
        )
        ws.cell(r, T("KPI値(自動)")).number_format = "0.0%"

        ws.cell(r, T("目標(入力/比率)")).number_format = "0.0%"

        ws.cell(r, T("達成率(自動)")).value = (
            f'=IFERROR({get_column_letter(T("KPI値(自動)"))}{r}/'
            f'{get_column_letter(T("目標(入力/比率)"))}{r},"")'
        )
        ws.cell(r, T("達成率(自動)")).number_format = "0.0%"

    # 列幅
    widths = {
        T("月(YYYY-MM)"): 11,
        T("課"): 12,
        T("主軸レバー"): 12,
        T("KPI名"): 28,
        T("分母定義"): 24,
        T("分母(実績)"): 10,
        T("分子(実績)"): 10,
        T("KPI値(自動)"): 10,
        T("目標(入力/比率)"): 12,
        T("達成率(自動)"): 10,
        T("①今月動かした対象（分母内のどこ？）"): 34,
        T("②実施した具体行動（何をどう変えた？）"): 36,
        T("③分子はどう動いた？（事実のみ）"): 28,
        T("④なぜ動いた？（仮説）"): 28,
        T("未達時の疑い軸（選択）"): 18,
        T("来月の修正（具体）"): 36,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # 行高
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42
    for r in range(FORCE_START, FORCE_END + 1):
        ws.row_dimensions[r].height = 60

    ws.freeze_panes = "A3"

    # DV（プルダウン）
    dv_dept  = DataValidation(type="list", formula1=f"={dept_range}", allow_blank=True)
    dv_lever = DataValidation(type="list", formula1=f"={lever_range}", allow_blank=True)
    dv_doubt = DataValidation(type="list", formula1=f"={doubt_range}", allow_blank=True)
    ws.add_data_validation(dv_dept)
    ws.add_data_validation(dv_lever)
    ws.add_data_validation(dv_doubt)

    dv_dept.add(f"{get_column_letter(T('課'))}{FORCE_START}:{get_column_letter(T('課'))}{FORCE_END}")
    dv_lever.add(f"{get_column_letter(T('主軸レバー'))}{FORCE_START}:{get_column_letter(T('主軸レバー'))}{FORCE_END}")
    dv_doubt.add(f"{get_column_letter(T('未達時の疑い軸（選択）'))}{FORCE_START}:{get_column_letter(T('未達時の疑い軸（選択）'))}{FORCE_END}")

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

def main():
    wb = openpyxl.load_workbook(src_path)
    build_template(wb)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print("✅ 完了：目標は月次入力（元から取得しない）で、全課テンプレ＋H〜J計算式/表示形式まで反映して保存しました:", out_path)

if __name__ == "__main__":
    main()